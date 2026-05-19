"""Smoke test the synthetic public demo dataset.

This script intentionally uses only the safe openpyxl-based export path. It does
not require Microsoft Excel COM and does not write generated reports to disk.
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
import sys

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

from tmc_processor.importer import load_detected_sheets
from tmc_processor.mapping import apply_saved_mapping_to_sheets, read_mapping_excel
from tmc_processor.pipeline import process_tmc


DEMO_DIR = ROOT / "samples" / "demo"
RAW_WORKBOOK = DEMO_DIR / "DEMO_TMC1_FourLeg.xlsx"
MAPPING_WORKBOOK = DEMO_DIR / "DEMO_TMC1_FourLeg_mapping.xlsx"


def main() -> int:
    raw_sheets = load_detected_sheets(RAW_WORKBOOK)
    if sorted(raw_sheets) != ["ทิศ 1", "ทิศ 2", "ทิศ 3", "ทิศ 4"]:
        raise AssertionError(f"Unexpected demo raw sheets: {sorted(raw_sheets)}")

    for sheet_name, rows in raw_sheets.items():
        if rows.empty:
            raise AssertionError(f"{sheet_name} parsed with no rows")
        vehicle_classes = set(rows["vehicle_class"].dropna())
        expected = {"Bicy", "MC", "PC<7", "PC>7", "LB", "MB", "HB", "LT", "MT", "HT", "TR", "STR"}
        if vehicle_classes != expected:
            raise AssertionError(f"{sheet_name} vehicle classes mismatch: {sorted(vehicle_classes)}")

    saved_mapping = read_mapping_excel(MAPPING_WORKBOOK)
    mapping = apply_saved_mapping_to_sheets(list(raw_sheets), saved_mapping)
    setup = {
        "project_name": "Synthetic Demo Project",
        "tmc_id": "DEMO-TMC1",
        "tmc_name": "Synthetic Four-Leg Demo Intersection",
        "survey_point": "Synthetic Four-Leg Demo Point",
        "survey_date": "2026-01-01",
        "survey_period": "07:00-19:00",
        "north_label": "Demo North",
        "south_label": "Demo South",
        "east_label": "Demo East",
        "west_label": "Demo West",
    }

    result = process_tmc(
        raw_sheets=raw_sheets,
        mapping=mapping,
        setup=setup,
        detected_sheets=list(raw_sheets),
        generate_workbook=True,
        use_template_report_layout=False,
        use_excel_com_native_charts=False,
    )

    if result.normalized.empty:
        raise AssertionError("Processing produced no normalized rows")
    if result.hourly.empty or result.movement.empty or result.vehicle.empty:
        raise AssertionError("Processing summaries were not generated")

    peak_periods = set(result.peaks["period"].astype(str))
    if {"AM", "PM"} - peak_periods:
        raise AssertionError(f"Expected AM and PM peaks, got {sorted(peak_periods)}")

    peak_lookup = result.peaks.set_index("period")
    am_start = str(peak_lookup.loc["AM", "peak_start"])[:5]
    pm_start = str(peak_lookup.loc["PM", "peak_start"])[:5]
    if am_start != "08:00" or pm_start != "17:00":
        raise AssertionError(f"Unexpected demo peak starts: AM={am_start}, PM={pm_start}")

    if not result.workbook_bytes:
        raise AssertionError("Export workbook bytes were not generated")
    workbook = load_workbook(BytesIO(result.workbook_bytes), read_only=True)
    required_sheets = {"Export_Metadata", "Setup", "Mapping", "Normalized_Data", "Peak_PHF", "TMC_Report"}
    missing = required_sheets - set(workbook.sheetnames)
    if missing:
        raise AssertionError(f"Export workbook missing sheets: {sorted(missing)}")

    print("Demo smoke test passed")
    print(f"Raw sheets: {len(raw_sheets)}")
    print(f"Normalized rows: {len(result.normalized)}")
    print(f"AM peak: {am_start}, PM peak: {pm_start}")
    print(f"Export workbook bytes: {len(result.workbook_bytes):,}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
