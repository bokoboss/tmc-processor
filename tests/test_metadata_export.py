from __future__ import annotations

from datetime import time
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook
import pandas as pd

from tmc_processor.batch import BatchItem, process_batch_files
from tmc_processor.exporter import export_workbook
from tmc_processor.mapping_preset import load_mapping_preset
from tmc_processor.pipeline import process_tmc
from tmc_processor.session import apply_session_to_state, build_project_session, session_from_json, session_to_json_bytes


ROOT = Path(__file__).resolve().parents[1]
DEMO_DIR = ROOT / "samples" / "demo"
DAY1 = DEMO_DIR / "DEMO_TMC1_FourLeg.xlsx"
PRESET = DEMO_DIR / "DEMO_TMC1_FourLeg.mapping.json"


def _mapping() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "raw_sheet": "S1",
                "raw_direction": "North",
                "movement_code": "NS",
                "source_stream": "mainline",
                "raw_movement_label": "North",
                "from_leg": "N",
                "to_leg": "S",
                "turn_type": "through",
                "facility_type": "at_grade",
                "include_in_peak": True,
                "include_in_report": True,
                "aggregation_method": "sum",
            }
        ]
    )


def _raw_hour() -> dict[str, pd.DataFrame]:
    rows = []
    for minute in (0, 15, 30, 45):
        start = time(7, minute)
        end = time(8, 0) if minute == 45 else time(7, minute + 15)
        rows.append(
            {
                "raw_sheet": "S1",
                "raw_direction": "North",
                "time_start": start,
                "time_end": end,
                "vehicle_class": "PC<7",
                "count": 10,
            }
        )
    return {"S1": pd.DataFrame(rows)}


def _setup() -> dict[str, object]:
    return {
        "project_name": "Metadata Project",
        "tmc_id": "META-01",
        "tmc_title": "รายงานทดสอบ Metadata Export",
        "survey_point": "จุดสำรวจทดสอบ",
        "survey_date_text": "วันทดสอบ 1 มกราคม 2569",
        "weather": "แจ่มใสทดสอบ",
        "responsible_party": "ผู้รับผิดชอบทดสอบ",
        "survey_period": "07:00-19:00",
        "north_label": "ปลายทางเหนือทดสอบ",
        "south_label": "ปลายทางใต้ทดสอบ",
        "east_label": "ปลายทางตะวันออกทดสอบ",
        "west_label": "ปลายทางตะวันตกทดสอบ",
        "north_road": "ถนนเหนือทดสอบ",
        "south_road": "ถนนใต้ทดสอบ",
        "east_road": "ถนนตะวันออกทดสอบ",
        "west_road": "ถนนตะวันตกทดสอบ",
        "caption_text": "คำบรรยายรายงานทดสอบ",
        "show_u_turn": False,
    }


def _processed():
    return process_tmc(
        raw_sheets=_raw_hour(),
        mapping=_mapping(),
        setup=_setup(),
        detected_sheets=["S1"],
        peak_windows={"AM": ("07:00", "08:00"), "PM": ("07:00", "08:00")},
        generate_workbook=False,
    )


def _sheet_records(workbook, sheet_name: str) -> dict[object, object]:
    return {
        row[0]: row[1]
        for row in workbook[sheet_name].iter_rows(min_row=2, values_only=True)
        if row[0] is not None
    }


def test_safe_png_workbook_exports_setup_metadata_to_metadata_and_setup_sheets() -> None:
    result = _processed()
    workbook_bytes = export_workbook(
        _setup(),
        _mapping(),
        result.normalized,
        result.qc,
        result.hourly,
        result.movement,
        result.vehicle,
        result.peaks,
        include_charts=False,
        include_diagram=False,
        export_mode="Safe PNG Export Mode",
        generated_at="2026-05-19T10:00:00Z",
    )

    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
    setup_rows = _sheet_records(workbook, "Setup")
    export_metadata = _sheet_records(workbook, "Export_Metadata")

    assert export_metadata["report_title"] == "รายงานทดสอบ Metadata Export"
    assert export_metadata["survey_date"] == "วันทดสอบ 1 มกราคม 2569"
    assert export_metadata["responsible_party"] == "ผู้รับผิดชอบทดสอบ"
    assert export_metadata["caption_text"] == "คำบรรยายรายงานทดสอบ"
    assert export_metadata["show_u_turn"] is False
    assert setup_rows["north_road"] == "ถนนเหนือทดสอบ"
    assert setup_rows["south_road"] == "ถนนใต้ทดสอบ"
    assert setup_rows["north_label"] == "ปลายทางเหนือทดสอบ"
    assert setup_rows["south_label"] == "ปลายทางใต้ทดสอบ"
    assert "tmc_title" in str(workbook["TMC_Report"]["A1"].value)


def test_template_layout_writes_mapped_metadata_and_direction_cells() -> None:
    result = _processed()
    workbook_bytes = export_workbook(
        _setup(),
        _mapping(),
        result.normalized,
        result.qc,
        result.hourly,
        result.movement,
        result.vehicle,
        result.peaks,
        include_charts=False,
        include_diagram=False,
        use_template_report_layout=True,
        export_mode="Excel Template Mode",
    )

    worksheet = load_workbook(BytesIO(workbook_bytes), data_only=False)["TMC_Report"]

    assert worksheet["B2"].value == "รายงานทดสอบ Metadata Export"
    assert worksheet["E5"].value == "จุดสำรวจทดสอบ"
    assert worksheet["K5"].value == "วันทดสอบ 1 มกราคม 2569"
    assert worksheet["E6"].value == "ผู้รับผิดชอบทดสอบ"
    assert worksheet["G12"].value == "ปลายทางเหนือทดสอบ"
    assert worksheet["O32"].value == "ปลายทางใต้ทดสอบ"
    assert worksheet["K11"].value == "ถนนเหนือทดสอบ"
    assert worksheet["K32"].value == "ถนนใต้ทดสอบ"
    assert worksheet["E35"].value == "คำบรรยายรายงานทดสอบ"


def test_project_session_preserves_setup_metadata_direction_labels_and_u_turn_toggle() -> None:
    session = build_project_session(metadata=_setup(), directions=_setup())
    loaded = session_from_json(session_to_json_bytes(session)).session
    state: dict[str, object] = {}

    changed = apply_session_to_state(loaded, state)

    assert loaded["metadata"]["tmc_title"] == "รายงานทดสอบ Metadata Export"
    assert loaded["directions"]["north_road"] == "ถนนเหนือทดสอบ"
    assert loaded["directions"]["caption_text"] == "คำบรรยายรายงานทดสอบ"
    assert loaded["directions"]["show_u_turn"] is False
    assert state["tmc_title_input"] == "รายงานทดสอบ Metadata Export"
    assert state["north_label_input"] == "ปลายทางเหนือทดสอบ"
    assert state["show_u_turn_checkbox"] is False
    assert "show_u_turn_checkbox" in changed


def test_batch_report_workbook_uses_per_file_date_and_shared_setup_metadata() -> None:
    preset = load_mapping_preset(PRESET.read_bytes()).preset
    result = process_batch_files(
        [
            BatchItem(
                file_name=DAY1.name,
                workbook_bytes=DAY1.read_bytes(),
                survey_date_text="วันสำรวจรายไฟล์ทดสอบ",
                output_stem="metadata-batch",
                notes="batch notes",
            )
        ],
        mapping_preset=preset,
        setup=_setup(),
        use_template_report_layout=False,
        generated_at="2026-05-19T10:00:00Z",
    )

    with ZipFile(BytesIO(result.package_bytes)) as archive:
        workbook_bytes = archive.read("file_01_metadata-batch/metadata-batch_report.xlsx")
        summary_text = archive.read("file_01_metadata-batch/metadata-batch_export_summary.txt").decode("utf-8")

    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
    setup_rows = _sheet_records(workbook, "Setup")
    export_metadata = _sheet_records(workbook, "Export_Metadata")

    assert setup_rows["survey_date_text"] == "วันสำรวจรายไฟล์ทดสอบ"
    assert export_metadata["survey_date"] == "วันสำรวจรายไฟล์ทดสอบ"
    assert setup_rows["north_road"] == "ถนนเหนือทดสอบ"
    assert setup_rows["north_label"] == "ปลายทางเหนือทดสอบ"
    assert "output_stem: metadata-batch" in summary_text
    assert "survey_date_text: วันสำรวจรายไฟล์ทดสอบ" in summary_text
