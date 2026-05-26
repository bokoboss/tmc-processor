from __future__ import annotations

from pathlib import Path
from io import BytesIO
from unittest.mock import patch
from zipfile import ZipFile

from openpyxl import load_workbook
import pandas as pd
import pytest

import tmc_processor.exporter as exporter_module
from tmc_processor.diagram import build_v2_movement_diagram_data, render_v2_movement_diagram_png
from tmc_processor.export_package import create_v2_generated_export_package_zip
from tmc_processor.exporter import export_v2_generated_workbook, export_v2_template_workbook, export_v2_template_workbook_com, export_workbook
from tmc_processor.importer import load_detected_sheets
from tmc_processor.mapping import apply_saved_mapping_to_sheets, read_mapping_excel_with_metadata, validate_mapping_scheme
from tmc_processor.mapping_preset import apply_mapping_preset_to_detected_sheets, load_mapping_preset
from tmc_processor.movement_scheme import APPROACH_MOVEMENT_CODES, MOVEMENT_SCHEME_V2, is_approach_movement_code
from tmc_processor.pipeline import process_tmc, process_tmc_dry_run_v2
from tmc_processor.batch import batch_processing_block_reason


ROOT = Path(__file__).resolve().parents[1]
DEMO_DIR = ROOT / "samples" / "demo"
RAW_WORKBOOK = DEMO_DIR / "DEMO_TMC1_FourLeg.xlsx"
V2_PRESET = DEMO_DIR / "DEMO_TMC1_FourLeg_approach_v2.mapping.json"
V2_MAPPING_XLSX = DEMO_DIR / "DEMO_TMC1_FourLeg_approach_v2_mapping.xlsx"
V2_TEMPLATE_WORKBOOK = ROOT / "templates" / "four_leg_tmc_report_template_approach_v2.xlsx"
V2_TEMPLATE_MAP = ROOT / "templates" / "four_leg_tmc_report_template_approach_v2_map.json"


def _setup() -> dict[str, object]:
    return {
        "project_name": "Synthetic Demo Project",
        "tmc_id": "DEMO-TMC1",
        "tmc_name": "Synthetic Four-Leg Demo Intersection",
        "survey_date": "2026-01-01",
        "movement_code_scheme": MOVEMENT_SCHEME_V2,
    }


def _custom_setup() -> dict[str, object]:
    return {
        **_setup(),
        "report_title": "Custom v2 TMC Report",
        "survey_point": "Approach V2 Test Point",
        "survey_date_text": "2026-05-26",
        "responsible_party": "Phase I2C QA",
        "north_label": "North Destination",
        "south_label": "South Destination",
        "east_label": "East Destination",
        "west_label": "West Destination",
        "north_road": "North Road",
        "south_road": "South Road",
        "east_road": "East Road",
        "west_road": "West Road",
        "caption_text": "Custom approach movement caption",
    }


def _raw_sheets() -> dict[str, pd.DataFrame]:
    return load_detected_sheets(RAW_WORKBOOK)


def _v2_preset_mapping(raw_sheets: dict[str, pd.DataFrame]) -> pd.DataFrame:
    preset = load_mapping_preset(V2_PRESET.read_bytes()).preset
    return apply_mapping_preset_to_detected_sheets(preset, list(raw_sheets)).mapping


def _v2_excel_mapping(raw_sheets: dict[str, pd.DataFrame]) -> pd.DataFrame:
    loaded = read_mapping_excel_with_metadata(V2_MAPPING_XLSX)
    assert loaded.movement_code_scheme == MOVEMENT_SCHEME_V2
    return apply_saved_mapping_to_sheets(list(raw_sheets), loaded.mapping)


def _dry_run_with_preset():
    raw_sheets = _raw_sheets()
    mapping = _v2_preset_mapping(raw_sheets)
    return process_tmc_dry_run_v2(
        raw_sheets=raw_sheets,
        mapping=mapping,
        setup=_setup(),
        detected_sheets=list(raw_sheets),
    )


def test_v2_demo_mapping_preset_loads_and_validates_for_dry_run() -> None:
    raw_sheets = _raw_sheets()
    mapping = _v2_preset_mapping(raw_sheets)

    assert validate_mapping_scheme(mapping, MOVEMENT_SCHEME_V2) == []
    assert mapping["movement_code"].tolist() == ["NT", "NT", "WT", "ET"]
    assert mapping["approach_direction"].tolist() == ["N", "N", "W", "E"]
    assert mapping["movement_type"].tolist() == ["T", "T", "T", "T"]


def test_v2_mapping_excel_loads_and_dry_runs() -> None:
    raw_sheets = _raw_sheets()
    mapping = _v2_excel_mapping(raw_sheets)
    result = process_tmc_dry_run_v2(
        raw_sheets=raw_sheets,
        mapping=mapping,
        setup=_setup(),
        detected_sheets=list(raw_sheets),
    )

    assert not result.normalized.empty
    assert set(result.normalized["movement_code"]) == {"NT", "WT", "ET"}
    assert result.workbook_bytes == b""


def test_v2_dry_run_normalizes_approach_movement_codes_and_metadata() -> None:
    result = _dry_run_with_preset()

    assert not result.normalized.empty
    assert set(result.normalized["movement_code"]) == {"NT", "WT", "ET"}
    assert set(result.normalized["output_movement_code"]) == {"NT", "WT", "ET"}
    assert set(result.normalized["movement_code_scheme"]) == {MOVEMENT_SCHEME_V2}
    assert set(result.normalized["approach_direction"]) == {"N", "W", "E"}
    assert set(result.normalized["movement_type"]) == {"T"}
    assert all(is_approach_movement_code(code) for code in result.normalized["movement_code"].unique())


def test_v2_hourly_movement_summary_uses_approach_movement_order() -> None:
    result = _dry_run_with_preset()

    assert list(result.hourly_movement_pcu.columns[1:-1]) == APPROACH_MOVEMENT_CODES
    assert result.hourly_movement_pcu.columns[-1] == "Total"
    assert set(result.hourly_movement_pcu.columns[1:-1]) == set(APPROACH_MOVEMENT_CODES)


def test_v2_movement_codes_in_outputs_are_valid_approach_movement() -> None:
    result = _dry_run_with_preset()

    normalized_codes = set(result.normalized["movement_code"].dropna().astype(str))
    movement_codes = set(result.movement["movement_code"].dropna().astype(str))
    summary_codes = set(result.hourly_movement_pcu.columns[1:-1])

    assert normalized_codes
    assert normalized_codes <= set(APPROACH_MOVEMENT_CODES)
    assert movement_codes <= set(APPROACH_MOVEMENT_CODES)
    assert all(is_approach_movement_code(code) for code in summary_codes)


def test_v2_peak_calculation_returns_am_pm_without_v1_movement_assumptions() -> None:
    result = _dry_run_with_preset()

    assert set(result.peaks["period"].astype(str)) == {"AM", "PM"}
    assert result.peaks["hourly_pcu"].gt(0).all()
    assert set(result.normalized["movement_code"]).isdisjoint({"NS", "WE", "EW"})


def test_v2_dry_run_does_not_enable_final_excel_export() -> None:
    result = _dry_run_with_preset()

    assert result.workbook_bytes == b""
    with pytest.raises(ValueError, match="v2 export/report generation is not supported"):
        export_workbook(
            _setup(),
            pd.DataFrame(),
            result.normalized,
            result.qc,
            result.hourly,
            result.movement,
            result.vehicle,
            result.peaks,
        )

    with pytest.raises(ValueError, match="v2 export/report generation is not supported"):
        export_workbook(
            _setup(),
            pd.DataFrame(),
            result.normalized,
            result.qc,
            result.hourly,
            result.movement,
            result.vehicle,
            result.peaks,
            use_template_report_layout=True,
            export_mode="Excel Template Mode",
        )

    with pytest.raises(ValueError, match="v2 export/report generation is not supported"):
        export_workbook(
            _setup(),
            pd.DataFrame(),
            result.normalized,
            result.qc,
            result.hourly,
            result.movement,
            result.vehicle,
            result.peaks,
            use_excel_com_native_charts=True,
            export_mode="Excel Template Mode",
        )

    with pytest.raises(ValueError, match="approach_movement v2"):
        process_tmc(
            raw_sheets=_raw_sheets(),
            mapping=_v2_preset_mapping(_raw_sheets()),
            setup=_setup(),
            detected_sheets=list(_raw_sheets()),
            generate_workbook=False,
        )


def test_mixed_v1_v2_mapping_codes_are_rejected() -> None:
    mixed = pd.DataFrame(
        [
            {"raw_sheet": "A", "movement_code": "NS", "include_in_report": True},
            {"raw_sheet": "B", "movement_code": "NT", "include_in_report": True},
        ]
    )

    assert validate_mapping_scheme(mixed, MOVEMENT_SCHEME_V2) == [
        "Mapping contains mixed from_to and approach_movement movement codes.",
        "Row 1 has invalid approach_movement output_movement_code 'NS'.",
    ]


def test_batch_v2_dry_run_is_supported_by_phase_k() -> None:
    assert batch_processing_block_reason(MOVEMENT_SCHEME_V2) == ""


def _sheet_records(workbook, sheet_name: str) -> dict[object, object]:
    return {
        row[0]: row[1]
        for row in workbook[sheet_name].iter_rows(min_row=2, values_only=True)
        if row[0] is not None
    }


def test_v2_dry_run_result_exports_generated_workbook_bytes() -> None:
    result = _dry_run_with_preset()

    workbook_bytes = export_v2_generated_workbook(
        result,
        setup=_setup(),
        mapping=_v2_preset_mapping(_raw_sheets()),
        generated_at="2026-05-19T10:00:00Z",
    )

    assert workbook_bytes.startswith(b"PK")
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
    assert "Export_Metadata" in workbook.sheetnames
    assert "Normalized_Data" in workbook.sheetnames
    assert "Hourly_Movement_PCU" in workbook.sheetnames
    assert "Movement_Summary" in workbook.sheetnames
    assert "Vehicle_Composition" in workbook.sheetnames
    assert "Peak_Summary" in workbook.sheetnames
    assert "QC_Check" in workbook.sheetnames

    metadata = _sheet_records(workbook, "Export_Metadata")
    assert metadata["movement_code_scheme"] == MOVEMENT_SCHEME_V2
    assert metadata["template_version"] == "generated_approach_movement_v2"
    assert metadata["export_template"] == "generated_approach_movement_v2"
    assert metadata["export_mode_used"] == "Safe PNG Export Mode"
    assert "openpyxl template helper is limited to structural/internal validation" in metadata["v2_export_limitation_notes"]


def test_v2_generated_hourly_movement_columns_follow_approach_order() -> None:
    result = _dry_run_with_preset()
    workbook = load_workbook(BytesIO(export_v2_generated_workbook(result, setup=_setup())), data_only=False)
    headers = [cell.value for cell in workbook["Hourly_Movement_PCU"][1]]

    assert headers[1:17] == APPROACH_MOVEMENT_CODES
    assert headers[-1] == "Total"


def test_v2_generated_workbook_contains_all_movement_code_references() -> None:
    result = _dry_run_with_preset()
    workbook = load_workbook(BytesIO(export_v2_generated_workbook(result, setup=_setup())), data_only=False)
    worksheet = workbook["Movement_Code_Reference"]
    rows = list(worksheet.iter_rows(min_row=2, values_only=True))

    assert [row[0] for row in rows] == APPROACH_MOVEMENT_CODES
    assert len(rows) == 16
    assert rows[0] == ("NL", "N", "Northbound", "L", "Left turn", "NL - Northbound Left turn")


def test_v2_diagram_data_helper_returns_ordered_approach_movements() -> None:
    result = _dry_run_with_preset()
    diagram = build_v2_movement_diagram_data(
        movement_summary=result.movement,
        hourly_movement_pcu=result.hourly_movement_pcu,
        peaks=result.peaks,
    )

    assert diagram["movement_code"].tolist() == APPROACH_MOVEMENT_CODES
    assert len(diagram) == 16


def test_v2_diagram_data_labels_nl_as_northbound_left() -> None:
    diagram = build_v2_movement_diagram_data()
    nl = diagram.loc[diagram["movement_code"] == "NL"].iloc[0]

    assert nl["approach_direction"] == "N"
    assert nl["approach_direction_label"] == "Northbound"
    assert nl["movement_type"] == "L"
    assert nl["movement_type_label"] == "Left turn"
    assert nl["display_label"] == "NL - Northbound Left turn"


def test_v2_diagram_data_excludes_v1_from_to_codes() -> None:
    diagram = build_v2_movement_diagram_data()

    assert set(diagram["movement_code"]).isdisjoint({"NS", "WE", "EN", "EW"})


def test_v2_diagram_data_totals_match_movement_summary() -> None:
    result = _dry_run_with_preset()
    diagram = build_v2_movement_diagram_data(
        movement_summary=result.movement,
        hourly_movement_pcu=result.hourly_movement_pcu,
        peaks=result.peaks,
    )
    summary = result.movement.groupby("movement_code", as_index=True)[["count", "pcu"]].sum()

    for code in ["NT", "WT", "ET"]:
        row = diagram.loc[diagram["movement_code"] == code].iloc[0]
        assert row["total_count"] == pytest.approx(summary.loc[code, "count"])
        assert row["total_pcu"] == pytest.approx(summary.loc[code, "pcu"])


def test_v2_visual_diagram_png_renders_headlessly() -> None:
    result = _dry_run_with_preset()
    diagram = build_v2_movement_diagram_data(
        movement_summary=result.movement,
        hourly_movement_pcu=result.hourly_movement_pcu,
        peaks=result.peaks,
    )

    png = render_v2_movement_diagram_png(diagram)

    assert len(png) > 1000
    assert png.startswith(b"\x89PNG\r\n\x1a\n")


def test_v2_generated_workbook_includes_movement_diagram_data_sheet() -> None:
    result = _dry_run_with_preset()
    workbook = load_workbook(BytesIO(export_v2_generated_workbook(result, setup=_setup())), data_only=False)

    assert "Movement_Diagram_Data" in workbook.sheetnames
    rows = list(workbook["Movement_Diagram_Data"].iter_rows(min_row=2, values_only=True))
    assert [row[0] for row in rows] == APPROACH_MOVEMENT_CODES
    assert rows[0][0:6] == ("NL", "N", "Northbound", "L", "Left turn", "NL - Northbound Left turn")
    assert {row[0] for row in rows}.isdisjoint({"NS", "WE", "EN"})


def test_v2_generated_workbook_keeps_v2_normalized_and_movement_labels() -> None:
    result = _dry_run_with_preset()
    workbook = load_workbook(BytesIO(export_v2_generated_workbook(result, setup=_setup())), data_only=False)

    normalized_headers = [cell.value for cell in workbook["Normalized_Data"][1]]
    movement_rows = list(workbook["Movement_Summary"].iter_rows(min_row=2, values_only=True))

    for column in [
        "movement_code_scheme",
        "movement_code",
        "output_movement_code",
        "approach_direction",
        "movement_type",
    ]:
        assert column in normalized_headers
    assert [row[1] for row in movement_rows] == APPROACH_MOVEMENT_CODES
    assert all(row[0] == MOVEMENT_SCHEME_V2 for row in movement_rows)


def test_v2_generated_export_does_not_use_v1_template_map() -> None:
    result = _dry_run_with_preset()

    with patch("tmc_processor.exporter.load_report_template_resources") as loader:
        workbook_bytes = export_v2_generated_workbook(result, setup=_setup())

    assert workbook_bytes.startswith(b"PK")
    loader.assert_not_called()


def test_v2_template_workbook_and_map_exist() -> None:
    assert V2_TEMPLATE_WORKBOOK.exists()
    assert V2_TEMPLATE_MAP.exists()


def test_v2_dry_run_result_exports_template_workbook_bytes() -> None:
    result = _dry_run_with_preset()

    workbook_bytes = export_v2_template_workbook(
        result,
        setup=_custom_setup(),
        mapping=_v2_preset_mapping(_raw_sheets()),
        generated_at="2026-05-26T08:00:00Z",
    )

    assert workbook_bytes.startswith(b"PK")
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
    assert "Summary" in workbook.sheetnames
    assert "Movement_Diagram_Data" in workbook.sheetnames

    summary = workbook["Summary"]
    assert [summary.cell(9, column).value for column in range(23, 39)] == APPROACH_MOVEMENT_CODES
    assert {"NS", "WE", "EN", "EW"}.isdisjoint({summary.cell(9, column).value for column in range(23, 39)})

    metadata = _sheet_records(workbook, "Export_Metadata")
    assert metadata["template_version"] == "four_leg_approach_movement_v2"
    assert metadata["movement_code_scheme"] == MOVEMENT_SCHEME_V2
    assert metadata["export_template"] == "four_leg_approach_movement_v2"
    assert metadata["excel_template_mode_supported"] is False
    assert metadata["native_template_export_supported"] is False

    assert summary["B2"].value == "Custom v2 TMC Report"
    assert summary["E5"].value == "Approach V2 Test Point"
    assert summary["K5"].value == "2026-05-26"
    assert summary["E6"].value == "Phase I2C QA"
    assert summary["G12"].value == "North Destination"
    assert summary["O32"].value == "South Destination"
    assert summary["Q17"].value == "East Destination"
    assert summary["D26"].value == "West Destination"
    assert summary["K11"].value == "North Road"
    assert summary["K32"].value == "South Road"
    assert summary["R19"].value == "East Road"
    assert summary["D24"].value == "West Road"
    assert summary["E35"].value == "Custom approach movement caption"

    hourly_headers = [cell.value for cell in workbook["Hourly_Movement_PCU"][1]]
    assert hourly_headers[1:17] == APPROACH_MOVEMENT_CODES
    movement_rows = list(workbook["Movement_Summary"].iter_rows(min_row=2, values_only=True))
    assert [row[1] for row in movement_rows] == APPROACH_MOVEMENT_CODES
    assert workbook["Movement_Diagram_Data"]["A2"].value == "NL"


def test_v2_template_export_uses_v2_template_map_only() -> None:
    result = _dry_run_with_preset()

    with patch("tmc_processor.exporter.load_report_template_resources", wraps=exporter_module.load_report_template_resources) as loader:
        workbook_bytes = export_v2_template_workbook(result, setup=_setup(), mapping=_v2_preset_mapping(_raw_sheets()))

    assert workbook_bytes.startswith(b"PK")
    args, _ = loader.call_args
    assert Path(args[0]).name == V2_TEMPLATE_WORKBOOK.name
    assert Path(args[1]).name == V2_TEMPLATE_MAP.name


def test_v2_com_template_export_selects_v2_resources_and_does_not_touch_source() -> None:
    result = _dry_run_with_preset()
    before = V2_TEMPLATE_WORKBOOK.stat().st_mtime_ns
    calls: list[dict[str, object]] = []

    def fake_com_export(**kwargs: object) -> bytes:
        calls.append(kwargs)
        return b"PK-v2-com"

    with patch("tmc_processor.exporter._export_workbook_with_excel_com", side_effect=fake_com_export):
        workbook_bytes = export_v2_template_workbook_com(
            result,
            setup=_custom_setup(),
            mapping=_v2_preset_mapping(_raw_sheets()),
            generated_at="2026-05-26T08:00:00Z",
        )

    after = V2_TEMPLATE_WORKBOOK.stat().st_mtime_ns
    assert workbook_bytes == b"PK-v2-com"
    assert before == after
    assert len(calls) == 1
    call = calls[0]
    assert Path(str(call["template_path"])).name == V2_TEMPLATE_WORKBOOK.name
    assert Path(str(call["template_map_path"])).name == V2_TEMPLATE_MAP.name
    assert call["movement_code_scheme"] == MOVEMENT_SCHEME_V2
    assert call["diagram_movement_codes"] == APPROACH_MOVEMENT_CODES
    assert "Movement_Diagram_Data" in call["sheets"]


def test_v2_com_template_export_never_selects_v1_template_files() -> None:
    result = _dry_run_with_preset()

    with pytest.raises(ValueError, match="must not use v1 template files"):
        export_v2_template_workbook_com(
            result,
            setup=_setup(),
            mapping=_v2_preset_mapping(_raw_sheets()),
            template_path=str(ROOT / "templates" / "four_leg_tmc_report_template.xlsx"),
        )

    with pytest.raises(ValueError, match="must not use v1 template files"):
        export_v2_template_workbook_com(
            result,
            setup=_setup(),
            mapping=_v2_preset_mapping(_raw_sheets()),
            template_map_path=str(ROOT / "templates" / "four_leg_tmc_report_template_map.json"),
        )


def test_v2_openpyxl_template_helper_is_limited_non_visual_export() -> None:
    result = _dry_run_with_preset()

    workbook_bytes = export_v2_template_workbook(result, setup=_setup(), mapping=_v2_preset_mapping(_raw_sheets()))
    metadata = _sheet_records(load_workbook(BytesIO(workbook_bytes), read_only=True), "Export_Metadata")

    assert metadata["excel_template_mode_supported"] is False
    assert metadata["native_template_export_supported"] is False
    assert "not safe for visual Excel Template Mode" in metadata["v2_export_limitation_notes"]


def test_v2_template_export_rejects_v1_template_files_and_excel_com() -> None:
    result = _dry_run_with_preset()

    with pytest.raises(ValueError, match="must not use v1 template files"):
        export_v2_template_workbook(
            result,
            setup=_setup(),
            mapping=_v2_preset_mapping(_raw_sheets()),
            template_map_path=str(ROOT / "templates" / "four_leg_tmc_report_template_map.json"),
        )

    with pytest.raises(ValueError, match="Excel COM/native"):
        export_v2_template_workbook(
            result,
            setup=_setup(),
            mapping=_v2_preset_mapping(_raw_sheets()),
            use_excel_com_native_charts=True,
        )


def test_v2_template_export_preserves_hlookup_formula_structure() -> None:
    result = _dry_run_with_preset()
    workbook = load_workbook(
        BytesIO(export_v2_template_workbook(result, setup=_setup(), mapping=_v2_preset_mapping(_raw_sheets()))),
        data_only=False,
    )
    summary = workbook["Summary"]
    formula_cells = ["J29", "K29", "M14", "E20", "Q23"]
    formulas = [str(summary[cell].value or "") for cell in formula_cells]

    assert all(formula.startswith("=") for formula in formulas)
    assert all("$W$9:$AM$22" in formula or "$W$9:$AL$22" in formula for formula in formulas)
    assert not any(code in formula for formula in formulas for code in ["NS", "WE", "EN", "EW"])


def test_v2_generated_package_excludes_raw_inputs_and_includes_summary() -> None:
    result = _dry_run_with_preset()
    workbook_bytes = export_v2_generated_workbook(result, setup=_setup())
    package = create_v2_generated_export_package_zip(
        workbook_bytes=workbook_bytes,
        setup=_setup(),
        peaks=result.peaks,
        qc=result.qc,
        workbook_filename="v2_generated.xlsx",
        source_file_name="raw_input.xlsx",
        generated_at="2026-05-19T10:00:00Z",
    )

    with ZipFile(BytesIO(package)) as archive:
        names = set(archive.namelist())
        summary = archive.read("export_summary.txt").decode("utf-8")
        diagram_csv = archive.read("diagram/movement_diagram_data.csv").decode("utf-8")
        diagram_png = archive.read("diagram/movement_diagram.png")

    assert "v2_generated.xlsx" in names
    assert "export_summary.txt" in names
    assert "diagram/movement_diagram_data.csv" in names
    assert "diagram/movement_diagram.png" in names
    assert "raw_input.xlsx" not in names
    assert "Template version: generated_approach_movement_v2" in summary
    assert "NL, N,Northbound" not in diagram_csv
    assert "NL,N,Northbound,L,Left turn" in diagram_csv
    assert "NS," not in diagram_csv
    assert diagram_png.startswith(b"\x89PNG\r\n\x1a\n")
    assert len(diagram_png) > 1000
