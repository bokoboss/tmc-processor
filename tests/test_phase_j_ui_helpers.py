from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
import pandas as pd
import pytest

import app
from tmc_processor.batch import batch_inputs_ready
from tmc_processor.constants import AM_WINDOW, DEFAULT_PEAK_MODE, PM_WINDOW
from tmc_processor.importer import load_detected_sheets
from tmc_processor.mapping import apply_saved_mapping_to_sheets, default_mapping_for_sheets
from tmc_processor.mapping_preset import apply_mapping_preset_to_detected_sheets, load_mapping_preset
from tmc_processor.movement_scheme import MOVEMENT_SCHEME_V1, MOVEMENT_SCHEME_V2
from tmc_processor.pipeline import ProcessingResult, V2DryRunResult


ROOT = Path(__file__).resolve().parents[1]
DEMO_DIR = ROOT / "samples" / "demo"
RAW_WORKBOOK = DEMO_DIR / "DEMO_TMC1_FourLeg.xlsx"
V1_MAPPING_XLSX = DEMO_DIR / "DEMO_TMC1_FourLeg_mapping.xlsx"
V2_PRESET = DEMO_DIR / "DEMO_TMC1_FourLeg_approach_v2.mapping.json"


def _setup(scheme: str) -> dict[str, object]:
    return {
        "project_name": "Phase J UI Helper",
        "tmc_id": "PHASE-J",
        "tmc_name": "Phase J Demo",
        "survey_date_text": "2026-05-26",
        "movement_code_scheme": scheme,
    }


def _peak_windows() -> dict[str, tuple[str, str]]:
    return {"AM": AM_WINDOW, "PM": PM_WINDOW}


def _raw_sheets() -> dict[str, pd.DataFrame]:
    return load_detected_sheets(RAW_WORKBOOK)


def _v2_mapping(raw_sheets: dict[str, pd.DataFrame]) -> pd.DataFrame:
    preset = load_mapping_preset(V2_PRESET.read_bytes()).preset
    return apply_mapping_preset_to_detected_sheets(preset, list(raw_sheets)).mapping


def _v1_mapping(raw_sheets: dict[str, pd.DataFrame]) -> pd.DataFrame:
    return apply_saved_mapping_to_sheets(list(raw_sheets), pd.read_excel(V1_MAPPING_XLSX, sheet_name="Mapping"))


def _v2_result() -> tuple[V2DryRunResult, pd.DataFrame]:
    raw_sheets = _raw_sheets()
    mapping = _v2_mapping(raw_sheets)
    result = app._process_single_file_for_ui(
        raw_sheets=raw_sheets,
        mapping=mapping,
        setup=_setup(MOVEMENT_SCHEME_V2),
        detected_sheets=list(raw_sheets),
        peak_mode=DEFAULT_PEAK_MODE,
        peak_windows=_peak_windows(),
        pce_factors={},
    )
    assert isinstance(result, V2DryRunResult)
    return result, mapping


def test_single_file_ui_helper_processes_v2_through_dry_run_path() -> None:
    result, _ = _v2_result()

    assert result.movement_code_scheme == MOVEMENT_SCHEME_V2
    assert not result.normalized.empty
    assert set(result.normalized["movement_code_scheme"]) == {MOVEMENT_SCHEME_V2}


def test_v2_result_feeds_peak_review_hourly_data_helper() -> None:
    result, mapping = _v2_result()

    hourly = app._hourly_movement_for_ui(result, mapping)

    assert not hourly.empty
    assert "NT" in hourly.columns
    assert "Total" in hourly.columns


def test_v2_ui_export_helper_produces_generated_workbook() -> None:
    result, mapping = _v2_result()

    workbook_bytes = app._export_single_file_for_ui(
        result=result,
        mapping=mapping,
        setup=_setup(MOVEMENT_SCHEME_V2),
        export_mode=app.SAFE_PNG_EXPORT_MODE,
        use_template_report_layout=False,
        use_excel_com_native_charts=False,
        source_file_name=RAW_WORKBOOK.name,
        generated_at="2026-05-26 12:00:00",
    )
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True)

    assert "Hourly_Movement_PCU" in workbook.sheetnames
    assert "Movement_Diagram_Data" in workbook.sheetnames


def test_v2_ui_export_helper_blocks_excel_template_mode_without_openpyxl_template_export(monkeypatch: pytest.MonkeyPatch) -> None:
    result, mapping = _v2_result()

    def fail_if_called(*args: object, **kwargs: object) -> bytes:
        raise AssertionError("v2 UI Excel Template Mode must not call COM export when COM is unavailable")

    monkeypatch.setattr(app, "export_v2_template_workbook_com", fail_if_called)

    with pytest.raises(ValueError, match="Excel Template Mode สำหรับ approach_movement ต้องใช้ Excel COM"):
        app._export_single_file_for_ui(
            result=result,
            mapping=mapping,
            setup=_setup(MOVEMENT_SCHEME_V2),
            export_mode=app.EXCEL_TEMPLATE_EXPORT_MODE,
            use_template_report_layout=True,
            use_excel_com_native_charts=False,
            source_file_name=RAW_WORKBOOK.name,
            generated_at="2026-05-26 12:00:00",
        )


def test_v2_excel_com_native_mode_uses_com_template_export(monkeypatch: pytest.MonkeyPatch) -> None:
    result, mapping = _v2_result()
    calls: list[dict[str, object]] = []

    def fake_com_export(*args: object, **kwargs: object) -> bytes:
        calls.append({"args": args, "kwargs": kwargs})
        return b"PK-v2-com"

    monkeypatch.setattr(app, "export_v2_template_workbook_com", fake_com_export)

    workbook_bytes = app._export_single_file_for_ui(
        result=result,
        mapping=mapping,
        setup=_setup(MOVEMENT_SCHEME_V2),
        export_mode=app.EXCEL_TEMPLATE_EXPORT_MODE,
        use_template_report_layout=True,
        use_excel_com_native_charts=True,
        source_file_name=RAW_WORKBOOK.name,
        generated_at="2026-05-26 12:00:00",
    )

    assert workbook_bytes == b"PK-v2-com"
    assert len(calls) == 1
    assert calls[0]["kwargs"]["setup"]["movement_code_scheme"] == MOVEMENT_SCHEME_V2


def test_v2_batch_analysis_remains_blocked() -> None:
    assert not batch_inputs_ready(
        uploaded_workbook_count=1,
        mapping_available=True,
        pce_factors_ready=True,
        movement_code_scheme=MOVEMENT_SCHEME_V2,
    )


def test_v1_single_file_ui_helper_still_uses_processing_result_path() -> None:
    raw_sheets = _raw_sheets()
    mapping = _v1_mapping(raw_sheets)

    result = app._process_single_file_for_ui(
        raw_sheets=raw_sheets,
        mapping=mapping,
        setup=_setup(MOVEMENT_SCHEME_V1),
        detected_sheets=list(raw_sheets),
        peak_mode=DEFAULT_PEAK_MODE,
        peak_windows=_peak_windows(),
        pce_factors={},
    )

    assert isinstance(result, ProcessingResult)
    assert not isinstance(result, V2DryRunResult)
    assert not result.normalized.empty


def test_v1_batch_readiness_unchanged() -> None:
    assert batch_inputs_ready(
        uploaded_workbook_count=1,
        mapping_available=True,
        pce_factors_ready=True,
        movement_code_scheme=MOVEMENT_SCHEME_V1,
    )


def test_project_session_preserves_v2_mapping_scheme() -> None:
    raw_sheets = _raw_sheets()
    mapping = _v2_mapping(raw_sheets)

    session = app.build_project_session(
        metadata={},
        directions={},
        mapping=mapping,
        movement_code_scheme=MOVEMENT_SCHEME_V2,
        detected_sheet_names=list(raw_sheets),
        pce_factors={},
    )

    assert session["mapping"]["movement_code_scheme"] == MOVEMENT_SCHEME_V2
