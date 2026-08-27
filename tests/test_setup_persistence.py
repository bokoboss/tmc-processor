from __future__ import annotations

from datetime import time
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
import pandas as pd
import streamlit as st
from streamlit.testing.v1 import AppTest

import app
from tmc_processor.excel_com_export import ExcelComStatus
from tmc_processor.exporter import export_workbook
from tmc_processor.pipeline import process_tmc
from tmc_processor.session import apply_session_to_state, build_project_session, session_from_json, session_to_json_bytes


CUSTOM_SETUP_VALUES = {
    "tmc_title": "รายงานทดสอบข้อมูลตั้งค่า",
    "tmc_id": "TMC-PERSIST-01",
    "survey_point": "แยกทดสอบ",
    "survey_date_text": "1 มกราคม 2569",
    "responsible_party": "ทีมทดสอบ",
    "caption_text": "คำอธิบายภาพทดสอบ",
    "north_road": "ถนนเหนือทดสอบ",
    "south_road": "ถนนใต้ทดสอบ",
    "east_road": "ถนนตะวันออกทดสอบ",
    "west_road": "ถนนตะวันตกทดสอบ",
    "north_label": "ปลายทางเหนือทดสอบ",
    "south_label": "ปลายทางใต้ทดสอบ",
    "east_label": "ปลายทางตะวันออกทดสอบ",
    "west_label": "ปลายทางตะวันตกทดสอบ",
    "show_u_turn": False,
}


def _run_app_for_apptest() -> None:
    import app as streamlit_app

    streamlit_app._run_streamlit_app()


def _widget_by_key(widgets, key: str):
    return next(widget for widget in widgets if widget.key == key)


def _widget_by_prefix(widgets, prefix: str):
    return next(widget for widget in widgets if widget.key == prefix or widget.key.startswith(f"{prefix}_"))


def _button_by_label(at: AppTest, label: str):
    return next(button for button in at.button if button.label == label)


def _mapping() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "raw_sheet": "S1",
                "raw_direction": "North",
                "movement_code": "NS",
                "source_stream": "mainline",
                "raw_movement_label": "North through",
                "from_leg": "N",
                "to_leg": "S",
                "turn_type": "through",
                "facility_type": "at_grade",
                "include_in_peak": True,
                "include_in_report": True,
                "aggregation_method": "sum",
            }
        ]
    )


def _raw_hour() -> dict[str, pd.DataFrame]:
    rows = []
    for minute in (0, 15, 30, 45):
        rows.append(
            {
                "raw_sheet": "S1",
                "raw_direction": "North",
                "time_start": time(7, minute),
                "time_end": time(8, 0) if minute == 45 else time(7, minute + 15),
                "vehicle_class": "PC<7",
                "count": 10,
            }
        )
    return {"S1": pd.DataFrame(rows)}


def _custom_widget_state(*, show_u_turn: bool = False) -> dict[str, object]:
    state: dict[str, object] = {}
    app.initialize_setup_state_once("demo-default", state)
    values = {**CUSTOM_SETUP_VALUES, "show_u_turn": show_u_turn}
    for field, value in values.items():
        state[app.SETUP_FIELD_WIDGET_KEYS[field]] = value
    return state


def _custom_setup(*, show_u_turn: bool = False) -> dict[str, object]:
    return app.build_setup_for_processing("demo.xlsx", _custom_widget_state(show_u_turn=show_u_turn))


def _sheet_records(workbook, sheet_name: str) -> dict[object, object]:
    return {
        row[0]: row[1]
        for row in workbook[sheet_name].iter_rows(min_row=2, values_only=True)
        if row[0] is not None
    }


def _processed_with_custom_setup(*, show_u_turn: bool = False):
    setup = _custom_setup(show_u_turn=show_u_turn)
    return process_tmc(
        raw_sheets=_raw_hour(),
        mapping=_mapping(),
        setup=setup,
        detected_sheets=["S1"],
        peak_windows={"AM": ("07:00", "08:00"), "PM": ("07:00", "08:00")},
        generate_workbook=False,
    )


def test_setup_state_initializes_once_without_overwriting_existing_values() -> None:
    state = {
        app.SETUP_STATE_KEY: {
            "tmc_title": CUSTOM_SETUP_VALUES["tmc_title"],
            "survey_point": CUSTOM_SETUP_VALUES["survey_point"],
            "show_u_turn": False,
        }
    }

    app.initialize_setup_state_once("uploaded-default", state)

    assert state[app.SETUP_STATE_KEY]["tmc_title"] == CUSTOM_SETUP_VALUES["tmc_title"]
    assert state[app.SETUP_STATE_KEY]["survey_point"] == CUSTOM_SETUP_VALUES["survey_point"]
    assert state[app.SETUP_FIELD_WIDGET_KEYS["tmc_title"]] == CUSTOM_SETUP_VALUES["tmc_title"]
    assert state[app.SETUP_FIELD_WIDGET_KEYS["show_u_turn"]] is False


def test_custom_setup_values_build_setup_for_processing() -> None:
    setup = _custom_setup(show_u_turn=True)

    assert setup["tmc_title"] == CUSTOM_SETUP_VALUES["tmc_title"]
    assert setup["survey_point"] == CUSTOM_SETUP_VALUES["survey_point"]
    assert setup["survey_date_text"] == CUSTOM_SETUP_VALUES["survey_date_text"]
    assert setup["responsible_party"] == CUSTOM_SETUP_VALUES["responsible_party"]
    assert setup["caption_text"] == CUSTOM_SETUP_VALUES["caption_text"]
    assert setup["north_road"] == CUSTOM_SETUP_VALUES["north_road"]
    assert setup["south_road"] == CUSTOM_SETUP_VALUES["south_road"]
    assert setup["east_road"] == CUSTOM_SETUP_VALUES["east_road"]
    assert setup["west_road"] == CUSTOM_SETUP_VALUES["west_road"]
    assert setup["north_label"] == CUSTOM_SETUP_VALUES["north_label"]
    assert setup["south_label"] == CUSTOM_SETUP_VALUES["south_label"]
    assert setup["east_label"] == CUSTOM_SETUP_VALUES["east_label"]
    assert setup["west_label"] == CUSTOM_SETUP_VALUES["west_label"]
    assert setup["show_u_turn"] is True


def test_processing_state_keeps_custom_setup_after_widget_keys_are_pruned() -> None:
    state = _custom_widget_state(show_u_turn=False)
    setup = app.build_setup_for_processing("demo.xlsx", state)
    result = process_tmc(
        raw_sheets=_raw_hour(),
        mapping=_mapping(),
        setup=setup,
        detected_sheets=["S1"],
        peak_windows={"AM": ("07:00", "08:00"), "PM": ("07:00", "08:00")},
        generate_workbook=False,
    )
    state["tmc_processed"] = {"result": result, "setup": setup}
    for widget_key in app.SETUP_FIELD_WIDGET_KEYS.values():
        state.pop(widget_key, None)

    app.initialize_setup_state_once("uploaded-default", state)
    restored_setup = app.build_setup_for_processing("demo.xlsx", state)

    assert state[app.SETUP_FIELD_WIDGET_KEYS["tmc_title"]] == CUSTOM_SETUP_VALUES["tmc_title"]
    assert state[app.SETUP_FIELD_WIDGET_KEYS["survey_point"]] == CUSTOM_SETUP_VALUES["survey_point"]
    assert state[app.SETUP_FIELD_WIDGET_KEYS["caption_text"]] == CUSTOM_SETUP_VALUES["caption_text"]
    assert restored_setup["north_label"] == CUSTOM_SETUP_VALUES["north_label"]
    assert restored_setup["show_u_turn"] is False


def test_safe_png_generated_workbook_contains_custom_setup_metadata() -> None:
    setup = _custom_setup(show_u_turn=True)
    result = _processed_with_custom_setup(show_u_turn=True)

    workbook_bytes = export_workbook(
        setup,
        _mapping(),
        result.normalized,
        result.qc,
        result.hourly,
        result.movement,
        result.vehicle,
        result.peaks,
        include_charts=False,
        include_diagram=False,
        export_mode="Safe PNG Export Mode",
        generated_at="2026-05-24T00:00:00Z",
    )

    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
    setup_rows = _sheet_records(workbook, "Setup")
    metadata_rows = _sheet_records(workbook, "Export_Metadata")

    assert metadata_rows["report_title"] == CUSTOM_SETUP_VALUES["tmc_title"]
    assert metadata_rows["survey_point"] == CUSTOM_SETUP_VALUES["survey_point"]
    assert metadata_rows["survey_date"] == CUSTOM_SETUP_VALUES["survey_date_text"]
    assert metadata_rows["responsible_party"] == CUSTOM_SETUP_VALUES["responsible_party"]
    assert metadata_rows["caption_text"] == CUSTOM_SETUP_VALUES["caption_text"]
    assert setup_rows["north_road"] == CUSTOM_SETUP_VALUES["north_road"]
    assert setup_rows["south_road"] == CUSTOM_SETUP_VALUES["south_road"]
    assert setup_rows["east_road"] == CUSTOM_SETUP_VALUES["east_road"]
    assert setup_rows["west_road"] == CUSTOM_SETUP_VALUES["west_road"]
    assert setup_rows["north_label"] == CUSTOM_SETUP_VALUES["north_label"]
    assert setup_rows["south_label"] == CUSTOM_SETUP_VALUES["south_label"]
    assert setup_rows["east_label"] == CUSTOM_SETUP_VALUES["east_label"]
    assert setup_rows["west_label"] == CUSTOM_SETUP_VALUES["west_label"]
    assert setup_rows["show_u_turn"] is True


def test_excel_template_openpyxl_path_writes_custom_setup_metadata() -> None:
    setup = _custom_setup(show_u_turn=False)
    result = _processed_with_custom_setup(show_u_turn=False)

    workbook_bytes = export_workbook(
        setup,
        _mapping(),
        result.normalized,
        result.qc,
        result.hourly,
        result.movement,
        result.vehicle,
        result.peaks,
        include_charts=False,
        include_diagram=False,
        use_template_report_layout=True,
        export_mode="Excel Template Mode",
    )

    worksheet = load_workbook(BytesIO(workbook_bytes), data_only=False)["TMC_Report"]

    assert worksheet["B2"].value == CUSTOM_SETUP_VALUES["tmc_title"]
    assert worksheet["E5"].value == CUSTOM_SETUP_VALUES["survey_point"]
    assert worksheet["K5"].value == CUSTOM_SETUP_VALUES["survey_date_text"]
    assert worksheet["E6"].value == CUSTOM_SETUP_VALUES["responsible_party"]
    assert worksheet["G12"].value == CUSTOM_SETUP_VALUES["north_label"]
    assert worksheet["O32"].value == CUSTOM_SETUP_VALUES["south_label"]
    assert worksheet["Q17"].value == CUSTOM_SETUP_VALUES["east_label"]
    assert worksheet["D26"].value == CUSTOM_SETUP_VALUES["west_label"]
    assert worksheet["K11"].value == CUSTOM_SETUP_VALUES["north_road"]
    assert worksheet["K32"].value == CUSTOM_SETUP_VALUES["south_road"]
    assert worksheet["R19"].value == CUSTOM_SETUP_VALUES["east_road"]
    assert worksheet["D24"].value == CUSTOM_SETUP_VALUES["west_road"]
    assert worksheet["E35"].value == CUSTOM_SETUP_VALUES["caption_text"]


def test_streamlit_apptest_keeps_analysis_actions_in_analyze_stage(monkeypatch) -> None:
    monkeypatch.setattr(
        app,
        "_probe_excel_com_for_ui",
        lambda force=False: ExcelComStatus(available=False, reason="mocked in AppTest", detail="", version=""),
    )

    single = AppTest.from_function(_run_app_for_apptest, default_timeout=20)
    single.run(timeout=30)
    stage_buttons = [button for button in single.button if button.key.startswith("workflow_tab_")]
    assert [button.label for button in stage_buttons] == app.workflow_stages_for_mode(app.WORKFLOW_SINGLE_MODE)
    assert "Clear source" in [button.label for button in single.button]

    _button_by_label(single, "Mapping").click()
    single.run(timeout=30)
    assert "Analyze TMC" not in [button.label for button in single.button]

    _button_by_label(single, "Analyze").click()
    single.run(timeout=30)
    assert "Analyze TMC" in [button.label for button in single.button]

    batch = AppTest.from_function(_run_app_for_apptest, default_timeout=20)
    batch.run(timeout=30)
    batch.radio[0].set_value(batch.radio[0].options[1])
    batch.run(timeout=30)
    assert "Clear Batch sources" in [button.label for button in batch.button]
    _button_by_label(batch, "Mapping").click()
    batch.run(timeout=30)
    assert _widget_by_key(batch.file_uploader, "batch_mapping_preset_upload")
    assert "Clear Mapping Preset" in [button.label for button in batch.button]
    assert "วิเคราะห์ Batch" not in [button.label for button in batch.button]

    _button_by_label(batch, "Analyze").click()
    batch.run(timeout=30)
    assert "วิเคราะห์ Batch" in [button.label for button in batch.button]


def _assert_analyze_setup_values(at: AppTest, expected: dict[str, object]) -> None:
    for field in ("am_peak_window_start", "am_peak_window_end", "pm_peak_window_start", "pm_peak_window_end"):
        widget_key = app.SETUP_FIELD_WIDGET_KEYS[field]
        assert _widget_by_key(at.time_input, widget_key).value == expected[field]
    assert _widget_by_key(at.selectbox, app.SETUP_FIELD_WIDGET_KEYS["peak_mode"]).value == expected["peak_mode"]
    assert _widget_by_key(at.text_input, app.SETUP_FIELD_WIDGET_KEYS["survey_period"]).value == expected["survey_period"]


def _exercise_analyze_setup_round_trip(at: AppTest, *, batch: bool = False) -> None:
    at.run(timeout=30)
    if batch:
        at.radio[0].set_value(at.radio[0].options[1])
        at.run(timeout=30)

    _button_by_label(at, "Analyze").click()
    at.run(timeout=30)
    defaults = {
        "am_peak_window_start": time(7, 0),
        "am_peak_window_end": time(12, 0),
        "pm_peak_window_start": time(15, 0),
        "pm_peak_window_end": time(19, 0),
        "peak_mode": app.DEFAULT_PEAK_MODE,
        "survey_period": app.DEFAULT_SURVEY_PERIOD,
    }
    _assert_analyze_setup_values(at, defaults)

    configured = {
        "am_peak_window_start": time(6, 0),
        "am_peak_window_end": time(10, 0),
        "pm_peak_window_start": time(15, 0),
        "pm_peak_window_end": time(19, 0),
        "peak_mode": "rolling_60min",
        "survey_period": "06.00 - 19.00",
    }
    for field in ("am_peak_window_start", "am_peak_window_end", "pm_peak_window_start", "pm_peak_window_end"):
        _widget_by_key(at.time_input, app.SETUP_FIELD_WIDGET_KEYS[field]).set_value(configured[field])
    _widget_by_key(at.selectbox, app.SETUP_FIELD_WIDGET_KEYS["peak_mode"]).set_value(configured["peak_mode"])
    _widget_by_key(at.text_input, app.SETUP_FIELD_WIDGET_KEYS["survey_period"]).set_value(configured["survey_period"])
    at.run(timeout=30)

    _button_by_label(at, "Data" if batch else "Mapping").click()
    at.run(timeout=30)
    _button_by_label(at, "Analyze").click()
    at.run(timeout=30)
    _assert_analyze_setup_values(at, configured)
    assert at.session_state[app.SETUP_STATE_KEY]["peak_mode"] == configured["peak_mode"]
    assert at.session_state[app.SETUP_STATE_KEY]["survey_period"] == configured["survey_period"]


def test_streamlit_apptest_analyze_setup_persists_across_single_stage_navigation(monkeypatch) -> None:
    monkeypatch.setattr(
        app,
        "_probe_excel_com_for_ui",
        lambda force=False: ExcelComStatus(available=False, reason="mocked in AppTest", detail="", version=""),
    )
    _exercise_analyze_setup_round_trip(AppTest.from_function(_run_app_for_apptest, default_timeout=20))


def test_streamlit_apptest_analyze_setup_persists_across_batch_stage_navigation(monkeypatch) -> None:
    monkeypatch.setattr(
        app,
        "_probe_excel_com_for_ui",
        lambda force=False: ExcelComStatus(available=False, reason="mocked in AppTest", detail="", version=""),
    )
    _exercise_analyze_setup_round_trip(AppTest.from_function(_run_app_for_apptest, default_timeout=20), batch=True)


def test_project_session_save_load_preserves_custom_setup_state_and_u_turn_toggle() -> None:
    st.session_state.clear()
    app.initialize_setup_state_once("demo.xlsx")
    for field, value in {**CUSTOM_SETUP_VALUES, "show_u_turn": True}.items():
        st.session_state[app.SETUP_FIELD_WIDGET_KEYS[field]] = value
    setup = app.build_setup_for_processing("demo.xlsx")

    session = build_project_session(metadata=setup, directions=setup)
    loaded = session_from_json(session_to_json_bytes(session)).session
    restored_state: dict[str, object] = {}
    changed = apply_session_to_state(loaded, restored_state)
    app.update_setup_state_from_widgets(restored_state)

    assert loaded["metadata"]["tmc_title"] == CUSTOM_SETUP_VALUES["tmc_title"]
    assert loaded["directions"]["caption_text"] == CUSTOM_SETUP_VALUES["caption_text"]
    assert loaded["directions"]["show_u_turn"] is True
    assert restored_state["tmc_title_input"] == CUSTOM_SETUP_VALUES["tmc_title"]
    assert restored_state["north_road_input"] == CUSTOM_SETUP_VALUES["north_road"]
    assert restored_state["show_u_turn_checkbox"] is True
    assert restored_state[app.SETUP_STATE_KEY]["show_u_turn"] is True
    assert "show_u_turn_checkbox" in changed


def test_streamlit_apptest_setup_values_survive_processing_and_export(monkeypatch) -> None:
    monkeypatch.setattr(
        app,
        "_probe_excel_com_for_ui",
        lambda force=False: ExcelComStatus(available=False, reason="mocked in AppTest", detail="", version=""),
    )
    root = Path(__file__).resolve().parents[1]
    at = AppTest.from_function(_run_app_for_apptest, default_timeout=20)
    at.run(timeout=30)
    at.file_uploader[0].set_value(
        (
            "DEMO_TMC1_FourLeg.xlsx",
            (root / "samples" / "demo" / "DEMO_TMC1_FourLeg.xlsx").read_bytes(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    )
    at.run(timeout=60)

    for field, value in {
        "tmc_title": CUSTOM_SETUP_VALUES["tmc_title"],
        "survey_point": CUSTOM_SETUP_VALUES["survey_point"],
        "survey_date_text": CUSTOM_SETUP_VALUES["survey_date_text"],
        "responsible_party": CUSTOM_SETUP_VALUES["responsible_party"],
        "north_label": CUSTOM_SETUP_VALUES["north_label"],
        "east_label": CUSTOM_SETUP_VALUES["east_label"],
        "south_label": CUSTOM_SETUP_VALUES["south_label"],
        "west_label": CUSTOM_SETUP_VALUES["west_label"],
        "north_road": CUSTOM_SETUP_VALUES["north_road"],
        "east_road": CUSTOM_SETUP_VALUES["east_road"],
        "south_road": CUSTOM_SETUP_VALUES["south_road"],
        "west_road": CUSTOM_SETUP_VALUES["west_road"],
        "caption_text": CUSTOM_SETUP_VALUES["caption_text"],
    }.items():
        _widget_by_key(at.text_input, app.SETUP_FIELD_WIDGET_KEYS[field]).set_value(value)
    _widget_by_key(at.checkbox, app.SETUP_FIELD_WIDGET_KEYS["show_u_turn"]).set_value(False)
    at.run(timeout=60)

    _button_by_label(at, "Mapping").click()
    at.run(timeout=60)
    _widget_by_key(at.file_uploader, "mapping_preset_upload").set_value(
        (
            "DEMO_TMC1_FourLeg.mapping.json",
            (root / "samples" / "demo" / "DEMO_TMC1_FourLeg.mapping.json").read_bytes(),
            "application/json",
        )
    )
    at.run(timeout=60)
    analyze_button = _button_by_label(at, "Analyze")
    assert analyze_button.disabled is False
    analyze_button.click()
    at.run(timeout=60)
    analyze_action = _button_by_label(at, "Analyze TMC")
    assert analyze_action.disabled is False
    analyze_action.click()
    at.run(timeout=90)
    assert at.session_state["active_workflow_tab"] == "Review"

    _button_by_label(at, "Data").click()
    at.run(timeout=60)
    for field in (
        "tmc_title",
        "survey_point",
        "survey_date_text",
        "responsible_party",
        "north_label",
        "north_road",
        "caption_text",
    ):
        assert _widget_by_key(at.text_input, app.SETUP_FIELD_WIDGET_KEYS[field]).value == CUSTOM_SETUP_VALUES[field]
    assert _widget_by_key(at.checkbox, app.SETUP_FIELD_WIDGET_KEYS["show_u_turn"]).value is False

    _button_by_label(at, "Export").click()
    at.run(timeout=60)
    export_action = _button_by_label(at, "สร้างรายงาน Excel")
    assert export_action.disabled is False
    export_action.click()
    at.run(timeout=120)

    output = at.session_state["tmc_output"]
    workbook = load_workbook(BytesIO(output["workbook_bytes"]), data_only=False)
    metadata_rows = _sheet_records(workbook, "Export_Metadata")
    setup_rows = _sheet_records(workbook, "Setup")
    assert metadata_rows["report_title"] == CUSTOM_SETUP_VALUES["tmc_title"]
    assert metadata_rows["survey_point"] == CUSTOM_SETUP_VALUES["survey_point"]
    assert metadata_rows["survey_date"] == CUSTOM_SETUP_VALUES["survey_date_text"]
    assert metadata_rows["responsible_party"] == CUSTOM_SETUP_VALUES["responsible_party"]
    assert metadata_rows["caption_text"] == CUSTOM_SETUP_VALUES["caption_text"]
    assert setup_rows["north_road"] == CUSTOM_SETUP_VALUES["north_road"]
    assert setup_rows["north_label"] == CUSTOM_SETUP_VALUES["north_label"]


def test_streamlit_apptest_explicit_single_clear_removes_upload_and_invalidates_state(monkeypatch) -> None:
    monkeypatch.setattr(
        app,
        "_probe_excel_com_for_ui",
        lambda force=False: ExcelComStatus(available=False, reason="mocked in AppTest", detail="", version=""),
    )
    root = Path(__file__).resolve().parents[1]
    at = AppTest.from_function(_run_app_for_apptest, default_timeout=20)
    at.run(timeout=30)
    _widget_by_key(at.file_uploader, "raw_tmc_upload").set_value(
        (
            "DEMO_TMC1_FourLeg.xlsx",
            (root / "samples" / "demo" / "DEMO_TMC1_FourLeg.xlsx").read_bytes(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    )
    at.run(timeout=60)

    assert app.SINGLE_SOURCE_UPLOAD_STATE_KEY in at.session_state
    clear_button = _button_by_label(at, "Clear source")
    assert clear_button.disabled is False
    clear_button.click()
    at.run(timeout=60)

    assert app.SINGLE_SOURCE_UPLOAD_STATE_KEY not in at.session_state
    workflow_state = at.session_state[app.WORKFLOW_STATE_KEY][app.WORKFLOW_SINGLE_MODE]
    assert workflow_state.readiness.source is False
    assert workflow_state.readiness.analysis is False
    assert workflow_state.readiness.review is False
    assert workflow_state.readiness.export is False
    assert _widget_by_prefix(at.file_uploader, "raw_tmc_upload").value is None

    _widget_by_prefix(at.file_uploader, "raw_tmc_upload").set_value(
        (
            "DEMO_TMC1_FourLeg.xlsx",
            (root / "samples" / "demo" / "DEMO_TMC1_FourLeg.xlsx").read_bytes(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    )
    at.run(timeout=60)
    assert at.session_state[app.SINGLE_SOURCE_UPLOAD_STATE_KEY]["name"] == "DEMO_TMC1_FourLeg.xlsx"


def test_streamlit_apptest_batch_clear_controls_remove_canonical_inputs(monkeypatch) -> None:
    monkeypatch.setattr(
        app,
        "_probe_excel_com_for_ui",
        lambda force=False: ExcelComStatus(available=False, reason="mocked in AppTest", detail="", version=""),
    )
    root = Path(__file__).resolve().parents[1]
    workbook_bytes = (root / "samples" / "demo" / "DEMO_TMC1_FourLeg.xlsx").read_bytes()
    preset_bytes = (root / "samples" / "demo" / "DEMO_TMC1_FourLeg.mapping.json").read_bytes()
    at = AppTest.from_function(_run_app_for_apptest, default_timeout=20)
    at.run(timeout=30)
    at.radio[0].set_value(at.radio[0].options[1])
    at.run(timeout=30)
    _widget_by_prefix(at.file_uploader, "batch_raw_tmc_uploads").set_value(
        [
            (
                "DEMO_TMC1_FourLeg.xlsx",
                workbook_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        ]
    )
    at.run(timeout=60)

    assert app.BATCH_SOURCE_UPLOAD_STATE_KEY in at.session_state
    _button_by_label(at, "Clear Batch sources").click()
    at.run(timeout=60)

    assert app.BATCH_SOURCE_UPLOAD_STATE_KEY not in at.session_state
    batch_state = at.session_state[app.WORKFLOW_STATE_KEY][app.WORKFLOW_BATCH_MODE]
    assert batch_state.readiness.source is False
    assert batch_state.readiness.analysis is False
    assert batch_state.readiness.review is False
    assert batch_state.readiness.export is False

    _widget_by_prefix(at.file_uploader, "batch_raw_tmc_uploads").set_value(
        [
            (
                "DEMO_TMC1_FourLeg.xlsx",
                workbook_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        ]
    )
    at.run(timeout=60)
    _button_by_label(at, "Mapping").click()
    at.run(timeout=60)
    _widget_by_key(at.file_uploader, "batch_mapping_preset_upload").set_value(
        (
            "DEMO_TMC1_FourLeg.mapping.json",
            preset_bytes,
            "application/json",
        )
    )
    at.run(timeout=60)

    assert app.BATCH_MAPPING_PRESET_UPLOAD_STATE_KEY in at.session_state
    _button_by_label(at, "Clear Mapping Preset").click()
    at.run(timeout=60)

    assert app.BATCH_MAPPING_PRESET_UPLOAD_STATE_KEY not in at.session_state
    batch_state = at.session_state[app.WORKFLOW_STATE_KEY][app.WORKFLOW_BATCH_MODE]
    assert batch_state.readiness.mapping is False
    assert batch_state.readiness.analysis is False
    assert batch_state.readiness.review is False
    assert batch_state.readiness.export is False
