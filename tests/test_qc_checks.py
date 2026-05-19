from __future__ import annotations

from datetime import time
from io import BytesIO

from openpyxl import load_workbook
import pandas as pd

from tmc_processor.exporter import export_workbook
from tmc_processor.pipeline import process_tmc


def _mapping(rows: list[dict[str, object]]) -> pd.DataFrame:
    defaults = {
        "raw_direction": "North",
        "source_stream": "mainline",
        "raw_movement_label": "North",
        "from_leg": "N",
        "to_leg": "S",
        "turn_type": "through",
        "facility_type": "at_grade",
        "include_in_peak": True,
        "include_in_report": True,
        "aggregation_method": "sum",
    }
    return pd.DataFrame([{**defaults, **row} for row in rows])


def _rows(sheet: str, movement_direction: str, vehicle_class: str, counts: list[int], start_hour: int = 7) -> list[dict[str, object]]:
    rows = []
    for index, count in enumerate(counts):
        minute = (index % 4) * 15
        hour = start_hour + (index // 4)
        end_hour = hour + 1 if minute == 45 else hour
        end_minute = 0 if minute == 45 else minute + 15
        rows.append(
            {
                "raw_sheet": sheet,
                "raw_direction": movement_direction,
                "time_start": time(hour, minute),
                "time_end": time(end_hour, end_minute),
                "vehicle_class": vehicle_class,
                "count": count,
            }
        )
    return rows


def _raw_sheet(sheet: str, movement_direction: str, vehicle_class: str, counts: list[int], start_hour: int = 7) -> pd.DataFrame:
    return pd.DataFrame(_rows(sheet, movement_direction, vehicle_class, counts, start_hour=start_hour))


def _process(raw_sheets: dict[str, pd.DataFrame], mapping: pd.DataFrame, **kwargs):
    return process_tmc(
        raw_sheets=raw_sheets,
        mapping=mapping,
        setup={"project_name": "QC test"},
        detected_sheets=list(raw_sheets),
        peak_windows={"AM": ("06:00", "10:00"), "PM": ("06:00", "10:00")},
        generate_workbook=False,
        **kwargs,
    )


def _qc_rows(result, check: str) -> pd.DataFrame:
    return result.qc[result.qc["check"] == check]


def test_zero_volume_mapped_movement_warning() -> None:
    mapping = _mapping([{"raw_sheet": "A", "movement_code": "NS"}])
    result = _process({"A": _raw_sheet("A", "North", "PC<7", [0, 0, 0, 0])}, mapping)

    rows = _qc_rows(result, "zero_volume_mapped_movement")
    assert not rows.empty
    assert set(rows["severity"]) == {"warning"}
    assert rows.iloc[0]["category"] == "movement_volume"


def test_high_movement_share_info_and_warning() -> None:
    mapping = _mapping(
        [
            {"raw_sheet": "A", "movement_code": "NS"},
            {"raw_sheet": "B", "movement_code": "EW", "raw_direction": "East", "from_leg": "E", "to_leg": "W"},
        ]
    )
    info_result = _process(
        {
            "A": _raw_sheet("A", "North", "PC<7", [13, 12, 13, 12]),
            "B": _raw_sheet("B", "East", "PC<7", [13, 12, 13, 12]),
        },
        mapping,
    )
    warning_result = _process(
        {
            "A": _raw_sheet("A", "North", "PC<7", [18, 17, 18, 17]),
            "B": _raw_sheet("B", "East", "PC<7", [8, 7, 8, 7]),
        },
        mapping,
    )

    assert "info" in set(_qc_rows(info_result, "high_single_movement_share")["severity"])
    assert "warning" in set(_qc_rows(warning_result, "high_single_movement_share")["severity"])


def test_heavy_vehicle_share_info_and_warning() -> None:
    mapping = _mapping([{"raw_sheet": "A", "movement_code": "NS"}])
    info_result = _process(
        {"A": pd.DataFrame(_rows("A", "North", "HT", [8, 7, 8, 7]) + _rows("A", "North", "PC<7", [18, 17, 18, 17]))},
        mapping,
    )
    warning_result = _process(
        {"A": pd.DataFrame(_rows("A", "North", "HT", [13, 12, 13, 12]) + _rows("A", "North", "PC<7", [13, 12, 13, 12]))},
        mapping,
    )

    assert "info" in set(_qc_rows(info_result, "high_heavy_vehicle_share")["severity"])
    assert "warning" in set(_qc_rows(warning_result, "high_heavy_vehicle_share")["severity"])


def test_low_phf_info_and_warning() -> None:
    mapping = _mapping([{"raw_sheet": "A", "movement_code": "NS"}])
    info_result = _process({"A": _raw_sheet("A", "North", "PC<7", [100, 100, 100, 20])}, mapping)
    warning_result = _process({"A": _raw_sheet("A", "North", "PC<7", [100, 0, 0, 0])}, mapping)

    assert "info" in set(_qc_rows(info_result, "low_peak_phf")["severity"])
    assert "warning" in set(_qc_rows(warning_result, "low_peak_phf")["severity"])


def test_peak_spike_info_and_warning() -> None:
    mapping = _mapping([{"raw_sheet": "A", "movement_code": "NS"}])
    info_counts = [20, 20, 20, 20, 40, 40, 40, 40, 20, 20, 20, 20]
    warning_counts = [20, 20, 20, 20, 60, 60, 60, 60, 20, 20, 20, 20]

    info_result = _process({"A": _raw_sheet("A", "North", "PC<7", info_counts, start_hour=6)}, mapping)
    warning_result = _process({"A": _raw_sheet("A", "North", "PC<7", warning_counts, start_hour=6)}, mapping)

    assert "info" in set(_qc_rows(info_result, "peak_spike_vs_adjacent_hours")["severity"])
    assert "warning" in set(_qc_rows(warning_result, "peak_spike_vs_adjacent_hours")["severity"])


def test_aggregation_info() -> None:
    mapping = _mapping(
        [
            {"raw_sheet": "A", "movement_code": "NS", "source_stream": "mainline"},
            {"raw_sheet": "B", "movement_code": "NS", "source_stream": "frontage"},
        ]
    )
    result = _process(
        {
            "A": _raw_sheet("A", "North", "PC<7", [10, 10, 10, 10]),
            "B": _raw_sheet("B", "North", "PC<7", [5, 5, 5, 5]),
        },
        mapping,
    )

    rows = _qc_rows(result, "aggregated_output_movement")
    assert not rows.empty
    assert rows.iloc[0]["severity"] == "info"
    assert rows.iloc[0]["category"] == "movement_aggregation"


def test_pce_override_info() -> None:
    mapping = _mapping([{"raw_sheet": "A", "movement_code": "NS"}])
    result = _process({"A": _raw_sheet("A", "North", "MC", [10, 10, 10, 10])}, mapping, pce_factors={"MC": 0.75})

    rows = _qc_rows(result, "pce_override")
    assert not rows.empty
    assert rows.iloc[0]["severity"] == "info"
    assert "MC" in rows.iloc[0]["detail"]


def test_qc_export_includes_new_messages() -> None:
    mapping = _mapping([{"raw_sheet": "A", "movement_code": "NS"}])
    result = _process({"A": _raw_sheet("A", "North", "PC<7", [0, 0, 0, 0])}, mapping)
    workbook_bytes = export_workbook(
        {"project_name": "QC export"},
        mapping,
        result.normalized,
        result.qc,
        result.hourly,
        result.movement,
        result.vehicle,
        result.peaks,
        include_charts=False,
        include_diagram=False,
    )

    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
    sheet = workbook["QC_Check"]
    headers = [cell.value for cell in sheet[1]]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    check_index = headers.index("check")

    assert "category" in headers
    assert any(row[check_index] == "zero_volume_mapped_movement" for row in rows)
