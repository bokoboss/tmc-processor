from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from tmc_processor.diagram import MOVEMENT_CODES
from tmc_processor.movement_scheme import APPROACH_MOVEMENT_CODES, MOVEMENT_SCHEME_V1, MOVEMENT_SCHEME_V2
from tmc_processor.template_integrity import (
    load_template_map_for_integrity,
    map_movement_code_leaks,
    mapped_approach_table_order,
    mapped_diagram_movement_order,
    mapped_hourly_movement_order,
    verify_template_integrity,
)


ROOT = Path(__file__).resolve().parents[1]
TEMPLATES = ROOT / "templates"
V1_TEMPLATE_PATH = TEMPLATES / "four_leg_tmc_report_template.xlsx"
V1_MAP_PATH = TEMPLATES / "four_leg_tmc_report_template_map.json"
V2_TEMPLATE_PATH = TEMPLATES / "four_leg_tmc_report_template_approach_v2.xlsx"
V2_MAP_PATH = TEMPLATES / "four_leg_tmc_report_template_approach_v2_map.json"
DESIGN_DOC_PATH = ROOT / "docs" / "APPROACH_MOVEMENT_TEMPLATE_DESIGN.md"


def _json(path: Path) -> dict[str, object]:
    return json.loads(path.read_text(encoding="utf-8"))


def _cell_in_range(cell_ref: str, cell_range: str) -> bool:
    cell_col, cell_row, _, _ = range_boundaries(cell_ref)
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    return min_col <= cell_col <= max_col and min_row <= cell_row <= max_row


def _cell_in_any_range(cell_ref: str, ranges: list[str]) -> bool:
    return any(_cell_in_range(cell_ref, cell_range) for cell_range in ranges)


def test_v2_template_map_exists_and_declares_integrity_metadata() -> None:
    assert V2_MAP_PATH.exists()
    mapping = _json(V2_MAP_PATH)

    assert mapping["template_version"] == "four_leg_approach_movement_v2"
    assert mapping["movement_code_scheme"] == MOVEMENT_SCHEME_V2
    assert mapping["movement_code_order"] == APPROACH_MOVEMENT_CODES
    assert mapping["template_workbook"] == V2_TEMPLATE_PATH.name


def test_v2_template_map_movement_headers_follow_approach_order() -> None:
    mapping = load_template_map_for_integrity(V2_MAP_PATH)

    assert mapped_hourly_movement_order(mapping) == APPROACH_MOVEMENT_CODES
    assert mapped_diagram_movement_order(mapping) == APPROACH_MOVEMENT_CODES
    assert mapped_approach_table_order(mapping) == APPROACH_MOVEMENT_CODES


def test_v2_template_map_has_no_v1_from_to_code_leakage() -> None:
    mapping = load_template_map_for_integrity(V2_MAP_PATH)
    v1_only_codes = set(MOVEMENT_CODES) - set(APPROACH_MOVEMENT_CODES)

    assert map_movement_code_leaks(mapping, v1_only_codes) == []
    assert {"NS", "WE", "EN", "EW"}.isdisjoint(set(mapped_hourly_movement_order(mapping)))
    assert {"NS", "WE", "EN", "EW"}.isdisjoint(set(mapped_diagram_movement_order(mapping)))


def test_v2_template_map_does_not_overwrite_v1_template_map() -> None:
    v1_mapping = _json(V1_MAP_PATH)
    v2_mapping = _json(V2_MAP_PATH)

    assert V1_MAP_PATH.name == "four_leg_tmc_report_template_map.json"
    assert V2_MAP_PATH.name == "four_leg_tmc_report_template_approach_v2_map.json"
    assert v1_mapping["template_workbook"] == "four_leg_tmc_report_template.xlsx"
    assert v2_mapping["template_workbook"] == "four_leg_tmc_report_template_approach_v2.xlsx"
    assert "template_version" not in v1_mapping
    assert v1_mapping["hourly_movement_table"]["columns"]["NS"] == "X"
    assert "NL" not in v1_mapping["hourly_movement_table"]["columns"]


def test_v1_template_integrity_still_matches_v1_map() -> None:
    result = verify_template_integrity(V1_TEMPLATE_PATH, V1_MAP_PATH, MOVEMENT_SCHEME_V1)

    assert result.errors == ()


def test_v2_template_integrity_map_is_ready_but_workbook_finalization_is_deferred() -> None:
    mapping = load_template_map_for_integrity(V2_MAP_PATH)
    assert mapped_hourly_movement_order(mapping) == APPROACH_MOVEMENT_CODES
    assert mapped_diagram_movement_order(mapping) == APPROACH_MOVEMENT_CODES

    if V2_TEMPLATE_PATH.exists():
        result = verify_template_integrity(V2_TEMPLATE_PATH, V2_MAP_PATH, MOVEMENT_SCHEME_V2)
        assert result.errors == ()
        workbook = load_workbook(V2_TEMPLATE_PATH, data_only=False)
        assert "Summary" in workbook.sheetnames
        worksheet = workbook["Summary"]
        columns = mapping["hourly_movement_table"]["columns"]
        header_row = mapping["hourly_movement_table"]["header_row"]
        assert [worksheet[f"{columns[code]}{header_row}"].value for code in APPROACH_MOVEMENT_CODES] == APPROACH_MOVEMENT_CODES
        assert {"NS", "WE", "EN", "EW"}.isdisjoint(
            {worksheet[f"{columns[code]}{header_row}"].value for code in APPROACH_MOVEMENT_CODES}
        )
    else:
        design_doc = DESIGN_DOC_PATH.read_text(encoding="utf-8")
        assert "v2 template XLSX finalization is deferred" in design_doc


def test_v2_template_workbook_declares_expected_marker_and_summary_headers() -> None:
    mapping = load_template_map_for_integrity(V2_MAP_PATH)
    columns = mapping["hourly_movement_table"]["columns"]
    header_row = mapping["hourly_movement_table"]["header_row"]
    header_cells = [f"{columns[code]}{header_row}" for code in APPROACH_MOVEMENT_CODES]

    assert f"{header_cells[0]}:{header_cells[-1]}" == "W9:AL9"

    for data_only in (False, True):
        workbook = load_workbook(V2_TEMPLATE_PATH, read_only=True, data_only=data_only)
        worksheet = workbook["Summary"]
        headers = [worksheet[cell].value for cell in header_cells]
        marker_values = {
            str(cell.value)
            for row in worksheet.iter_rows(min_row=1, max_row=80, min_col=1, max_col=60)
            for cell in row
            if cell.value is not None
        }
        workbook.close()

        assert headers == APPROACH_MOVEMENT_CODES
        assert set(headers).isdisjoint(set(MOVEMENT_CODES) - set(APPROACH_MOVEMENT_CODES))
        assert "Template version: four_leg_approach_movement_v2" in marker_values
        assert "Movement code scheme: approach_movement" in marker_values


def test_v2_template_map_structure_matches_manual_workbook() -> None:
    mapping = load_template_map_for_integrity(V2_MAP_PATH)
    workbook = load_workbook(V2_TEMPLATE_PATH, data_only=False)
    worksheet = workbook["Summary"]

    assert workbook.sheetnames == mapping["workbook_structure"]["sheets"]
    assert worksheet.calculate_dimension() == mapping["workbook_structure"]["used_range"]
    assert sorted(str(merged_range) for merged_range in worksheet.merged_cells.ranges) == sorted(
        mapping["workbook_structure"]["merged_ranges"]
    )


def test_v2_template_road_and_destination_label_cells_remain_metadata_mapped() -> None:
    mapping = load_template_map_for_integrity(V2_MAP_PATH)
    movement_info = mapping["movement_diagram_cells"]
    writable_ranges = mapping["writable_ranges"]

    label_cells = {
        label_key: label_info["cell"]
        for label_group in ("direction_labels", "road_labels")
        for label_key, label_info in movement_info[label_group].items()
    }

    assert label_cells == {
        "north_label": "G12",
        "south_label": "O32",
        "east_label": "Q17",
        "west_label": "D26",
        "north_road": "K11",
        "south_road": "K32",
        "east_road": "R19",
        "west_road": "D24",
    }
    assert all(_cell_in_any_range(cell, writable_ranges) for cell in label_cells.values())

    workbook = load_workbook(V2_TEMPLATE_PATH, read_only=True, data_only=False)
    worksheet = workbook["Summary"]
    for cell in label_cells.values():
        assert worksheet[cell].value not in APPROACH_MOVEMENT_CODES
    workbook.close()


def test_missing_v2_template_workbook_is_reported_by_integrity_checker() -> None:
    if V2_TEMPLATE_PATH.exists():
        return

    result = verify_template_integrity(V2_TEMPLATE_PATH, V2_MAP_PATH, MOVEMENT_SCHEME_V2)

    assert not result.ok
    assert any("Template workbook not found" in error for error in result.errors)
