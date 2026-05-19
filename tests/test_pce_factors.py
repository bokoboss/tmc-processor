from __future__ import annotations

from datetime import time
from io import BytesIO

from openpyxl import load_workbook
import pandas as pd

from tmc_processor.constants import DEFAULT_PCE_FACTORS
from tmc_processor.diagram import movement_diagram_values
from tmc_processor.exporter import export_workbook
from tmc_processor.metadata import APP_VERSION, TEMPLATE_VERSION
from tmc_processor.normalizer import normalize
from tmc_processor.pcu import get_default_pce_factors, normalize_pce_factors, pce_factors_equal, validate_pce_factors
from tmc_processor.pipeline import process_tmc
from tmc_processor.session import apply_session_to_state, build_project_session, session_from_json, session_to_json_bytes
from tmc_processor.summaries import hourly_movement_pcu


def _mapping() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "raw_sheet": "S1",
                "raw_direction": "North",
                "movement_code": "NS",
                "source_stream": "mainline",
                "raw_movement_label": "North",
                "from_leg": "N",
                "to_leg": "S",
                "turn_type": "through",
                "facility_type": "at_grade",
                "include_in_peak": True,
                "include_in_report": True,
                "aggregation_method": "sum",
            }
        ]
    )


def _raw_hour(mc_count: int = 10, pc_count: int = 5) -> dict[str, pd.DataFrame]:
    rows = []
    for minute in (0, 15, 30, 45):
        start = time(7, minute)
        end = time(8, 0) if minute == 45 else time(7, minute + 15)
        rows.extend(
            [
                {
                    "raw_sheet": "S1",
                    "raw_direction": "North",
                    "time_start": start,
                    "time_end": end,
                    "vehicle_class": "MC",
                    "count": mc_count,
                },
                {
                    "raw_sheet": "S1",
                    "raw_direction": "North",
                    "time_start": start,
                    "time_end": end,
                    "vehicle_class": "PC<7",
                    "count": pc_count,
                },
            ]
        )
    return {"S1": pd.DataFrame(rows)}


def _process(**kwargs):
    return process_tmc(
        raw_sheets=_raw_hour(),
        mapping=_mapping(),
        setup={"project_name": "PCE test"},
        detected_sheets=["S1"],
        peak_windows={"AM": ("07:00", "08:00"), "PM": ("07:00", "08:00")},
        generate_workbook=False,
        **kwargs,
    )


def test_default_pce_behavior_matches_explicit_defaults() -> None:
    implicit = normalize(_raw_hour(), _mapping(), setup={})
    explicit = normalize(_raw_hour(), _mapping(), setup={}, pce_factors=get_default_pce_factors())

    pd.testing.assert_series_equal(implicit["pcu"], explicit["pcu"], check_names=False)
    assert implicit.loc[implicit["vehicle_class"] == "MC", "pce_factor"].iloc[0] == DEFAULT_PCE_FACTORS["MC"]
    # Bicy=0.0 is an intentional historical default, not a general allowance for zero PCE overrides.
    assert normalize_pce_factors({"Bicy": 0})["Bicy"] == 0


def test_custom_pce_changes_normalized_pcu() -> None:
    default_result = _process()
    custom_result = _process(pce_factors={"MC": 1.0})

    assert custom_result.normalized["pcu"].sum() > default_result.normalized["pcu"].sum()
    assert custom_result.normalized.loc[custom_result.normalized["vehicle_class"] == "MC", "pce_factor"].eq(1.0).all()


def test_missing_pce_factor_falls_back_to_default() -> None:
    factors = normalize_pce_factors({"MC": 0.5})

    assert factors["MC"] == 0.5
    assert factors["PC<7"] == DEFAULT_PCE_FACTORS["PC<7"]


def test_invalid_pce_factor_falls_back_safely() -> None:
    validation = validate_pce_factors({"MC": 0, "PC<7": "bad", "Bicy": -1})

    assert validation.factors["MC"] == DEFAULT_PCE_FACTORS["MC"]
    assert validation.factors["PC<7"] == DEFAULT_PCE_FACTORS["PC<7"]
    assert validation.factors["Bicy"] == DEFAULT_PCE_FACTORS["Bicy"]
    assert len(validation.warnings) == 3


def test_pce_factor_equality_detects_stale_results() -> None:
    assert pce_factors_equal({"MC": DEFAULT_PCE_FACTORS["MC"]}, get_default_pce_factors())
    assert not pce_factors_equal({"MC": DEFAULT_PCE_FACTORS["MC"]}, {"MC": 0.75})


def test_selected_pce_affects_hourly_movement_pcu() -> None:
    default_hourly = hourly_movement_pcu(_process().normalized, _mapping())
    custom_hourly = hourly_movement_pcu(_process(pce_factors={"MC": 1.0}).normalized, _mapping())

    default_total = int(default_hourly["Total"].iloc[-1])
    custom_total = int(custom_hourly["Total"].iloc[-1])
    assert custom_total > default_total


def test_selected_pce_affects_peak_phf() -> None:
    default_peak = _process().peaks.loc[lambda frame: frame["period"] == "AM", "hourly_pcu"].iloc[0]
    custom_peak = _process(pce_factors={"MC": 1.0}).peaks.loc[lambda frame: frame["period"] == "AM", "hourly_pcu"].iloc[0]

    assert custom_peak > default_peak


def test_selected_pce_affects_diagram_data_values() -> None:
    custom_result = _process(pce_factors={"MC": 1.0})
    hourly = hourly_movement_pcu(custom_result.normalized, _mapping())
    values = movement_diagram_values(hourly, custom_result.peaks)

    assert values["NS"]["total"] == 60


def test_pce_factors_round_trip_through_project_session() -> None:
    session = build_project_session(pce_factors={"MC": 0.75})
    loaded = session_from_json(session_to_json_bytes(session))
    state: dict[str, object] = {}

    apply_session_to_state(loaded.session, state)

    assert loaded.session["pce_factors"]["MC"] == 0.75
    assert any(row["vehicle_class"] == "MC" and row["pce_factor"] == 0.75 for row in state["pce_factors_table"])

    older_session = session_from_json(b'{"schema_version": 1}')
    assert older_session.session["pce_factors"]["MC"] == DEFAULT_PCE_FACTORS["MC"]
    assert older_session.session["app_version"] == ""
    assert older_session.session["template_version"] == ""


def test_project_session_includes_version_metadata() -> None:
    session = build_project_session(pce_factors={"MC": 0.75})

    assert session["app_version"] == APP_VERSION
    assert session["template_version"] == TEMPLATE_VERSION


def test_export_workbook_includes_pce_factor_traceability() -> None:
    result = _process(pce_factors={"MC": 0.75})
    workbook_bytes = export_workbook(
        {"project_name": "PCE export"},
        _mapping(),
        result.normalized,
        result.qc,
        result.hourly,
        result.movement,
        result.vehicle,
        result.peaks,
        include_charts=False,
        include_diagram=False,
        pce_factors={"MC": 0.75},
    )

    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
    assert "PCE_Factors" in workbook.sheetnames
    rows = list(workbook["PCE_Factors"].iter_rows(min_row=2, values_only=True))
    assert ("MC", 0.75, "user_override") in rows
    assert ("PC<7", DEFAULT_PCE_FACTORS["PC<7"], "default") in rows


def test_export_workbook_includes_export_metadata() -> None:
    result = _process(pce_factors={"MC": 0.75})
    workbook_bytes = export_workbook(
        {"project_name": "Metadata export", "survey_point": "Main & 1st", "tmc_title": "TMC-01"},
        _mapping(),
        result.normalized,
        result.qc,
        result.hourly,
        result.movement,
        result.vehicle,
        result.peaks,
        include_charts=False,
        include_diagram=False,
        export_mode="Safe PNG Export Mode",
        source_file_name=r"C:\private\raw_input.xlsx",
        generated_at="2026-05-19T10:00:00Z",
    )

    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
    assert "Export_Metadata" in workbook.sheetnames
    metadata = {
        row[0]: row[1]
        for row in workbook["Export_Metadata"].iter_rows(min_row=2, values_only=True)
        if row[0]
    }
    assert metadata["app_version"] == APP_VERSION
    assert metadata["template_version"] == TEMPLATE_VERSION
    assert metadata["generated_at"] == "2026-05-19T10:00:00Z"
    assert metadata["export_mode"] == "Safe PNG Export Mode"
    assert metadata["source_file_name"] == "raw_input.xlsx"
    assert metadata["survey_point"] == "Main & 1st"


def test_template_layout_export_includes_export_metadata() -> None:
    result = _process(pce_factors={"MC": 0.75})
    workbook_bytes = export_workbook(
        {"project_name": "Template metadata export", "survey_point": "Template Point"},
        _mapping(),
        result.normalized,
        result.qc,
        result.hourly,
        result.movement,
        result.vehicle,
        result.peaks,
        include_charts=False,
        include_diagram=False,
        use_template_report_layout=True,
        export_mode="Excel Template Mode",
        source_file_name="template_input.xlsx",
        generated_at="2026-05-19T10:00:00Z",
    )

    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
    metadata = {
        row[0]: row[1]
        for row in workbook["Export_Metadata"].iter_rows(min_row=2, values_only=True)
        if row[0]
    }
    assert metadata["app_version"] == APP_VERSION
    assert metadata["template_version"] == TEMPLATE_VERSION
    assert metadata["export_mode"] == "Excel Template Mode"
