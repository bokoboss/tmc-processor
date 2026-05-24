from __future__ import annotations

from io import BytesIO
import json
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook

from tmc_processor.batch import (
    BATCH_EXCEL_TEMPLATE_EXPORT_MODE,
    BATCH_QC_COLUMNS,
    BATCH_SAFE_PNG_EXPORT_MODE,
    BATCH_SUMMARY_COLUMNS,
    BatchItem,
    analyze_batch_files,
    batch_change_invalidates,
    batch_file_metadata_defaults,
    batch_inputs_ready,
    batch_qc_frame,
    batch_selected_file_preview,
    batch_zip_contents_preview,
    batch_zip_generation_block_reason,
    BATCH_STALE_MESSAGE_TH,
    derive_survey_date_text_from_filename,
    generate_batch_zip_from_reviewed_peaks,
    process_batch_files,
    safe_output_stem,
    unique_safe_output_stems,
)
from tmc_processor.mapping_preset import load_mapping_preset


ROOT = Path(__file__).resolve().parents[1]
DEMO_DIR = ROOT / "samples" / "demo"
DAY1 = DEMO_DIR / "DEMO_TMC1_FourLeg.xlsx"
DAY2 = DEMO_DIR / "DEMO_TMC1_FourLeg_Day2.xlsx"
PRESET = DEMO_DIR / "DEMO_TMC1_FourLeg.mapping.json"


def _demo_items() -> list[BatchItem]:
    return [
        BatchItem(file_name=DAY1.name, workbook_bytes=DAY1.read_bytes()),
        BatchItem(file_name=DAY2.name, workbook_bytes=DAY2.read_bytes()),
    ]


def _preset() -> dict[str, object]:
    return load_mapping_preset(PRESET.read_bytes()).preset


def _setup() -> dict[str, object]:
    return {
        "project_name": "Synthetic Demo Project",
        "tmc_id": "DEMO-TMC1",
        "tmc_title": "Synthetic Four-Leg Demo Intersection",
        "survey_point": "Synthetic Four-Leg Demo Point",
        "survey_date_text": "2026-01-01",
        "survey_period": "07:00-19:00",
        "north_label": "Demo North",
        "south_label": "Demo South",
        "east_label": "Demo East",
        "west_label": "Demo West",
    }


def test_processes_two_demo_workbooks_with_one_mapping_preset() -> None:
    result = process_batch_files(
        _demo_items(),
        mapping_preset=_preset(),
        setup=_setup(),
        generated_at="2026-05-19T10:00:00Z",
    )

    assert [row.status for row in result.summary_rows] == ["success", "success"]
    assert all(row.AM_peak for row in result.summary_rows)
    assert all(row.PM_peak for row in result.summary_rows)
    assert all(row.total_vehicles > 0 for row in result.summary_rows)
    assert all(row.total_PCU > 0 for row in result.summary_rows)
    assert all(row.QC_errors >= 0 and row.QC_warnings >= 0 and row.QC_info >= 0 for row in result.summary_rows)


def test_batch_inputs_ready_requires_workbooks_and_mapping() -> None:
    assert batch_inputs_ready(uploaded_workbook_count=2, mapping_available=True, pce_factors_ready=True)
    assert not batch_inputs_ready(uploaded_workbook_count=0, mapping_available=True, pce_factors_ready=True)
    assert not batch_inputs_ready(uploaded_workbook_count=2, mapping_available=False, pce_factors_ready=True)
    assert not batch_inputs_ready(uploaded_workbook_count=2, mapping_available=True, pce_factors_ready=False)


def test_filename_date_extraction_and_safe_output_stem() -> None:
    assert derive_survey_date_text_from_filename("TMC_Main_2026-02-03.xlsx") == "2026-02-03"
    assert derive_survey_date_text_from_filename("TMC_Main_20260203.xlsx") == "2026-02-03"
    assert derive_survey_date_text_from_filename("TMC_Main_03-02-2026.xlsx") == "2026-02-03"
    assert derive_survey_date_text_from_filename("TMC_Main_no_date.xlsx") == ""
    assert safe_output_stem("../bad path/TMC Main:Day 1.xlsx") == "TMC_Main_Day_1"


def test_batch_file_metadata_defaults_do_not_include_raw_paths() -> None:
    defaults = batch_file_metadata_defaults([r"C:\raw\TMC Main 20260203.xlsx"])

    assert defaults == [
        {
            "file_name": "TMC Main 20260203.xlsx",
            "survey_date_text": "2026-02-03",
            "output_stem": "TMC_Main_20260203",
            "notes": "",
        }
    ]


def test_batch_input_changes_invalidate_after_analysis() -> None:
    assert batch_change_invalidates(("files-v1",), ("files-v2",), has_analysis=True)
    assert not batch_change_invalidates(("files-v1",), ("files-v2",), has_analysis=False)
    assert not batch_change_invalidates(("files-v1",), ("files-v1",), has_analysis=True)


def test_batch_zip_generation_is_blocked_when_batch_is_stale() -> None:
    assert (
        batch_zip_generation_block_reason(
            has_successful_files=True,
            peaks_ready=True,
            batch_stale=True,
        )
        == BATCH_STALE_MESSAGE_TH
    )
    assert not batch_zip_generation_block_reason(
        has_successful_files=True,
        peaks_ready=True,
        batch_stale=False,
    )


def test_unique_safe_output_stems_fallback_and_collision_handling() -> None:
    items = [
        BatchItem(file_name=r"C:\raw\Day 1.xlsx", workbook_bytes=b"", output_stem="../bad:stem"),
        BatchItem(file_name="Day 2.xlsx", workbook_bytes=b"", output_stem="../bad:stem"),
        BatchItem(file_name="Unsafe Name.xlsx", workbook_bytes=b"", output_stem="..."),
    ]

    assert unique_safe_output_stems(items) == ["bad_stem", "bad_stem_02", "Unsafe_Name"]


def test_batch_analysis_returns_suggested_and_default_confirmed_peaks() -> None:
    analysis = analyze_batch_files(
        _demo_items(),
        mapping_preset=_preset(),
        setup=_setup(),
        generated_at="2026-05-19T10:00:00Z",
    )

    assert [item.status for item in analysis.items] == ["success", "success"]
    for item in analysis.successful_items:
        assert item.suggested_AM_peak
        assert item.suggested_PM_peak
        assert item.confirmed_AM_peak == item.suggested_AM_peak
        assert item.confirmed_PM_peak == item.suggested_PM_peak
        assert item.suggested_AM_peak in item.hourly_period_options
        assert item.suggested_PM_peak in item.hourly_period_options
        assert not item.hourly_movement_pcu.empty
        assert "Total" in item.hourly_movement_pcu.columns


def test_selected_file_preview_helper_exposes_compact_values() -> None:
    analysis = analyze_batch_files(
        _demo_items(),
        mapping_preset=_preset(),
        setup=_setup(),
        generated_at="2026-05-19T10:00:00Z",
    )

    preview = batch_selected_file_preview(analysis.successful_items[0])

    assert preview["file_name"] == DAY1.name
    assert preview["output_stem"] == "DEMO_TMC1_FourLeg"
    assert preview["status"] == "success"
    assert preview["total_vehicles"] > 0
    assert preview["total_PCU"] > 0
    assert preview["suggested_AM_peak"]
    assert preview["confirmed_AM_peak"] == preview["suggested_AM_peak"]
    assert {"QC_errors", "QC_warnings", "QC_info"}.issubset(preview)


def test_batch_zip_contains_summary_and_one_folder_per_success_without_raw_inputs() -> None:
    result = process_batch_files(
        _demo_items(),
        mapping_preset=_preset(),
        setup=_setup(),
        generated_at="2026-05-19T10:00:00Z",
    )

    with ZipFile(BytesIO(result.package_bytes)) as archive:
        name_list = archive.namelist()
        names = set(name_list)

    assert "batch_summary.xlsx" in names
    assert name_list.count("batch_summary.xlsx") == 1
    success_folders = {row.folder_name for row in result.summary_rows if row.status == "success"}
    for row in result.summary_rows:
        if row.status != "success":
            continue
        folder = row.folder_name
        assert f"{folder}/{row.output_stem}_report.xlsx" in names
        assert f"{folder}/{row.output_stem}_export_summary.txt" in names
        assert f"{folder}/{row.output_stem}_session.tmcproj.json" in names
        assert f"{folder}/{row.output_stem}.mapping.json" in names
    assert len(success_folders) == 2
    assert DAY1.name not in names
    assert DAY2.name not in names
    assert all(not name.endswith(".xlsm") and name not in {DAY1.name, DAY2.name} for name in names)


def test_batch_summary_workbook_contains_summary_and_qc_sheets() -> None:
    result = process_batch_files(
        _demo_items(),
        mapping_preset=_preset(),
        setup=_setup(),
        generated_at="2026-05-19T10:00:00Z",
    )

    with ZipFile(BytesIO(result.package_bytes)) as archive:
        summary_bytes = archive.read("batch_summary.xlsx")

    workbook = load_workbook(BytesIO(summary_bytes), read_only=True, data_only=True)

    assert {"Batch_Summary", "Batch_QC"}.issubset(set(workbook.sheetnames))
    summary_headers = [cell.value for cell in next(workbook["Batch_Summary"].iter_rows(max_row=1))]
    qc_headers = [cell.value for cell in next(workbook["Batch_QC"].iter_rows(max_row=1))]
    assert summary_headers == BATCH_SUMMARY_COLUMNS
    assert qc_headers == BATCH_QC_COLUMNS


def test_batch_qc_sheet_collects_qc_rows_per_file() -> None:
    result = process_batch_files(
        _demo_items(),
        mapping_preset=_preset(),
        setup=_setup(),
        generated_at="2026-05-19T10:00:00Z",
    )

    with ZipFile(BytesIO(result.package_bytes)) as archive:
        workbook = load_workbook(BytesIO(archive.read("batch_summary.xlsx")), read_only=True, data_only=True)

    rows = list(workbook["Batch_QC"].iter_rows(values_only=True))
    records = [dict(zip(rows[0], row)) for row in rows[1:]]

    assert records
    assert {record["file_name"] for record in records} == {DAY1.name, DAY2.name}
    assert all(record["output_stem"] for record in records)
    assert all(record["severity"] in {"error", "warning", "info"} for record in records)


def test_failed_file_creates_batch_qc_failure_row() -> None:
    items = [
        BatchItem(file_name=DAY1.name, workbook_bytes=DAY1.read_bytes()),
        BatchItem(file_name="broken_input.xlsx", workbook_bytes=b"not an excel workbook"),
    ]
    result = process_batch_files(
        items,
        mapping_preset=_preset(),
        setup=_setup(),
        generated_at="2026-05-19T10:00:00Z",
    )

    qc = batch_qc_frame(result.qc_rows)
    failed = qc[qc["file_name"] == "broken_input.xlsx"]

    assert not failed.empty
    assert failed.iloc[0]["severity"] == "error"
    assert failed.iloc[0]["category"] == "batch_processing"
    assert failed.iloc[0]["check"] == "processing_failed"
    assert failed.iloc[0]["message"]


def test_batch_safe_png_export_records_export_mode_in_summaries() -> None:
    result = process_batch_files(
        _demo_items(),
        mapping_preset=_preset(),
        setup=_setup(),
        export_mode=BATCH_SAFE_PNG_EXPORT_MODE,
        generated_at="2026-05-19T10:00:00Z",
    )

    with ZipFile(BytesIO(result.package_bytes)) as archive:
        first = result.summary_rows[0]
        summary_text = archive.read(f"{first.folder_name}/{first.output_stem}_export_summary.txt").decode("utf-8")
        summary_bytes = archive.read("batch_summary.xlsx")

    workbook = load_workbook(BytesIO(summary_bytes), read_only=True, data_only=True)
    rows = list(workbook["Batch_Summary"].iter_rows(values_only=True))
    first_record = dict(zip(rows[0], rows[1]))

    assert first_record["export_mode_requested"] == BATCH_SAFE_PNG_EXPORT_MODE
    assert first_record["export_mode_used"] == BATCH_SAFE_PNG_EXPORT_MODE
    assert first_record["export_status"] == "success"
    assert "export_mode_requested: Safe PNG Export Mode" in summary_text
    assert "export_mode_used: Safe PNG Export Mode" in summary_text
    assert "export_status: success" in summary_text


def test_batch_excel_template_mode_records_requested_and_used_mode() -> None:
    result = process_batch_files(
        _demo_items(),
        mapping_preset=_preset(),
        setup=_setup(),
        export_mode=BATCH_EXCEL_TEMPLATE_EXPORT_MODE,
        use_excel_com_native_charts=False,
        generated_at="2026-05-19T10:00:00Z",
    )

    with ZipFile(BytesIO(result.package_bytes)) as archive:
        first = result.summary_rows[0]
        summary_text = archive.read(f"{first.folder_name}/{first.output_stem}_export_summary.txt").decode("utf-8")
        summary_bytes = archive.read("batch_summary.xlsx")

    workbook = load_workbook(BytesIO(summary_bytes), read_only=True, data_only=True)
    rows = list(workbook["Batch_Summary"].iter_rows(values_only=True))
    first_record = dict(zip(rows[0], rows[1]))

    assert first_record["export_mode_requested"] == BATCH_EXCEL_TEMPLATE_EXPORT_MODE
    assert first_record["export_mode_used"] == BATCH_SAFE_PNG_EXPORT_MODE
    assert first_record["export_status"] == "success"
    assert first_record["notes"]
    assert "export_mode_requested: Excel Template Mode" in summary_text
    assert "export_mode_used: Safe PNG Export Mode" in summary_text


def test_batch_zip_contents_preview_lists_expected_artifacts() -> None:
    result = process_batch_files(
        _demo_items(),
        mapping_preset=_preset(),
        setup=_setup(),
        generated_at="2026-05-19T10:00:00Z",
    )

    preview = batch_zip_contents_preview(result.summary_rows)

    assert preview[0] == "batch_summary.xlsx"
    assert "file_01_DEMO_TMC1_FourLeg/DEMO_TMC1_FourLeg_report.xlsx" in preview
    assert "file_01_DEMO_TMC1_FourLeg/DEMO_TMC1_FourLeg_export_summary.txt" in preview
    assert "file_01_DEMO_TMC1_FourLeg/DEMO_TMC1_FourLeg_session.tmcproj.json" in preview
    assert "file_01_DEMO_TMC1_FourLeg/DEMO_TMC1_FourLeg.mapping.json" in preview
    assert "file_01_DEMO_TMC1_FourLeg/charts/" in preview


def test_custom_confirmed_peak_overrides_suggested_in_final_zip() -> None:
    analysis = analyze_batch_files(
        _demo_items(),
        mapping_preset=_preset(),
        setup=_setup(),
        mapping_preset_name="Demo preset",
        generated_at="2026-05-19T10:00:00Z",
    )
    first = analysis.successful_items[0]
    custom_am = next(option for option in first.hourly_period_options if option != first.suggested_AM_peak)
    first.confirmed_AM_peak = custom_am

    result = generate_batch_zip_from_reviewed_peaks(
        analysis,
        setup=_setup(),
    )

    with ZipFile(BytesIO(result.package_bytes)) as archive:
        first_row = result.summary_rows[0]
        summary_text = archive.read(f"{first_row.folder_name}/{first_row.output_stem}_export_summary.txt").decode("utf-8")
        summary_bytes = archive.read("batch_summary.xlsx")

    workbook = load_workbook(BytesIO(summary_bytes), read_only=True, data_only=True)
    rows = list(workbook["Batch_Summary"].iter_rows(values_only=True))
    records = [dict(zip(rows[0], row)) for row in rows[1:]]
    first_record = records[0]

    assert first_record["suggested_AM_peak"] == first.suggested_AM_peak
    assert first_record["confirmed_AM_peak"] == custom_am
    assert first_record["AM_peak"] == custom_am
    assert f"AM peak period: {custom_am}" in summary_text
    assert "Peak selection source: user_confirmed_batch" in summary_text


def test_edited_batch_metadata_is_used_in_exports_and_summary() -> None:
    items = [
        BatchItem(
            file_name=DAY1.name,
            workbook_bytes=DAY1.read_bytes(),
            survey_date_text="2026-02-03",
            output_stem="demo-day-one",
        )
    ]
    result = process_batch_files(
        items,
        mapping_preset=_preset(),
        setup=_setup(),
        generated_at="2026-05-19T10:00:00Z",
    )
    row = result.summary_rows[0]

    with ZipFile(BytesIO(result.package_bytes)) as archive:
        names = set(archive.namelist())
        summary_text = archive.read("file_01_demo-day-one/demo-day-one_export_summary.txt").decode("utf-8")
        session = json.loads(archive.read("file_01_demo-day-one/demo-day-one_session.tmcproj.json").decode("utf-8"))
        summary_bytes = archive.read("batch_summary.xlsx")

    workbook = load_workbook(BytesIO(summary_bytes), read_only=True, data_only=True)
    rows = list(workbook["Batch_Summary"].iter_rows(values_only=True))
    first_record = dict(zip(rows[0], rows[1]))

    assert row.folder_name == "file_01_demo-day-one"
    assert row.export_file == "file_01_demo-day-one/demo-day-one_report.xlsx"
    assert row.generated_report_filename == "demo-day-one_report.xlsx"
    assert "file_01_demo-day-one/demo-day-one_report.xlsx" in names
    assert "survey_date_text: 2026-02-03" in summary_text
    assert "output_stem: demo-day-one" in summary_text
    assert session["metadata"]["survey_date_text"] == "2026-02-03"
    assert first_record["survey_date_text"] == "2026-02-03"
    assert first_record["output_stem"] == "demo-day-one"


def test_failed_file_does_not_stop_batch_and_summary_contains_success_and_failure_rows() -> None:
    items = [
        BatchItem(file_name=DAY1.name, workbook_bytes=DAY1.read_bytes()),
        BatchItem(file_name="broken_input.xlsx", workbook_bytes=b"not an excel workbook"),
    ]
    result = process_batch_files(
        items,
        mapping_preset=_preset(),
        setup=_setup(),
        generated_at="2026-05-19T10:00:00Z",
    )

    assert [row.status for row in result.summary_rows] == ["success", "failed"]
    assert result.summary_rows[1].notes

    with ZipFile(BytesIO(result.package_bytes)) as archive:
        names = set(archive.namelist())
        summary_bytes = archive.read("batch_summary.xlsx")

    assert "file_01_DEMO_TMC1_FourLeg/DEMO_TMC1_FourLeg_report.xlsx" in names
    assert "file_02_broken_input/broken_input_report.xlsx" not in names

    workbook = load_workbook(BytesIO(summary_bytes), read_only=True, data_only=True)
    rows = list(workbook["Batch_Summary"].iter_rows(values_only=True))
    headers = list(rows[0])
    records = [dict(zip(headers, row)) for row in rows[1:]]

    assert [record["status"] for record in records] == ["success", "failed"]
    assert headers == BATCH_SUMMARY_COLUMNS
    assert {"QC_errors", "QC_warnings", "QC_info"}.issubset(headers)
    assert {"export_mode_requested", "export_mode_used", "export_status", "export_error"}.issubset(headers)
    assert {"output_stem", "generated_report_filename"}.issubset(headers)
    assert records[0]["export_status"] == "success"
    assert records[1]["export_status"] == "failed"
    assert records[1]["export_error"]
    assert records[1]["notes"]


def test_failed_analysis_item_does_not_block_reviewed_batch_zip() -> None:
    items = [
        BatchItem(file_name=DAY1.name, workbook_bytes=DAY1.read_bytes()),
        BatchItem(file_name="broken_input.xlsx", workbook_bytes=b"not an excel workbook"),
    ]
    analysis = analyze_batch_files(
        items,
        mapping_preset=_preset(),
        setup=_setup(),
        generated_at="2026-05-19T10:00:00Z",
    )

    assert [item.status for item in analysis.items] == ["success", "failed"]

    result = generate_batch_zip_from_reviewed_peaks(analysis, setup=_setup())

    assert [row.status for row in result.summary_rows] == ["success", "failed"]
    with ZipFile(BytesIO(result.package_bytes)) as archive:
        names = set(archive.namelist())
        summary_bytes = archive.read("batch_summary.xlsx")

    assert "file_01_DEMO_TMC1_FourLeg/DEMO_TMC1_FourLeg_report.xlsx" in names
    assert "file_02_broken_input/broken_input_report.xlsx" not in names
    workbook = load_workbook(BytesIO(summary_bytes), read_only=True, data_only=True)
    rows = list(workbook["Batch_Summary"].iter_rows(values_only=True))
    records = [dict(zip(rows[0], row)) for row in rows[1:]]
    assert records[1]["status"] == "failed"
    assert records[1]["notes"]


def test_batch_summary_workbook_has_metadata_sheet() -> None:
    result = process_batch_files(
        _demo_items(),
        mapping_preset=_preset(),
        setup=_setup(),
        mapping_preset_name="Demo preset",
        generated_at="2026-05-19T10:00:00Z",
    )

    with ZipFile(BytesIO(result.package_bytes)) as archive:
        summary_bytes = archive.read("batch_summary.xlsx")

    workbook = load_workbook(BytesIO(summary_bytes), read_only=True, data_only=True)
    metadata_rows = list(workbook["metadata"].iter_rows(values_only=True))
    metadata = {row[0]: row[1] for row in metadata_rows[1:]}

    assert metadata["app_version"]
    assert metadata["template_version"] == "four_leg_v1"
    assert metadata["generated_at"] == "2026-05-19T10:00:00Z"
    assert metadata["mapping_preset_name"] == "Demo preset"


def test_batch_summary_workbook_has_confirmed_peak_columns() -> None:
    result = process_batch_files(
        _demo_items(),
        mapping_preset=_preset(),
        setup=_setup(),
        generated_at="2026-05-19T10:00:00Z",
    )

    with ZipFile(BytesIO(result.package_bytes)) as archive:
        summary_bytes = archive.read("batch_summary.xlsx")

    workbook = load_workbook(BytesIO(summary_bytes), read_only=True, data_only=True)
    headers = [cell.value for cell in next(workbook["Batch_Summary"].iter_rows(max_row=1))]

    assert "confirmed_AM_peak" in headers
    assert "confirmed_PM_peak" in headers


def test_duplicate_output_stems_use_predictable_safe_zip_paths() -> None:
    items = [
        BatchItem(file_name=DAY1.name, workbook_bytes=DAY1.read_bytes(), output_stem="same:name"),
        BatchItem(file_name=DAY2.name, workbook_bytes=DAY2.read_bytes(), output_stem="same:name"),
    ]

    result = process_batch_files(
        items,
        mapping_preset=_preset(),
        setup=_setup(),
        generated_at="2026-05-19T10:00:00Z",
    )

    with ZipFile(BytesIO(result.package_bytes)) as archive:
        names = set(archive.namelist())

    assert [row.folder_name for row in result.summary_rows] == ["file_01_same_name", "file_02_same_name_02"]
    assert "file_01_same_name/same_name_report.xlsx" in names
    assert "file_02_same_name_02/same_name_02_report.xlsx" in names
    assert DAY1.name not in names
    assert DAY2.name not in names
