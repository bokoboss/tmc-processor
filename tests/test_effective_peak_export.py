from __future__ import annotations

from datetime import time
from io import BytesIO

from openpyxl import load_workbook
import pandas as pd

import app
from tmc_processor.exporter import export_workbook
from tmc_processor.peaks import (
    PEAK_SELECTION_AUTO,
    PEAK_SELECTION_TEMPLATE_DEFAULT,
    PEAK_SELECTION_USER_CONFIRMED,
    resolve_effective_peak_periods,
)
from tmc_processor.pipeline import process_tmc


def _mapping() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "raw_sheet": "S1",
                "raw_direction": "North",
                "movement_code": "NS",
                "source_stream": "mainline",
                "raw_movement_label": "North",
                "from_leg": "N",
                "to_leg": "S",
                "turn_type": "through",
                "facility_type": "at_grade",
                "include_in_peak": True,
                "include_in_report": True,
                "aggregation_method": "sum",
            }
        ]
    )


def _raw_sheets() -> dict[str, pd.DataFrame]:
    rows = []
    for hour, counts in ((7, [25, 25, 25, 25]), (17, [249, 250, 250, 250])):
        for minute, count in zip((0, 15, 30, 45), counts):
            rows.append(
                {
                    "raw_sheet": "S1",
                    "raw_direction": "North",
                    "time_start": time(hour, minute),
                    "time_end": time(hour + 1, 0) if minute == 45 else time(hour, minute + 15),
                    "vehicle_class": "PC<7",
                    "count": count,
                }
            )
    return {"S1": pd.DataFrame(rows)}


def _processed():
    return process_tmc(
        raw_sheets=_raw_sheets(),
        mapping=_mapping(),
        setup={"project_name": "Effective peak test"},
        detected_sheets=["S1"],
        peak_windows={"AM": ("07:00", "08:00"), "PM": ("17:00", "18:00")},
        generate_workbook=False,
    )


def _records(workbook, sheet_name: str) -> dict[object, object]:
    return {row[0]: row[1] for row in workbook[sheet_name].iter_rows(min_row=2, values_only=True) if row[0]}


def _export(setup: dict[str, object]):
    result = _processed()
    return load_workbook(
        BytesIO(
            export_workbook(
                setup,
                _mapping(),
                result.normalized,
                result.qc,
                result.hourly,
                result.movement,
                result.vehicle,
                result.peaks,
                include_charts=False,
                include_diagram=False,
            )
        ),
        data_only=False,
    )


def _export_template_layout(setup: dict[str, object]):
    result = _processed()
    return load_workbook(
        BytesIO(
            export_workbook(
                setup,
                _mapping(),
                result.normalized,
                result.qc,
                result.hourly,
                result.movement,
                result.vehicle,
                result.peaks,
                include_charts=False,
                include_diagram=False,
                use_template_report_layout=True,
            )
        ),
        data_only=False,
    )


def test_effective_peak_resolution_manual_override() -> None:
    periods, source = resolve_effective_peak_periods(
        selected_peak_periods={"AM": ("17:00", "18:00")},
        recommended_peak_periods={"AM": ("07:00", "08:00")},
        template_default_peak_periods={"AM": ("07:00", "08:00")},
    )

    assert periods["AM"] == ("17:00", "18:00")
    assert source == PEAK_SELECTION_USER_CONFIRMED


def test_effective_peak_resolution_recommendation_overrides_template_default() -> None:
    periods, source = resolve_effective_peak_periods(
        selected_peak_periods=None,
        recommended_peak_periods={"PM": ("17:00", "18:00")},
        template_default_peak_periods={"PM": ("07:00", "08:00")},
    )

    assert periods["PM"] == ("17:00", "18:00")
    assert source == PEAK_SELECTION_AUTO


def test_effective_peak_resolution_recommendation_fallback() -> None:
    periods, source = resolve_effective_peak_periods(
        selected_peak_periods=None,
        recommended_peak_periods={"AM": ("07:00", "08:00")},
        template_default_peak_periods={"AM": ("17:00", "18:00")},
    )

    assert periods["AM"] == ("07:00", "08:00")
    assert source == PEAK_SELECTION_AUTO


def test_effective_peak_resolution_template_fallback_only() -> None:
    periods, source = resolve_effective_peak_periods(
        selected_peak_periods=None,
        recommended_peak_periods=None,
        template_default_peak_periods={"AM": ("07:00", "08:00")},
    )

    assert periods["AM"] == ("07:00", "08:00")
    assert source == PEAK_SELECTION_TEMPLATE_DEFAULT


def test_export_payload_and_summary_use_selected_peak() -> None:
    workbook = _export(
        {
            "peak_selection_source": PEAK_SELECTION_USER_CONFIRMED,
            "am_peak_start": "17:00",
            "am_peak_end": "18:00",
            "pm_peak_start": "17:00",
            "pm_peak_end": "18:00",
        }
    )

    metadata = _records(workbook, "Export_Metadata")
    peak = workbook["Peak_PHF"]
    assert metadata["effective_am_peak"] == "17:00-18:00"
    assert metadata["effective_peak_source"] == PEAK_SELECTION_USER_CONFIRMED
    assert peak["D2"].value == "17:00"
    assert peak["F2"].value == 999
    assert "AM 17:00-18:00" in workbook["TMC_Report"]["A3"].value


def test_export_payload_and_summary_use_recommended_peak_over_template_default() -> None:
    workbook = _export(
        {
            "peak_selection_source": PEAK_SELECTION_AUTO,
            "am_peak_start": "17:00",
            "am_peak_end": "18:00",
            "pm_peak_start": "07:00",
            "pm_peak_end": "08:00",
        }
    )

    metadata = _records(workbook, "Export_Metadata")
    peak = workbook["Peak_PHF"]
    assert metadata["effective_am_peak"] == "07:00-08:00"
    assert metadata["effective_pm_peak"] == "17:00-18:00"
    assert metadata["effective_peak_source"] == PEAK_SELECTION_AUTO
    assert peak["F2"].value == 100
    assert peak["J2"].value == 999
    assert "PM 17:00-18:00" in workbook["TMC_Report"]["A3"].value


def test_formula_linked_diagram_cells_use_effective_peak_data() -> None:
    workbook = _export(
        {
            "peak_selection_source": PEAK_SELECTION_USER_CONFIRMED,
            "am_peak_start": "17:00",
            "am_peak_end": "18:00",
            "pm_peak_start": "17:00",
            "pm_peak_end": "18:00",
        }
    )

    assert workbook["Diagram_Data"]["D2"].value.startswith("=IFERROR(INDEX(")
    assert "'Peak_PHF'!$D$2" in workbook["Diagram_Data"]["D2"].value
    assert workbook["Peak_PHF"]["F2"].value == 999


def test_template_layout_summary_label_uses_effective_peak() -> None:
    workbook = _export_template_layout(
        {
            "peak_selection_source": PEAK_SELECTION_USER_CONFIRMED,
            "am_peak_start": "17:00",
            "am_peak_end": "18:00",
        }
    )

    assert "AM 17:00-18:00" in workbook["TMC_Report"]["A3"].value
    assert workbook["Peak_PHF"]["F2"].value == 999


def test_changing_selected_peak_changes_export() -> None:
    am_workbook = _export(
        {
            "peak_selection_source": PEAK_SELECTION_USER_CONFIRMED,
            "am_peak_start": "07:00",
            "am_peak_end": "08:00",
        }
    )
    pm_workbook = _export(
        {
            "peak_selection_source": PEAK_SELECTION_USER_CONFIRMED,
            "am_peak_start": "17:00",
            "am_peak_end": "18:00",
        }
    )

    assert am_workbook["Peak_PHF"]["F2"].value == 100
    assert pm_workbook["Peak_PHF"]["F2"].value == 999
    assert am_workbook["TMC_Report"]["A3"].value != pm_workbook["TMC_Report"]["A3"].value


def test_streamlit_effective_peak_state_prefers_selected_over_recommended() -> None:
    app.st.session_state.clear()
    app.st.session_state["tmc_processed"] = {"result": _processed()}
    app.st.session_state["am_peak_period_select"] = "17:00-18:00"
    app.st.session_state["tmc_confirmed_am_peak_start"] = "17:00"
    app.st.session_state["tmc_confirmed_am_peak_end"] = "18:00"

    state = app._single_effective_peak_state()

    assert state["source"] == PEAK_SELECTION_USER_CONFIRMED
    assert state["values"]["am_peak_start"] == "17:00"


def test_streamlit_effective_peak_state_keeps_manual_source_after_leaving_review_tab() -> None:
    app.st.session_state.clear()
    app.st.session_state["tmc_processed"] = {"result": _processed()}
    app.st.session_state["tmc_confirmed_am_peak_start"] = "17:00"
    app.st.session_state["tmc_confirmed_am_peak_end"] = "18:00"
    app.st.session_state["tmc_confirmed_pm_peak_start"] = "17:00"
    app.st.session_state["tmc_confirmed_pm_peak_end"] = "18:00"

    state = app._single_effective_peak_state()

    assert state["source"] == PEAK_SELECTION_USER_CONFIRMED
    assert state["values"]["am_peak_start"] == "17:00"
