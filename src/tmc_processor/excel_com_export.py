"""Optional Microsoft Excel COM export for native template charts.

This module must remain import-safe on machines without pywin32.  The normal
openpyxl export path imports no COM modules and remains the default.
"""

from __future__ import annotations

import platform
import shutil
from dataclasses import dataclass, field
from datetime import date, datetime, time, timezone
import importlib
import os
from pathlib import Path
import site
import sys
import sysconfig
from typing import Any
import warnings

from openpyxl.utils.cell import get_column_letter, range_boundaries
import pandas as pd

from .diagram import MOVEMENT_CODES
from .report_template import (
    ReportTemplateUnavailable,
    _column_for_key,
    _excel_value,
    _mapped_label_value,
    _metadata_values,
    _normalise_time_label,
)
from .template_audit import audit_template_formulas


class ExcelComUnavailable(RuntimeError):
    """Raised when Microsoft Excel COM automation cannot be used."""


@dataclass(frozen=True)
class ExcelComStatus:
    available: bool
    reason: str
    version: str = ""
    detail: str = ""
    probed_at: str = ""
    pywintypes_file: str = ""
    pywintypes_error: str = ""
    pywin32_system32: str = ""
    sys_path_matches: tuple[str, ...] = ()


@dataclass
class ExcelComExportDiagnostics:
    formula_cells_preserved: list[str] = field(default_factory=list)
    formula_cells_overwritten: list[str] = field(default_factory=list)
    formula_cells_skipped: list[str] = field(default_factory=list)
    formulas_with_external_links: list[str] = field(default_factory=list)
    formulas_with_ref_errors: list[str] = field(default_factory=list)
    formulas_with_missing_sheet_references: list[str] = field(default_factory=list)
    mapped_formula_cells: list[str] = field(default_factory=list)
    calculation_warnings: list[str] = field(default_factory=list)

    def summary(self) -> dict[str, list[str]]:
        return {
            "formula_cells_preserved": list(self.formula_cells_preserved),
            "formula_cells_overwritten": list(self.formula_cells_overwritten),
            "formula_cells_skipped": list(self.formula_cells_skipped),
            "formulas_with_external_links": list(self.formulas_with_external_links),
            "formulas_with_ref_errors": list(self.formulas_with_ref_errors),
            "formulas_with_missing_sheet_references": list(self.formulas_with_missing_sheet_references),
            "mapped_formula_cells": list(self.mapped_formula_cells),
            "calculation_warnings": list(self.calculation_warnings),
        }


_LAST_EXPORT_DIAGNOSTICS = ExcelComExportDiagnostics()


def get_last_export_diagnostics() -> ExcelComExportDiagnostics:
    return _LAST_EXPORT_DIAGNOSTICS


def _import_com_modules():
    import win32com.client  # type: ignore[import-not-found]
    import pythoncom  # type: ignore[import-not-found]

    return pythoncom, win32com.client


def _probe_timestamp() -> str:
    return datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")


def _unique_existing_dirs(paths: list[Path]) -> list[Path]:
    existing: list[Path] = []
    seen: set[str] = set()
    for path in paths:
        try:
            resolved = path.resolve()
        except OSError:
            resolved = path
        key = str(resolved).casefold()
        if key not in seen and path.exists() and path.is_dir():
            seen.add(key)
            existing.append(path)
    return existing


def _site_package_paths() -> list[Path]:
    candidates = []
    try:
        candidates.extend(site.getsitepackages())
    except Exception:
        pass
    try:
        candidates.append(site.getusersitepackages())
    except Exception:
        pass
    try:
        paths = sysconfig.get_paths()
        candidates.extend([paths.get("purelib"), paths.get("platlib")])
    except Exception:
        pass
    candidates.extend(sys.path)

    return _unique_existing_dirs([Path(base) for base in candidates if base])


def _pywin32_candidate_dirs() -> list[Path]:
    paths: list[Path] = []
    for root in _site_package_paths():
        paths.extend([root / "pywin32_system32", root / "win32", root / "win32" / "lib"])
    return _unique_existing_dirs(paths)


def _sys_path_pywin32_matches(limit: int = 8) -> tuple[str, ...]:
    matches = [path for path in sys.path if "win32" in path.lower() or "pywin32" in path.lower()]
    return tuple(matches[:limit])


def _pywin32_system32_path() -> str:
    for path in _pywin32_candidate_dirs():
        if path.name.lower() == "pywin32_system32":
            return str(path)
    return ""


def _bootstrap_pywin32_paths() -> dict[str, Any]:
    """Register pywin32 directories for hosts that did not process pywin32 .pth files."""

    candidate_dirs = _pywin32_candidate_dirs()
    pywin32_system32 = ""
    for path in candidate_dirs:
        path_text = str(path)
        if path_text not in sys.path:
            sys.path.append(path_text)
        if path.name.lower() == "pywin32_system32":
            pywin32_system32 = path_text
            current_path = os.environ.get("PATH", "")
            path_entries = current_path.split(os.pathsep) if current_path else []
            if path_text not in path_entries:
                os.environ["PATH"] = path_text + (os.pathsep + current_path if current_path else "")
        if path.name.lower() == "pywin32_system32" and hasattr(os, "add_dll_directory"):
            try:
                os.add_dll_directory(path_text)
            except (FileNotFoundError, OSError):
                pass
    importlib.invalidate_caches()
    return {
        "pywin32_system32": pywin32_system32,
        "sys_path_matches": _sys_path_pywin32_matches(),
    }


def _import_com_modules_with_repair():
    _bootstrap_pywin32_paths()
    return _import_com_modules()


def _import_pywintypes_diagnostic() -> tuple[str, str]:
    try:
        pywintypes = importlib.import_module("pywintypes")
    except Exception as exc:
        return "", str(exc)
    return str(getattr(pywintypes, "__file__", "") or ""), ""


def _status(
    available: bool,
    reason: str,
    *,
    version: str = "",
    detail: str = "",
    probed_at: str,
    pywintypes_file: str = "",
    pywintypes_error: str = "",
    pywin32_system32: str = "",
    sys_path_matches: tuple[str, ...] | None = None,
) -> ExcelComStatus:
    return ExcelComStatus(
        available,
        reason,
        version=version,
        detail=detail,
        probed_at=probed_at or _probe_timestamp(),
        pywintypes_file=pywintypes_file,
        pywintypes_error=pywintypes_error,
        pywin32_system32=pywin32_system32 or _pywin32_system32_path(),
        sys_path_matches=sys_path_matches if sys_path_matches is not None else _sys_path_pywin32_matches(),
    )


def probe_excel_com() -> ExcelComStatus:
    """Run a real Excel COM smoke check and return a diagnostic status."""

    probed_at = _probe_timestamp()
    if platform.system() != "Windows":
        return _status(False, "NOT_WINDOWS", detail="Excel COM automation is Windows-only.", probed_at=probed_at)

    pythoncom = None
    coinitialized = False
    excel = None
    bootstrap = _bootstrap_pywin32_paths()
    pywintypes_file, pywintypes_error = _import_pywintypes_diagnostic()
    try:
        pythoncom, win32_client = _import_com_modules()
    except ImportError as exc:
        detail = str(exc)
        if pywintypes_error:
            detail = f"{detail}; pywintypes import warning: {pywintypes_error}"
        return _status(
            False,
            "PYWIN32_MISSING",
            detail=detail,
            probed_at=probed_at,
            pywintypes_file=pywintypes_file,
            pywintypes_error=pywintypes_error,
            pywin32_system32=str(bootstrap.get("pywin32_system32", "")),
            sys_path_matches=tuple(bootstrap.get("sys_path_matches", ()) or _sys_path_pywin32_matches()),
        )

    try:
        pythoncom.CoInitialize()
        coinitialized = True
        excel = win32_client.Dispatch("Excel.Application")
        version = str(excel.Version)
        detail = f"pywintypes import warning: {pywintypes_error}" if pywintypes_error else ""
        return _status(
            True,
            "COM_AVAILABLE",
            version=version,
            detail=detail,
            probed_at=probed_at,
            pywintypes_file=pywintypes_file,
            pywintypes_error=pywintypes_error,
            pywin32_system32=str(bootstrap.get("pywin32_system32", "")),
            sys_path_matches=tuple(bootstrap.get("sys_path_matches", ()) or _sys_path_pywin32_matches()),
        )
    except Exception as exc:
        return _status(
            False,
            "COM_DISPATCH_FAILED",
            detail=str(exc),
            probed_at=probed_at,
            pywintypes_file=pywintypes_file,
            pywintypes_error=pywintypes_error,
            pywin32_system32=str(bootstrap.get("pywin32_system32", "")),
            sys_path_matches=tuple(bootstrap.get("sys_path_matches", ()) or _sys_path_pywin32_matches()),
        )
    finally:
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        if pythoncom is not None and coinitialized:
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass


def is_excel_com_available() -> bool:
    """Return True when a real Excel COM dispatch succeeds."""

    return probe_excel_com().available


def require_excel_com() -> ExcelComStatus:
    status = probe_excel_com()
    if not status.available:
        detail = f": {status.detail}" if status.detail else ""
        raise ExcelComUnavailable(f"{status.reason}{detail}")
    return status


def _coerce_excel_value(value: Any) -> Any:
    value = _excel_value(value)
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if isinstance(value, datetime):
        return value
    if isinstance(value, date) and not isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if value is pd.NaT:
        return ""
    return value


def _strip_sheet_reference(reference: str) -> str:
    text = str(reference).strip()
    if "!" in text:
        text = text.rsplit("!", 1)[1]
    return text.replace("$", "")


def _range_cells(cell_range: str) -> list[str]:
    cleaned = _strip_sheet_reference(cell_range)
    min_col, min_row, max_col, max_row = range_boundaries(cleaned)
    return [
        f"{get_column_letter(column)}{row}"
        for row in range(min_row, max_row + 1)
        for column in range(min_col, max_col + 1)
    ]


def _cell_in_ranges(cell_ref: str, ranges: list[str]) -> bool:
    cell = _strip_sheet_reference(cell_ref)
    cell_col, cell_row, _, _ = range_boundaries(cell)
    for cell_range in ranges:
        try:
            cleaned_range = _strip_sheet_reference(cell_range)
            min_col, min_row, max_col, max_row = range_boundaries(cleaned_range)
        except ValueError:
            continue
        if min_col <= cell_col <= max_col and min_row <= cell_row <= max_row:
            return True
    return False


def _cell_has_formula(cell) -> bool:
    try:
        return bool(cell.HasFormula)
    except Exception:
        pass
    try:
        formula = cell.Formula
    except Exception:
        return False
    return isinstance(formula, str) and formula.startswith("=")


def _cell_formula_text(cell) -> str:
    try:
        formula = cell.Formula
    except Exception:
        return ""
    return str(formula or "")


class _FormulaWriteGuard:
    def __init__(self, template_map: dict[str, Any], diagnostics: ExcelComExportDiagnostics) -> None:
        self.diagnostics = diagnostics
        self.writable_ranges = list(template_map.get("writable_ranges") or [])
        self.protected_formula_ranges = list(template_map.get("protected_formula_ranges") or [])
        self.formula_overwrite_ranges = list(template_map.get("formula_overwrite_ranges") or [])
        self.enforce_writable_ranges = bool(self.writable_ranges)

    def can_write(self, cell_ref: str, cell, intent: str) -> bool:
        address = _strip_sheet_reference(cell_ref)
        has_formula = _cell_has_formula(cell)
        in_writable_range = not self.enforce_writable_ranges or _cell_in_ranges(address, self.writable_ranges)
        if has_formula:
            formula = _cell_formula_text(cell)
            can_overwrite_formula = _cell_in_ranges(address, self.formula_overwrite_ranges)
            protected = not self.protected_formula_ranges or _cell_in_ranges(address, self.protected_formula_ranges)
            if can_overwrite_formula:
                self.diagnostics.formula_cells_overwritten.append(f"{address}: {intent}: {formula}")
                return True
            if protected or not in_writable_range:
                message = f"{address}: {intent}: {formula}"
                self.diagnostics.formula_cells_skipped.append(message)
                warnings.warn(f"Skipped protected template formula cell {message}", RuntimeWarning, stacklevel=3)
                return False
        if not in_writable_range:
            return False
        return True


def _set_cell(
    worksheet,
    cell_ref: str | None,
    value: Any,
    guard: _FormulaWriteGuard | None = None,
    intent: str = "mapped cell",
    force: bool = False,
) -> None:
    if not cell_ref:
        return
    cell = worksheet.Range(cell_ref)
    if guard is not None and force:
        existing_formula = _cell_formula_text(cell)
        if existing_formula:
            guard.diagnostics.formula_cells_overwritten.append(f"{_strip_sheet_reference(cell_ref)}: {intent}: {existing_formula}")
    elif guard is not None and not guard.can_write(cell_ref, cell, intent):
        return
    value = _coerce_excel_value(value)
    if isinstance(value, str) and value.startswith("="):
        cell.Formula = value
    else:
        cell.Value = value


def _write_effective_peak_label(worksheet, metadata: dict[str, Any]) -> None:
    def label(period: str) -> str:
        prefix = period.lower()
        start = _coerce_excel_value(metadata.get(f"{prefix}_peak_start"))
        end = _coerce_excel_value(metadata.get(f"{prefix}_peak_end"))
        return f"{start}-{end}" if start and end else "-"

    source = str(metadata.get("peak_selection_source") or "").strip() or "-"
    _set_cell(worksheet, "A3", f"Peak used for export: AM {label('AM')}; PM {label('PM')} ({source})")


def _column_number(column: str) -> int:
    number = 0
    for char in column.upper():
        if not ("A" <= char <= "Z"):
            continue
        number = number * 26 + (ord(char) - ord("A") + 1)
    return number


def _split_cell(cell_ref: str) -> tuple[int, int]:
    letters = "".join(char for char in cell_ref if char.isalpha())
    digits = "".join(char for char in cell_ref if char.isdigit())
    if not letters or not digits:
        raise ValueError(f"Invalid cell reference: {cell_ref}")
    return int(digits), _column_number(letters)


def _range_bounds(cell_range: str) -> tuple[int, int, int, int]:
    start, end = cell_range.split(":", 1) if ":" in cell_range else (cell_range, cell_range)
    start_row, start_col = _split_cell(start)
    end_row, end_col = _split_cell(end)
    return start_row, start_col, end_row, end_col


def _write_range(worksheet, cell_range: str | None, values: list[Any], guard: _FormulaWriteGuard | None = None, intent: str = "mapped range") -> None:
    if not cell_range:
        return
    start_row, start_col, end_row, end_col = _range_bounds(cell_range)
    rows = end_row - start_row + 1
    columns = end_col - start_col + 1
    padded = [_coerce_excel_value(values[index]) if index < len(values) else "" for index in range(rows * columns)]
    for offset, value in enumerate(padded):
        row = start_row + (offset // columns)
        column = start_col + (offset % columns)
        _set_cell(worksheet, f"{get_column_letter(column)}{row}", value, guard, intent)


def _write_metadata(worksheet, template_map: dict[str, Any], metadata: dict[str, Any], guard: _FormulaWriteGuard | None = None) -> None:
    values = _metadata_values(metadata)
    for key, info in template_map.get("metadata_cells", {}).items():
        _set_cell(worksheet, info.get("value_cell") or info.get("cell"), values.get(key, ""), guard, f"metadata.{key}")

    movement_info = template_map.get("movement_diagram_cells", {})
    _set_cell(worksheet, movement_info.get("diagram_title", {}).get("cell"), values.get("survey_point", ""), guard, "movement.diagram_title")
    _set_cell(worksheet, movement_info.get("diagram_date", {}).get("cell"), values.get("survey_date", ""), guard, "movement.diagram_date")
    for direction, info in movement_info.get("direction_labels", {}).items():
        _set_cell(
            worksheet,
            info.get("cell"),
            _mapped_label_value(metadata, direction),
            guard,
            f"movement.direction_labels.{direction}",
        )
    for label_key, info in movement_info.get("road_labels", {}).items():
        _set_cell(
            worksheet,
            info.get("cell"),
            _mapped_label_value(metadata, label_key),
            guard,
            f"movement.road_labels.{label_key}",
        )
    _set_cell(worksheet, movement_info.get("caption", {}).get("cell"), metadata.get("caption_text") or "", guard, "movement.caption")


def _diagram_formula(movement_code: str, value_key: str) -> str:
    column = {"total_12_hour": "B", "pm_peak": "C", "am_peak": "D"}[value_key]
    return f'=IFERROR(INDEX(Diagram_Data!${column}:${column},MATCH("{movement_code}",Diagram_Data!$A:$A,0)),0)'


def _movement_formula_targets(template_map: dict[str, Any]):
    approach_tables = template_map.get("movement_diagram_cells", {}).get("approach_tables", {})
    value_sets = [
        ("total_12_hour_cells", "total_12_hour"),
        ("pm_peak_hour_cells", "pm_peak"),
        ("am_peak_hour_cells", "am_peak"),
    ]
    for approach in ("north", "south", "west", "east"):
        table = approach_tables.get(approach, {})
        movement_codes = table.get("movement_codes", ())
        for movement_code in movement_codes:
            for cell_set_key, value_key in value_sets:
                cell_ref = (table.get(cell_set_key) or {}).get(movement_code)
                if cell_ref:
                    yield movement_code, cell_ref, value_key


def _write_movement_formulas(worksheet, template_map: dict[str, Any], guard: _FormulaWriteGuard | None = None) -> None:
    if template_map.get("movement_diagram_cells", {}).get("formula_write_mode") == "preserve_template":
        return
    for movement_code, cell_ref, value_key in _movement_formula_targets(template_map):
        _set_cell(
            worksheet,
            cell_ref,
            _diagram_formula(movement_code, value_key),
            guard,
            f"movement.{movement_code}.{value_key}",
            force=True,
        )


def _write_summary_formulas(worksheet, template_map: dict[str, Any], guard: _FormulaWriteGuard | None = None) -> None:
    if template_map.get("summary_formula_write_mode", "exporter") == "preserve_template":
        return
    summary = template_map.get("summary_box_cells", {})
    _set_cell(worksheet, summary.get("am_peak_total_pcu", {}).get("value_cell"), '=IFERROR(Peak_PHF!$F$2,"")', guard, "summary.am_peak_total_pcu")
    _set_cell(worksheet, summary.get("pm_peak_total_pcu", {}).get("value_cell"), '=IFERROR(Peak_PHF!$J$2,"")', guard, "summary.pm_peak_total_pcu")
    _set_cell(worksheet, summary.get("total_12_hour_pcu", {}).get("value_cell"), "=SUM(Diagram_Data!$B$2:$B$17)", guard, "summary.total_12_hour_pcu")


def _source_rows_by_label(dataframe: pd.DataFrame) -> tuple[dict[str, pd.Series], pd.Series | None]:
    label_column = _column_for_key(dataframe, "time")
    source_by_label: dict[str, pd.Series] = {}
    source_total: pd.Series | None = None
    if not label_column:
        return source_by_label, source_total

    for _, source_row in dataframe.iterrows():
        normalised_label = _normalise_time_label(source_row[label_column])
        if normalised_label in {"à¸£à¸§à¸¡", "total"}:
            source_total = source_row
        elif normalised_label:
            source_by_label[normalised_label] = source_row
    return source_by_label, source_total


def _write_table(worksheet, table_map: dict[str, Any], dataframe: pd.DataFrame, guard: _FormulaWriteGuard | None = None) -> None:
    if not table_map:
        return
    header_row = int(table_map["header_row"])
    first_data_row = int(table_map["first_data_row"])
    total_row = int(table_map.get("total_row") or first_data_row + max(len(dataframe), 1) - 1)
    columns = table_map.get("columns", {})
    time_column_letter = table_map.get("time_column")

    target_labels: dict[int, Any] = {}
    if time_column_letter:
        time_column_index = _column_number(time_column_letter)
        for row in range(first_data_row, total_row + 1):
            target_labels[row] = worksheet.Cells(row, time_column_index).Value

    source_by_label, source_total = _source_rows_by_label(dataframe)

    for key, column_letter in columns.items():
        source_column = _column_for_key(dataframe, key)
        column_index = _column_number(column_letter)
        _set_cell(worksheet, f"{column_letter}{header_row}", source_column or key, guard, f"table.header.{key}")
        for row in range(first_data_row, total_row + 1):
            _set_cell(worksheet, f"{column_letter}{row}", "", guard, f"table.clear.{key}")

        if target_labels:
            for target_row, label in target_labels.items():
                is_total = target_row == total_row
                source_row = source_total if is_total else source_by_label.get(_normalise_time_label(label))
                if key == "time":
                    value = label
                elif source_row is not None and source_column and source_column in dataframe.columns:
                    value = source_row[source_column]
                else:
                    value = 0
                _set_cell(worksheet, f"{column_letter}{target_row}", value, guard, f"table.value.{key}")
            continue

        for offset, (_, source_row) in enumerate(dataframe.iterrows()):
            target_row = first_data_row + offset
            if target_row > total_row:
                break
            value = source_row[source_column] if source_column and source_column in dataframe.columns else ""
            _set_cell(worksheet, f"{column_letter}{target_row}", value, guard, f"table.value.{key}")


def _sheet_exists(workbook, sheet_name: str) -> bool:
    for index in range(1, workbook.Worksheets.Count + 1):
        if workbook.Worksheets(index).Name == sheet_name:
            return True
    return False


def _replace_sheet(workbook, sheet_name: str):
    if _sheet_exists(workbook, sheet_name):
        workbook.Worksheets(sheet_name).Delete()
    sheet = workbook.Worksheets.Add(After=workbook.Worksheets(workbook.Worksheets.Count))
    sheet.Name = sheet_name[:31]
    return sheet


def _dataframe_values(dataframe: pd.DataFrame) -> tuple[tuple[Any, ...], ...]:
    rows: list[tuple[Any, ...]] = [tuple(str(column) for column in dataframe.columns)]
    for _, row in dataframe.iterrows():
        rows.append(tuple(_coerce_excel_value(value) for value in row.tolist()))
    return tuple(rows)


def _write_dataframe_sheet(workbook, sheet_name: str, dataframe: pd.DataFrame) -> None:
    worksheet = _replace_sheet(workbook, sheet_name)
    if len(dataframe.columns) == 0:
        worksheet.Cells(1, 1).Value = ""
        return
    payload = _dataframe_values(dataframe)
    worksheet.Range(
        worksheet.Cells(1, 1),
        worksheet.Cells(len(payload), len(payload[0])),
    ).Value = payload
    worksheet.Rows(1).Font.Bold = True
    worksheet.Rows(1).Interior.Color = 0x784E1F
    worksheet.Rows(1).Font.Color = 0xFFFFFF
    worksheet.Columns.AutoFit()


def _diagram_data_formula(value_column: str, movement_cell: str, hourly_rows: int, hourly_columns: int) -> str:
    last_letter = _column_name(hourly_columns)
    data_range = f"'Hourly_Movement_PCU'!$B$2:${last_letter}${hourly_rows}"
    time_range = f"'Hourly_Movement_PCU'!$A$2:$A${hourly_rows}"
    movement_headers = f"'Hourly_Movement_PCU'!$B$1:${last_letter}$1"
    movement_match = f"MATCH({movement_cell},{movement_headers},0)"
    if value_column == "total":
        return f"=IFERROR(INDEX({data_range},ROWS({time_range}),{movement_match}),0)"

    peak_cell = "'Peak_PHF'!$H$2" if value_column == "pm" else "'Peak_PHF'!$D$2"
    peak_key = f'IF(ISNUMBER({peak_cell}),TEXT({peak_cell},"hh:mm"),LEFT({peak_cell},5))&"*"'
    return f"=IFERROR(INDEX({data_range},MATCH({peak_key},{time_range},0),{movement_match}),0)"


def _column_name(number: int) -> str:
    name = ""
    while number:
        number, remainder = divmod(number - 1, 26)
        name = chr(65 + remainder) + name
    return name


def _write_diagram_data_sheet(
    workbook,
    hourly_movement: pd.DataFrame,
    *,
    sheet_name: str = "Diagram_Data",
    movement_codes: tuple[str, ...] = MOVEMENT_CODES,
) -> None:
    worksheet = _replace_sheet(workbook, sheet_name)
    headers = ("movement_code", "total_pcu", "pm_peak_pcu", "am_peak_pcu")
    worksheet.Range("A1:D1").Value = (headers,)
    hourly_rows = max(len(hourly_movement) + 1, 2)
    hourly_columns = max(len(hourly_movement.columns), 2)
    for row_index, movement_code in enumerate(movement_codes, start=2):
        worksheet.Cells(row_index, 1).Value = movement_code
        worksheet.Cells(row_index, 2).Formula = _diagram_data_formula("total", f"$A{row_index}", hourly_rows, hourly_columns)
        worksheet.Cells(row_index, 3).Formula = _diagram_data_formula("pm", f"$A{row_index}", hourly_rows, hourly_columns)
        worksheet.Cells(row_index, 4).Formula = _diagram_data_formula("am", f"$A{row_index}", hourly_rows, hourly_columns)
    worksheet.Columns.AutoFit()


def _write_support_sheets(workbook, report_data: dict[str, Any]) -> None:
    sheets = report_data.get("sheets", {})
    for sheet_name, dataframe in sheets.items():
        _write_dataframe_sheet(workbook, sheet_name, dataframe)
    diagram_data_sheet_name = str(report_data.get("diagram_data_sheet_name") or "Diagram_Data")
    movement_codes = tuple(report_data.get("diagram_movement_codes") or MOVEMENT_CODES)
    _write_diagram_data_sheet(
        workbook,
        report_data.get("hourly_movement_pcu", pd.DataFrame()),
        sheet_name=diagram_data_sheet_name,
        movement_codes=movement_codes,
    )


def _write_native_chart_sources(worksheet, template_map: dict[str, Any], chart_source_data: dict[str, Any], guard: _FormulaWriteGuard | None = None) -> None:
    anchors = template_map.get("chart_anchors", {})
    hourly_source = anchors.get("native_hourly_chart_source", {})
    hourly_data = chart_source_data.get("hourly_pcu", {})
    _write_range(worksheet, hourly_source.get("categories"), list(hourly_data.get("categories", [])), guard, "chart.hourly.categories")
    _write_range(worksheet, hourly_source.get("values"), list(hourly_data.get("values", [])), guard, "chart.hourly.values")

    vehicle_source = anchors.get("native_vehicle_composition_chart_source", {})
    vehicle_data = chart_source_data.get("vehicle_composition", {})
    _write_range(worksheet, vehicle_source.get("categories"), list(vehicle_data.get("categories", [])), guard, "chart.vehicle.categories")
    _write_range(worksheet, vehicle_source.get("values"), list(vehicle_data.get("values", [])), guard, "chart.vehicle.values")


def _build_export_diagnostics(template_path: Path, template_map: dict[str, Any]) -> ExcelComExportDiagnostics:
    formula_audit = audit_template_formulas(template_path, template_map, template_map.get("template_sheet", "Summary"))
    diagnostics = ExcelComExportDiagnostics(
        formula_cells_preserved=[
            f"{cell.sheet}!{cell.cell}: {cell.formula}" for cell in formula_audit.formula_cells
        ],
        formulas_with_external_links=[
            f"{issue.sheet}!{issue.cell}: {issue.formula}" for issue in formula_audit.external_links
        ],
        formulas_with_ref_errors=[
            f"{issue.sheet}!{issue.cell}: {issue.formula}" for issue in formula_audit.ref_errors
        ],
        formulas_with_missing_sheet_references=[
            f"{issue.sheet}!{issue.cell}: {issue.referenced_sheet}: {issue.formula}"
            for issue in formula_audit.missing_sheet_issues
        ],
        mapped_formula_cells=[
            f"{issue.sheet}!{issue.cell}: {issue.mapped_range}: {issue.formula}"
            for issue in formula_audit.mapped_formula_cells
        ],
    )
    return diagnostics


def _try_set_calculation(target: Any, attribute: str, value: Any, diagnostics: ExcelComExportDiagnostics) -> None:
    try:
        setattr(target, attribute, value)
    except Exception as exc:
        diagnostics.calculation_warnings.append(f"{attribute}: {exc}")


def _try_call_calculation(target: Any, method_name: str, diagnostics: ExcelComExportDiagnostics) -> None:
    try:
        method = getattr(target, method_name)
    except Exception as exc:
        diagnostics.calculation_warnings.append(f"{method_name}: {exc}")
        return
    try:
        method()
    except Exception as exc:
        diagnostics.calculation_warnings.append(f"{method_name}: {exc}")


def export_with_excel_com(
    template_path,
    output_path,
    template_map,
    report_data,
    metadata,
    chart_source_data,
) -> Path:
    """Export a template workbook through Microsoft Excel COM.

    Excel opens a copy of the template, data is written only to mapped
    cells/ranges, native chart objects are left untouched, and Excel recalculates
    and saves the output workbook.
    """

    global _LAST_EXPORT_DIAGNOSTICS

    require_excel_com()

    pythoncom, win32_client = _import_com_modules_with_repair()
    source = Path(template_path).resolve()
    target = Path(output_path).resolve()
    if source == target:
        raise ValueError("COM export output_path must not be the source template path.")
    if not source.exists():
        raise ReportTemplateUnavailable(f"Report template workbook not found: {source}")

    target.parent.mkdir(parents=True, exist_ok=True)
    try:
        shutil.copy2(source, target)
    except PermissionError as exc:
        raise PermissionError(
            "ไม่สามารถเขียนไฟล์รายงานได้ กรุณาปิดไฟล์ Excel ที่เปิดค้างไว้ก่อนส่งออกใหม่ "
            f"(target: {target})"
        ) from exc
    except OSError as exc:
        raise OSError(f"Unable to prepare Excel COM export workbook at {target}: {exc}") from exc
    diagnostics = _build_export_diagnostics(source, template_map)
    _LAST_EXPORT_DIAGNOSTICS = diagnostics

    excel = None
    workbook = None
    coinitialized = False
    pythoncom.CoInitialize()
    coinitialized = True
    try:
        excel = win32_client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.EnableEvents = False
        excel.AskToUpdateLinks = False
        _try_set_calculation(excel, "Calculation", -4105, diagnostics)  # xlCalculationAutomatic
        workbook = excel.Workbooks.Open(
            Filename=str(target),
            UpdateLinks=0,
            ReadOnly=False,
            IgnoreReadOnlyRecommended=True,
            AddToMru=False,
            CorruptLoad=0,
        )

        _write_support_sheets(workbook, report_data)
        summary_name = template_map.get("template_sheet", "Summary")
        if not _sheet_exists(workbook, summary_name):
            raise ReportTemplateUnavailable(f"Template sheet not found: {summary_name}")
        summary = workbook.Worksheets(summary_name)
        guard = _FormulaWriteGuard(template_map, diagnostics)
        _write_metadata(summary, template_map, metadata, guard)
        _write_effective_peak_label(summary, metadata)
        _write_movement_formulas(summary, template_map, guard)
        _write_summary_formulas(summary, template_map, guard)
        _write_table(summary, template_map.get("hourly_movement_table", {}), report_data.get("hourly_movement_pcu", pd.DataFrame()), guard)
        _write_table(summary, template_map.get("hourly_vehicle_class_table", {}), report_data.get("hourly_vehicle_class", pd.DataFrame()), guard)
        _write_native_chart_sources(summary, template_map, chart_source_data, guard)

        _try_set_calculation(workbook, "ForceFullCalculation", True, diagnostics)
        _try_call_calculation(excel, "CalculateFullRebuild", diagnostics)
        _try_call_calculation(excel, "Calculate", diagnostics)
        workbook.Save()
    finally:
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        if coinitialized:
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    return target
