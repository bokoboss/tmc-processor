"""Read-only checks for Excel report templates and template maps."""

from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from .diagram import MOVEMENT_CODES
from .movement_scheme import APPROACH_MOVEMENT_CODES, MOVEMENT_SCHEME_V1, MOVEMENT_SCHEME_V2


@dataclass(frozen=True)
class TemplateIntegrityResult:
    errors: tuple[str, ...]
    warnings: tuple[str, ...] = ()

    @property
    def ok(self) -> bool:
        return not self.errors


def load_template_map_for_integrity(map_path: str | Path) -> dict[str, Any]:
    path = Path(map_path)
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def expected_movement_codes_for_scheme(movement_code_scheme: str) -> list[str]:
    if movement_code_scheme == MOVEMENT_SCHEME_V2:
        return list(APPROACH_MOVEMENT_CODES)
    if movement_code_scheme == MOVEMENT_SCHEME_V1:
        return list(MOVEMENT_CODES)
    raise ValueError(f"Unsupported movement_code_scheme: {movement_code_scheme!r}.")


def mapped_hourly_movement_order(mapping: dict[str, Any]) -> list[str]:
    columns = mapping.get("hourly_movement_table", {}).get("columns", {})
    return [str(code) for code in columns if code not in {"time", "Total"}]


def mapped_diagram_movement_order(mapping: dict[str, Any]) -> list[str]:
    movements = mapping.get("movement_diagram_cells", {}).get("diagram_movements", {})
    return [str(code) for code in movements]


def mapped_approach_table_order(mapping: dict[str, Any]) -> list[str]:
    tables = mapping.get("movement_diagram_cells", {}).get("approach_tables", {})
    ordered: list[str] = []
    for approach in ("north", "south", "east", "west"):
        ordered.extend(str(code) for code in tables.get(approach, {}).get("movement_codes", []))
    return ordered


def map_movement_code_leaks(mapping: dict[str, Any], forbidden_codes: set[str]) -> list[str]:
    found: set[str] = set()
    found.update(mapped_hourly_movement_order(mapping))
    found.update(mapped_diagram_movement_order(mapping))
    found.update(mapped_approach_table_order(mapping))
    return sorted(found & forbidden_codes)


def _range_is_valid(cell_range: str) -> bool:
    try:
        range_boundaries(str(cell_range))
    except (TypeError, ValueError):
        return False
    return True


def _collect_mapped_ranges(mapping: dict[str, Any]) -> list[str]:
    ranges: list[str] = []
    for key in ("writable_ranges", "protected_formula_ranges", "formula_overwrite_ranges"):
        ranges.extend(str(value) for value in mapping.get(key, []) if value)
    for section in ("metadata_cells", "summary_box_cells"):
        for info in mapping.get(section, {}).values():
            for key in ("cell", "label_cell", "value_cell", "merged_range", "label_range", "value_range"):
                if isinstance(info, dict) and info.get(key):
                    ranges.append(str(info[key]))
    movement_info = mapping.get("movement_diagram_cells", {})
    for section in ("diagram_title", "diagram_date", "caption"):
        info = movement_info.get(section, {})
        for key in ("cell", "merged_range"):
            if isinstance(info, dict) and info.get(key):
                ranges.append(str(info[key]))
    for section in ("direction_labels", "road_labels"):
        for info in movement_info.get(section, {}).values():
            if isinstance(info, dict) and info.get("cell"):
                ranges.append(str(info["cell"]))
    for movement in movement_info.get("diagram_movements", {}).values():
        if not isinstance(movement, dict):
            continue
        for key in ("header_cell", "total_12_hour_cell", "pm_peak_hour_cell", "am_peak_hour_cell"):
            if movement.get(key):
                ranges.append(str(movement[key]))
    return ranges


def verify_template_integrity(
    template_path: str | Path,
    map_path: str | Path,
    movement_code_scheme: str,
) -> TemplateIntegrityResult:
    """Validate a template workbook and map without saving or modifying the workbook."""

    errors: list[str] = []
    warnings: list[str] = []
    template = Path(template_path)
    mapping = load_template_map_for_integrity(map_path)
    expected_codes = expected_movement_codes_for_scheme(movement_code_scheme)

    if movement_code_scheme == MOVEMENT_SCHEME_V2:
        template_version = mapping.get("template_version")
        if template_version and template_version != "four_leg_approach_movement_v2":
            errors.append(f"Unexpected v2 template_version: {template_version!r}.")
        if mapping.get("movement_code_scheme") not in {None, MOVEMENT_SCHEME_V2}:
            errors.append(f"Unexpected v2 movement_code_scheme: {mapping.get('movement_code_scheme')!r}.")

    declared_order = mapping.get("movement_code_order")
    if declared_order is not None and list(declared_order) != expected_codes:
        errors.append("Template map movement_code_order does not match the expected movement order.")

    hourly_order = mapped_hourly_movement_order(mapping)
    if hourly_order != expected_codes:
        errors.append("Hourly movement table headers do not match the expected movement order.")

    diagram_order = mapped_diagram_movement_order(mapping)
    if set(diagram_order) != set(expected_codes):
        errors.append("Diagram movement codes do not match the expected movement set.")

    approach_order = mapped_approach_table_order(mapping)
    if movement_code_scheme == MOVEMENT_SCHEME_V2 and approach_order and approach_order != expected_codes:
        errors.append("Approach table movement headers do not match the expected movement order.")
    if movement_code_scheme == MOVEMENT_SCHEME_V1 and approach_order and set(approach_order) != set(expected_codes):
        errors.append("Approach table movement headers do not match the expected movement set.")

    if movement_code_scheme == MOVEMENT_SCHEME_V2:
        leaked = map_movement_code_leaks(mapping, set(MOVEMENT_CODES) - set(APPROACH_MOVEMENT_CODES))
        if leaked:
            errors.append(f"Template map contains movement codes from the wrong scheme: {', '.join(leaked)}.")

    invalid_ranges = [cell_range for cell_range in _collect_mapped_ranges(mapping) if not _range_is_valid(cell_range)]
    if invalid_ranges:
        errors.append(f"Template map contains invalid cell/range references: {', '.join(invalid_ranges)}.")

    if not template.exists():
        errors.append(f"Template workbook not found: {template}")
        return TemplateIntegrityResult(tuple(errors), tuple(warnings))

    try:
        workbook = load_workbook(template, data_only=False)
    except Exception as exc:  # pragma: no cover - defensive around malformed binary workbooks
        errors.append(f"Template workbook cannot be opened by openpyxl: {exc}")
        return TemplateIntegrityResult(tuple(errors), tuple(warnings))

    required_sheets = mapping.get("workbook_structure", {}).get("sheets") or [mapping.get("template_sheet", "Summary")]
    missing_sheets = [sheet for sheet in required_sheets if sheet not in workbook.sheetnames]
    if missing_sheets:
        errors.append(f"Template workbook is missing required sheets: {', '.join(missing_sheets)}.")

    template_sheet = mapping.get("template_sheet", required_sheets[0] if required_sheets else "Summary")
    if template_sheet not in workbook.sheetnames:
        return TemplateIntegrityResult(tuple(errors), tuple(warnings))

    worksheet = workbook[template_sheet]
    header_row = int(mapping.get("hourly_movement_table", {}).get("header_row", 0) or 0)
    for code, column in mapping.get("hourly_movement_table", {}).get("columns", {}).items():
        if code in {"time", "Total"} or not header_row:
            continue
        value = worksheet[f"{column}{header_row}"].value
        if value != code:
            errors.append(f"Hourly movement header {column}{header_row} expected {code!r}, found {value!r}.")

    for code, info in mapping.get("movement_diagram_cells", {}).get("diagram_movements", {}).items():
        if not isinstance(info, dict):
            continue
        header_cell = info.get("header_cell")
        if header_cell and worksheet[str(header_cell)].value != code:
            errors.append(f"Diagram movement header {header_cell} expected {code!r}, found {worksheet[str(header_cell)].value!r}.")

    return TemplateIntegrityResult(tuple(errors), tuple(warnings))
