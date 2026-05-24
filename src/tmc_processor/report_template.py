"""Template-driven four-leg TMC report sheet export helpers."""

from __future__ import annotations

import json
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, time
from io import BytesIO
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

import pandas as pd
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries

from .constants import VEHICLE_CLASSES
from .metadata import direction_label_value, metadata_cell_values
from .template_audit import validate_template_before_export


ROOT_DIR = Path(__file__).resolve().parents[2]
DEFAULT_TEMPLATE_PATH = ROOT_DIR / "templates" / "four_leg_tmc_report_template.xlsx"
DEFAULT_TEMPLATE_MAP_PATH = ROOT_DIR / "templates" / "four_leg_tmc_report_template_map.json"


@dataclass(frozen=True)
class ReportTemplateResources:
    template_path: Path
    map_path: Path
    mapping: dict[str, Any]
    warnings: tuple[str, ...] = ()


class ReportTemplateUnavailable(RuntimeError):
    """Raised when template-driven report export cannot be used safely."""


CHART_NS = {"c": "http://schemas.openxmlformats.org/drawingml/2006/chart"}
NATIVE_CHART_KEYS = ("hourly_pcu_line_chart", "vehicle_composition_bar_chart")
DESTINATION_LABEL_KEYS = ("north_label", "south_label", "east_label", "west_label")
ROAD_LABEL_KEYS = ("north_road", "south_road", "east_road", "west_road")


def load_template_map(map_path: str | Path = DEFAULT_TEMPLATE_MAP_PATH) -> dict[str, Any]:
    path = Path(map_path)
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def load_report_template_resources(
    template_path: str | Path = DEFAULT_TEMPLATE_PATH,
    map_path: str | Path = DEFAULT_TEMPLATE_MAP_PATH,
) -> ReportTemplateResources:
    template = Path(template_path)
    mapping_path = Path(map_path)
    if not template.exists():
        raise ReportTemplateUnavailable(f"Report template workbook not found: {template}")
    if not mapping_path.exists():
        raise ReportTemplateUnavailable(f"Report template map not found: {mapping_path}")

    validation = validate_template_before_export(template)
    if validation.should_fallback_to_generated_report:
        raise ReportTemplateUnavailable(
            "Report template has blocking validation warnings: "
            + "; ".join(validation.blocking_warnings)
        )
    return ReportTemplateResources(
        template_path=template,
        map_path=mapping_path,
        mapping=load_template_map(mapping_path),
        warnings=validation.non_blocking_warnings,
    )


def validate_native_chart_sources(template_path: str | Path, mapping: dict[str, Any]) -> tuple[str, ...]:
    """Return warnings for missing or unsafe native chart sources in the template package."""

    warnings: list[str] = []
    anchors = mapping.get("chart_anchors", {})
    with zipfile.ZipFile(template_path) as archive:
        names = set(archive.namelist())
        for chart_key in NATIVE_CHART_KEYS:
            info = anchors.get(chart_key, {})
            chart_xml = info.get("chart_xml")
            if not chart_xml:
                warnings.append(f"{chart_key}: chart_xml is not mapped")
                continue
            if chart_xml not in names:
                warnings.append(f"{chart_key}: chart part not found: {chart_xml}")
                continue

            root = ET.fromstring(archive.read(chart_xml))
            formulas = [node.text or "" for node in root.findall(".//c:f", CHART_NS)]
            expected = list((info.get("source_ranges") or {}).values())
            if expected and formulas[: len(expected)] != expected:
                warnings.append(
                    f"{chart_key}: chart source ranges differ from map: "
                    f"expected {expected}, found {formulas[:len(expected)]}"
                )
            for formula in formulas:
                if "#REF!" in formula:
                    warnings.append(f"{chart_key}: chart source contains #REF!: {formula}")
                if "[" in formula and "]" in formula:
                    warnings.append(f"{chart_key}: chart source contains external workbook link: {formula}")
    return tuple(warnings)


def restore_template_native_drawings(
    workbook_bytes: bytes,
    template_path: str | Path,
    target_sheet_name: str = "TMC_Report",
    source_sheet_name: str = "Summary",
) -> bytes:
    """Deprecated no-op for the disabled native template chart restoration path."""

    # Native chart preservation via OOXML package patching is disabled because it can corrupt worksheet XML.
    return workbook_bytes


def _blank_if_missing(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return value


def _first_present(setup: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        value = setup.get(key)
        if value is None:
            continue
        if isinstance(value, float) and pd.isna(value):
            continue
        if isinstance(value, str) and value.strip() == "":
            continue
        return value
    return ""


def _time_text(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, pd.Timestamp):
        value = value.time()
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    return str(value)


def _date_text(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, pd.Timestamp):
        value = value.date()
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _survey_point(setup: dict[str, Any]) -> str:
    explicit = _first_present(setup, "survey_point")
    if explicit:
        return str(explicit).strip()
    fallback = " ".join(str(setup.get(key) or "").strip() for key in ("tmc_id", "tmc_name")).strip()
    if fallback:
        return fallback
    return str(_first_present(setup, "intersection_name", "tmc_title")).strip()


def _metadata_values(setup: dict[str, Any]) -> dict[str, Any]:
    return metadata_cell_values(setup)


def _first_data_column(dataframe: pd.DataFrame) -> str | None:
    return str(dataframe.columns[0]) if len(dataframe.columns) else None


def _column_for_key(dataframe: pd.DataFrame, key: str) -> str | None:
    if dataframe.empty and len(dataframe.columns) == 0:
        return None
    if key == "time":
        return _first_data_column(dataframe)
    explicit = {
        "total_pcu": "Total (PCU)",
        "total_vehicles": "Total (คัน)",
    }
    wanted = explicit.get(key, key)
    if wanted in dataframe.columns:
        return wanted
    lowered = {str(column).strip().lower(): str(column) for column in dataframe.columns}
    return lowered.get(str(wanted).strip().lower())


def _normalise_time_label(value: Any) -> str:
    return str(value or "").strip().replace(".", ":").lower()


def _excel_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    if pd.isna(value) if not isinstance(value, (list, tuple, dict, str)) else False:
        return ""
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    return value


def _strip_sheet_reference(reference: str) -> str:
    text = str(reference).strip()
    if "!" in text:
        text = text.rsplit("!", 1)[1]
    return text.replace("$", "")


def _cell_in_ranges(cell_ref: str, ranges: list[str]) -> bool:
    cleaned_cell = _strip_sheet_reference(cell_ref)
    cell_min_col, cell_min_row, _, _ = range_boundaries(cleaned_cell)
    for cell_range in ranges:
        try:
            cleaned_range = _strip_sheet_reference(cell_range)
            min_col, min_row, max_col, max_row = range_boundaries(cleaned_range)
        except ValueError:
            continue
        if min_col <= cell_min_col <= max_col and min_row <= cell_min_row <= max_row:
            return True
    return False


def _is_protected_formula_cell(worksheet, cell_ref: str, protected_formula_ranges: list[str]) -> bool:
    if not protected_formula_ranges:
        return False
    if not _cell_in_ranges(cell_ref, protected_formula_ranges):
        return False
    value = worksheet[cell_ref].value
    return isinstance(value, str) and value.startswith("=")


def _set_cell(
    worksheet,
    cell_ref: str | None,
    value: Any,
    protected_formula_ranges: list[str] | None = None,
    formula_overwrite_ranges: list[str] | None = None,
) -> None:
    if cell_ref:
        cell = worksheet[cell_ref]
        if (
            _is_protected_formula_cell(worksheet, cell_ref, protected_formula_ranges or [])
            and not _cell_in_ranges(cell_ref, formula_overwrite_ranges or [])
        ):
            return
        if not isinstance(cell, MergedCell):
            cell.value = _excel_value(value)


def _diagram_formula(movement_code: str, value_key: str) -> str:
    column = {"total_12_hour": "B", "pm_peak": "C", "am_peak": "D"}[value_key]
    return f'=IFERROR(INDEX(Diagram_Data!${column}:${column},MATCH("{movement_code}",Diagram_Data!$A:$A,0)),0)'


def _movement_formula_targets(mapping: dict[str, Any]):
    approach_tables = mapping.get("movement_diagram_cells", {}).get("approach_tables", {})
    value_sets = [
        ("total_12_hour_cells", "total_12_hour"),
        ("pm_peak_hour_cells", "pm_peak"),
        ("am_peak_hour_cells", "am_peak"),
    ]
    for approach in ("north", "south", "west", "east"):
        table = approach_tables.get(approach, {})
        movement_codes = table.get("movement_codes", ())
        for movement_code in movement_codes:
            for cell_set_key, value_key in value_sets:
                cell_ref = (table.get(cell_set_key) or {}).get(movement_code)
                if cell_ref:
                    yield movement_code, cell_ref, value_key


def _write_metadata(worksheet, mapping: dict[str, Any], setup: dict[str, Any]) -> None:
    values = _metadata_values(setup)
    protected_formula_ranges = list(mapping.get("protected_formula_ranges") or [])
    formula_overwrite_ranges = list(mapping.get("formula_overwrite_ranges") or [])
    for key, info in mapping.get("metadata_cells", {}).items():
        _set_cell(
            worksheet,
            info.get("value_cell") or info.get("cell"),
            values.get(key, ""),
            protected_formula_ranges,
            formula_overwrite_ranges,
        )

    movement_info = mapping.get("movement_diagram_cells", {})
    diagram_title = movement_info.get("diagram_title", {}).get("cell")
    diagram_date = movement_info.get("diagram_date", {}).get("cell")
    caption = movement_info.get("caption", {}).get("cell")
    _set_cell(worksheet, diagram_title, values.get("survey_point", ""), protected_formula_ranges, formula_overwrite_ranges)
    _set_cell(worksheet, diagram_date, values.get("survey_date", ""), protected_formula_ranges, formula_overwrite_ranges)
    for direction, info in movement_info.get("direction_labels", {}).items():
        _set_cell(
            worksheet,
            info.get("cell"),
            _mapped_label_value(setup, direction),
            protected_formula_ranges,
            formula_overwrite_ranges,
        )
    for label_key, info in movement_info.get("road_labels", {}).items():
        _set_cell(
            worksheet,
            info.get("cell"),
            _mapped_label_value(setup, label_key),
            protected_formula_ranges,
            formula_overwrite_ranges,
        )
    _set_cell(worksheet, caption, setup.get("caption_text") or "", protected_formula_ranges, formula_overwrite_ranges)


def _direction_label_value(setup: dict[str, Any], direction: str) -> str:
    return direction_label_value(setup, direction)


def _mapped_label_value(setup: dict[str, Any], label_key: str) -> str:
    if label_key in DESTINATION_LABEL_KEYS or label_key in ROAD_LABEL_KEYS:
        return str(_first_present(setup, label_key) or "")
    return _direction_label_value(setup, label_key)


def _write_movement_formulas(worksheet, mapping: dict[str, Any]) -> None:
    for movement_code, cell_ref, value_key in _movement_formula_targets(mapping):
        _set_cell(worksheet, cell_ref, _diagram_formula(movement_code, value_key), [])


def _write_summary_formulas(worksheet, mapping: dict[str, Any]) -> None:
    if mapping.get("summary_formula_write_mode", "exporter") == "preserve_template":
        return
    protected_formula_ranges = list(mapping.get("protected_formula_ranges") or [])
    summary = mapping.get("summary_box_cells", {})
    _set_cell(worksheet, summary.get("am_peak_total_pcu", {}).get("value_cell"), '=IFERROR(Peak_PHF!$F$2,"")', protected_formula_ranges)
    _set_cell(worksheet, summary.get("pm_peak_total_pcu", {}).get("value_cell"), '=IFERROR(Peak_PHF!$J$2,"")', protected_formula_ranges)
    _set_cell(worksheet, summary.get("total_12_hour_pcu", {}).get("value_cell"), "=SUM(Diagram_Data!$B$2:$B$17)", protected_formula_ranges)


def _clear_range_values(worksheet, cell_range: str) -> None:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in range(min_row, max_row + 1):
        for column in range(min_col, max_col + 1):
            cell = worksheet.cell(row=row, column=column)
            if not isinstance(cell, MergedCell):
                cell.value = ""


def _write_table(
    worksheet,
    table_map: dict[str, Any],
    dataframe: pd.DataFrame,
    protected_formula_ranges: list[str] | None = None,
) -> None:
    if not table_map:
        return
    header_row = int(table_map["header_row"])
    first_data_row = int(table_map["first_data_row"])
    total_row = int(table_map.get("total_row") or first_data_row + max(len(dataframe), 1) - 1)
    columns = table_map.get("columns", {})
    time_column_letter = table_map.get("time_column")

    target_labels: dict[int, Any] = {}
    source_by_label: dict[str, pd.Series] = {}
    source_total: pd.Series | None = None
    source_label_column = _column_for_key(dataframe, "time")
    if time_column_letter and source_label_column:
        time_column_index = column_index_from_string(time_column_letter)
        target_labels = {
            row: worksheet.cell(row=row, column=time_column_index).value
            for row in range(first_data_row, total_row + 1)
        }
        for _, source_row in dataframe.iterrows():
            label = source_row[source_label_column]
            normalised_label = _normalise_time_label(label)
            if normalised_label in {"รวม", "total"}:
                source_total = source_row
            elif normalised_label:
                source_by_label[normalised_label] = source_row

    for key, column_letter in columns.items():
        source_column = _column_for_key(dataframe, key)
        column_index = column_index_from_string(column_letter)
        header_cell = worksheet.cell(row=header_row, column=column_index)
        if not isinstance(header_cell, MergedCell):
            header_cell.value = source_column or key
        for row in range(first_data_row, total_row + 1):
            _set_cell(
                worksheet,
                f"{column_letter}{row}",
                "",
                protected_formula_ranges,
            )

        if target_labels:
            for target_row, label in target_labels.items():
                is_total = target_row == total_row
                source_row = source_total if is_total else source_by_label.get(_normalise_time_label(label))
                if key == "time":
                    value = label
                elif source_row is not None and source_column and source_column in dataframe.columns:
                    value = source_row[source_column]
                else:
                    value = 0
                _set_cell(
                    worksheet,
                    f"{column_letter}{target_row}",
                    value,
                    protected_formula_ranges,
                )
            continue

        for offset, (_, source_row) in enumerate(dataframe.iterrows()):
            target_row = first_data_row + offset
            if target_row > total_row:
                break
            value = source_row[source_column] if source_column and source_column in dataframe.columns else ""
            _set_cell(
                worksheet,
                f"{column_letter}{target_row}",
                value,
                protected_formula_ranges,
            )


def _write_horizontal_range(worksheet, cell_range: str | None, values: list[Any]) -> None:
    if not cell_range:
        return
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    if min_row != max_row:
        raise ReportTemplateUnavailable(f"Native chart source range must be one row: {cell_range}")
    for offset, column in enumerate(range(min_col, max_col + 1)):
        value = values[offset] if offset < len(values) else ""
        cell = worksheet.cell(row=min_row, column=column)
        if not isinstance(cell, MergedCell):
            cell.value = _excel_value(value)


def _vehicle_chart_percentages(vehicle_composition_report: pd.DataFrame) -> tuple[list[str], list[Any]]:
    if vehicle_composition_report.empty:
        return VEHICLE_CLASSES, [0 for _ in VEHICLE_CLASSES]

    class_column = "vehicle_class" if "vehicle_class" in vehicle_composition_report.columns else None
    percent_column = next(
        (
            column
            for column in ["สัดส่วน (%)", "percent", "count_share"]
            if column in vehicle_composition_report.columns
        ),
        None,
    )
    if class_column is None or percent_column is None:
        return VEHICLE_CLASSES, [0 for _ in VEHICLE_CLASSES]

    rows = vehicle_composition_report.copy()
    rows[class_column] = rows[class_column].astype(str)
    lookup = {
        row[class_column]: _excel_value(row[percent_column])
        for _, row in rows.iterrows()
        if row[class_column] != "Total"
    }
    return VEHICLE_CLASSES, [lookup.get(vehicle_class, 0) for vehicle_class in VEHICLE_CLASSES]


def _write_native_chart_sources(
    worksheet,
    mapping: dict[str, Any],
    vehicle_composition_report: pd.DataFrame,
) -> None:
    chart_sources = mapping.get("chart_anchors", {})
    vehicle_source = chart_sources.get("native_vehicle_composition_chart_source", {})
    categories, percentages = _vehicle_chart_percentages(vehicle_composition_report)
    _write_horizontal_range(worksheet, vehicle_source.get("categories"), categories)
    _write_horizontal_range(worksheet, vehicle_source.get("values"), percentages)


def _add_png(worksheet, png: bytes | None, anchor: str | None, width: int, height: int) -> None:
    if not png or not anchor:
        return
    image = OpenpyxlImage(BytesIO(png))
    image.width = width
    image.height = height
    worksheet.add_image(image, anchor)


def _insert_png_charts(worksheet, mapping: dict[str, Any], chart_pngs: dict[str, bytes]) -> None:
    anchors = mapping.get("chart_anchors", {})
    hourly_anchor = anchors.get("hourly_pcu_png_anchor", {}).get("cell")
    vehicle_anchor = anchors.get("vehicle_composition_png_anchor", {}).get("cell")
    _add_png(worksheet, chart_pngs.get("hourly_pcu"), hourly_anchor, 600, 310)
    _add_png(worksheet, chart_pngs.get("vehicle_composition"), vehicle_anchor, 600, 340)


def populate_template_report_sheet(
    workbook,
    mapping: dict[str, Any],
    setup: dict[str, Any],
    hourly_movement_pcu: pd.DataFrame,
    hourly_vehicle_class: pd.DataFrame,
    vehicle_composition_report: pd.DataFrame,
    chart_pngs: dict[str, bytes],
    report_sheet_name: str = "TMC_Report",
    use_native_template_charts: bool = False,
) -> None:
    template_sheet_name = mapping.get("template_sheet", "Summary")
    if template_sheet_name not in workbook.sheetnames:
        raise ReportTemplateUnavailable(f"Template sheet not found: {template_sheet_name}")

    worksheet = workbook[template_sheet_name]
    if worksheet.title != report_sheet_name:
        if report_sheet_name in workbook.sheetnames:
            del workbook[report_sheet_name]
        worksheet.title = report_sheet_name

    _write_metadata(worksheet, mapping, setup)
    _write_movement_formulas(worksheet, mapping)
    _write_summary_formulas(worksheet, mapping)
    protected_formula_ranges = list(mapping.get("protected_formula_ranges") or [])
    _write_table(worksheet, mapping.get("hourly_movement_table", {}), hourly_movement_pcu, protected_formula_ranges)
    _write_table(worksheet, mapping.get("hourly_vehicle_class_table", {}), hourly_vehicle_class, protected_formula_ranges)
    if use_native_template_charts:
        _write_native_chart_sources(worksheet, mapping, vehicle_composition_report)
    else:
        _insert_png_charts(worksheet, mapping, chart_pngs)
