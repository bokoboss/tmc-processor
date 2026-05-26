"""Traceability summary and ZIP package helpers for report exports."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
import re
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook
import pandas as pd

from .mapping import clean_mapping, mapping_to_excel_bytes, movement_aggregation_messages
from .metadata import APP_VERSION, TEMPLATE_VERSION, generated_timestamp_text, get_app_version
from .movement_scheme import MOVEMENT_SCHEME_V2
from .pcu import pce_factor_traceability_frame


PACKAGE_MIME = "application/zip"
V2_GENERATED_TEMPLATE_VERSION = "generated_approach_movement_v2"
V2_MOVEMENT_DIAGRAM_DATA_SHEET_NAME = "Movement_Diagram_Data"


def app_version() -> str:
    """Return the application version from the central metadata helper."""

    return get_app_version()


def _timestamp_text(generated_at: datetime | str | None = None) -> str:
    return generated_timestamp_text(generated_at)


def _safe_member_name(name: str | None, default: str) -> str:
    base = Path(str(name or default)).name
    base = re.sub(r"[^A-Za-z0-9._-]+", "_", base).strip("._")
    return base or default


def safe_package_filename(workbook_filename: str | None, default: str = "tmc_export_package.zip") -> str:
    stem = Path(_safe_member_name(workbook_filename, "tmc_report.xlsx")).stem
    cleaned = re.sub(r"[^A-Za-z0-9_-]+", "_", stem).strip("_")
    return f"{cleaned[:60] or 'tmc_export'}_package.zip" if cleaned else default


def _text_value(value: Any) -> str:
    if value is None:
        return ""
    if pd.isna(value) if not isinstance(value, (dict, list, tuple, set)) else False:
        return ""
    text = str(value).strip()
    return text[:5] if len(text) >= 5 and text[2:3] == ":" else text


def _peak_period(setup: dict[str, Any], peaks: pd.DataFrame | None, period: str) -> str:
    prefix = period.casefold()
    start = _text_value(setup.get(f"{prefix}_peak_start"))
    end = _text_value(setup.get(f"{prefix}_peak_end"))
    if (not start or not end) and peaks is not None and not peaks.empty and "period" in peaks.columns:
        rows = peaks[peaks["period"].astype(str).str.upper() == period.upper()]
        if not rows.empty:
            row = rows.iloc[0]
            start = start or _text_value(row.get("peak_start"))
            end = end or _text_value(row.get("peak_end"))
    return f"{start}-{end}" if start and end else ""


def _template_version(template_version: str | None, export_settings: dict[str, Any] | None) -> str:
    if template_version:
        return str(template_version)
    export_settings = export_settings or {}
    return str(export_settings.get("template_version") or TEMPLATE_VERSION)


def _mapping_summary(mapping: pd.DataFrame | None) -> list[str]:
    if mapping is None or mapping.empty:
        return ["Mapping rows: 0"]
    cleaned = clean_mapping(mapping)
    included = cleaned[cleaned["include_in_report"]]
    lines = [
        f"Mapping rows: {len(cleaned)}",
        f"Included in report: {len(included)}",
        f"Included in peak: {int(cleaned['include_in_peak'].sum())}",
    ]
    messages = movement_aggregation_messages(cleaned)
    if messages:
        lines.append("Aggregated movements:")
        lines.extend(f"- {message}" for message in messages)
    else:
        lines.append("Aggregated movements: none")
    return lines


def _qc_count_summary(qc: pd.DataFrame | None) -> str:
    if qc is None or qc.empty or "severity" not in qc.columns:
        return "total=0, error=0, warning=0, info=0"
    counts = qc["severity"].fillna("").astype(str).str.casefold().value_counts()
    return (
        f"total={len(qc)}, "
        f"error={int(counts.get('error', 0))}, "
        f"warning={int(counts.get('warning', 0))}, "
        f"info={int(counts.get('info', 0))}"
    )


def build_export_summary_text(
    *,
    setup: dict[str, Any] | None = None,
    source_file_name: str | None = None,
    export_mode: str | None = None,
    peaks: pd.DataFrame | None = None,
    mapping: pd.DataFrame | None = None,
    qc: pd.DataFrame | None = None,
    workbook_filename: str | None = None,
    pce_factors: dict[str, float] | None = None,
    export_settings: dict[str, Any] | None = None,
    template_version: str | None = None,
    generated_at: datetime | str | None = None,
) -> str:
    """Build a plain-text export summary suitable for traceability packages."""

    setup = setup or {}
    pce_rows = pce_factor_traceability_frame(pce_factors)
    survey_title = setup.get("survey_point") or setup.get("tmc_title") or setup.get("tmc_name") or ""
    peak_selection_source = setup.get("peak_selection_source") or ""
    if not peak_selection_source and peaks is not None and not peaks.empty and "peak_selection_source" in peaks.columns:
        peak_selection_source = str(peaks["peak_selection_source"].dropna().iloc[0]) if peaks["peak_selection_source"].notna().any() else ""

    lines = [
        "TMC Processor Export Summary",
        "",
        f"App version: {APP_VERSION}",
        f"Template version: {_template_version(template_version, export_settings)}",
        f"Generated at: {_timestamp_text(generated_at)}",
        f"Source file name: {Path(str(source_file_name or '')).name}",
        f"Survey point / TMC title: {survey_title}",
        f"Export mode: {export_mode or ''}",
        f"AM peak period: {_peak_period(setup, peaks, 'AM')}",
        f"PM peak period: {_peak_period(setup, peaks, 'PM')}",
        f"Peak selection source: {peak_selection_source}",
        "",
        "PCE factors:",
    ]
    lines.extend(
        f"- {row.vehicle_class}: {float(row.pce_factor):g} ({row.source})"
        for row in pce_rows.itertuples(index=False)
    )
    lines.extend(["", "Mapping aggregation summary:"])
    lines.extend(_mapping_summary(mapping))
    lines.extend(
        [
            "",
            f"QC warnings / info count: {_qc_count_summary(qc)}",
            f"Output workbook name: {_safe_member_name(workbook_filename, 'tmc_report.xlsx')}",
            "",
        ]
    )
    return "\n".join(lines)


def create_export_package_zip(
    *,
    workbook_bytes: bytes,
    workbook_filename: str,
    export_summary_text: str,
    project_session_bytes: bytes | None = None,
    project_session_filename: str | None = None,
    mapping_preset_bytes: bytes | None = None,
    mapping_preset_filename: str | None = None,
    mapping: pd.DataFrame | None = None,
    chart_pngs: dict[str, bytes] | None = None,
    diagram_png: bytes | None = None,
) -> bytes:
    """Create an in-memory ZIP package from explicitly provided export artifacts."""

    output = BytesIO()
    with ZipFile(output, "w", compression=ZIP_DEFLATED) as archive:
        archive.writestr(_safe_member_name(workbook_filename, "tmc_report.xlsx"), bytes(workbook_bytes))
        archive.writestr("export_summary.txt", export_summary_text.encode("utf-8"))
        if project_session_bytes:
            archive.writestr(
                _safe_member_name(project_session_filename, "tmc_session.tmcproj.json"),
                bytes(project_session_bytes),
            )
        if mapping_preset_bytes:
            archive.writestr(
                _safe_member_name(mapping_preset_filename, "mapping_preset.mapping.json"),
                bytes(mapping_preset_bytes),
            )
        if mapping is not None and not mapping.empty:
            archive.writestr("mapping_table.xlsx", mapping_to_excel_bytes(mapping))
        for name, png_bytes in sorted((chart_pngs or {}).items()):
            if png_bytes:
                archive.writestr(f"charts/{_safe_member_name(name, 'chart')}.png", bytes(png_bytes))
        if diagram_png:
            archive.writestr("charts/tmc_movement_diagram.png", bytes(diagram_png))
    return output.getvalue()


def _workbook_sheet_csv_bytes(workbook_bytes: bytes, sheet_name: str) -> bytes | None:
    try:
        workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
    except Exception:
        return None
    try:
        if sheet_name not in workbook.sheetnames:
            return None
        rows = list(workbook[sheet_name].iter_rows(values_only=True))
        if not rows:
            return None
        return pd.DataFrame(rows[1:], columns=list(rows[0])).to_csv(index=False).encode("utf-8")
    finally:
        workbook.close()


def create_v2_generated_export_package_zip(
    *,
    workbook_bytes: bytes,
    workbook_filename: str = "approach_movement_v2_generated_workbook.xlsx",
    setup: dict[str, Any] | None = None,
    peaks: pd.DataFrame | None = None,
    mapping: pd.DataFrame | None = None,
    qc: pd.DataFrame | None = None,
    mapping_preset_bytes: bytes | None = None,
    mapping_preset_filename: str | None = None,
    source_file_name: str | None = None,
    export_mode: str = "Safe PNG Export Mode",
    generated_at: datetime | str | None = None,
) -> bytes:
    """Package a v2 generated workbook and traceability text without raw input files."""

    package_setup = {**(setup or {}), "movement_code_scheme": MOVEMENT_SCHEME_V2}
    summary = build_export_summary_text(
        setup=package_setup,
        source_file_name=source_file_name,
        export_mode=export_mode,
        peaks=peaks,
        mapping=mapping,
        qc=qc,
        workbook_filename=workbook_filename,
        export_settings={
            "movement_code_scheme": MOVEMENT_SCHEME_V2,
            "template_version": V2_GENERATED_TEMPLATE_VERSION,
            "v2_export_scope": "generated workbook with table-based movement diagram data; template/native export unsupported",
        },
        template_version=V2_GENERATED_TEMPLATE_VERSION,
        generated_at=generated_at,
    )
    output = BytesIO()
    with ZipFile(output, "w", compression=ZIP_DEFLATED) as archive:
        archive.writestr(_safe_member_name(workbook_filename, "tmc_report.xlsx"), bytes(workbook_bytes))
        archive.writestr("export_summary.txt", summary.encode("utf-8"))
        if mapping_preset_bytes:
            archive.writestr(
                _safe_member_name(mapping_preset_filename, "mapping_preset.mapping.json"),
                bytes(mapping_preset_bytes),
            )
        if mapping is not None and not mapping.empty:
            archive.writestr("mapping_table.xlsx", mapping_to_excel_bytes(mapping))
        diagram_csv = _workbook_sheet_csv_bytes(workbook_bytes, V2_MOVEMENT_DIAGRAM_DATA_SHEET_NAME)
        if diagram_csv:
            archive.writestr("diagram/movement_diagram_data.csv", diagram_csv)
    return output.getvalue()
