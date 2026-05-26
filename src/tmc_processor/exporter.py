"""Excel workbook export."""

from __future__ import annotations

import warnings
from pathlib import Path
from tempfile import TemporaryDirectory
from io import BytesIO
from datetime import datetime, time
from typing import Any

from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import pandas as pd

from .charts import report_chart_pngs
from .constants import DEFAULT_PCE_FACTORS, DEFAULT_PEAK_MODE, VEHICLE_CLASSES
from .diagram import DiagramConfig, MOVEMENT_CODES, generate_four_leg_tmc_diagram
from .metadata import APP_VERSION, TEMPLATE_VERSION, generated_timestamp_text, metadata_cell_values, setup_with_metadata
from .mapping import clean_mapping
from .movement_scheme import MOVEMENT_SCHEME_V2, normalize_movement_code_scheme
from .pcu import pce_factor_traceability_frame
from .peaks import PEAK_SELECTION_USER_CONFIRMED, confirmed_peak_periods_from_setup, confirmed_peak_phf
from .report_template import (
    DEFAULT_TEMPLATE_MAP_PATH,
    DEFAULT_TEMPLATE_PATH,
    ReportTemplateUnavailable,
    load_report_template_resources,
    populate_template_report_sheet,
)
from .summaries import (
    hourly_movement_pcu,
    hourly_vehicle_class,
    movement_aggregation_audit,
    phf_15min,
    vehicle_composition_report,
    vehicle_group_pce,
)
from .time_utils import hourly_interval_rows


EXPORT_SHEETS = [
    "Export_Metadata",
    "Setup",
    "PCE_Factors",
    "Mapping",
    "Movement_Aggregation_Audit",
    "Normalized_Data",
    "QC_Check",
    "Hourly_Summary",
    "Movement_Summary",
    "Vehicle_Composition",
    "Hourly_Movement_PCU",
    "Hourly_Vehicle_Class",
    "Vehicle_Composition_Report",
    "Vehicle_Group_PCE",
    "PHF_15min",
    "Peak_PHF",
    "Report_Text",
]

CHART_SHEET_NAME = "Charts"
DIAGRAM_DATA_SHEET_NAME = "Diagram_Data"
TMC_REPORT_SHEET_NAME = "TMC_Report"
DIAGRAM_SHEET_NAME = "Diagram"
DEFAULT_CREATE_EXCEL_TABLES = False


def _export_metadata_frame(
    setup: dict[str, Any],
    *,
    export_mode: str | None = None,
    source_file_name: str | None = None,
    generated_at: datetime | str | None = None,
    template_version: str = TEMPLATE_VERSION,
) -> pd.DataFrame:
    metadata_values = metadata_cell_values(setup)
    rows = [
        ("app_version", APP_VERSION),
        ("template_version", template_version),
        ("generated_at", generated_timestamp_text(generated_at)),
        ("export_mode", export_mode or ""),
        ("source_file_name", Path(str(source_file_name or "")).name),
        ("report_title", metadata_values.get("report_title", "")),
        ("project", metadata_values.get("project", "")),
        ("survey_point", metadata_values.get("survey_point", "")),
        ("survey_date", metadata_values.get("survey_date", "")),
        ("weather", metadata_values.get("weather", "")),
        ("responsible_party", metadata_values.get("responsible_party", "")),
        ("survey_period", metadata_values.get("survey_period", "")),
        ("north_label", setup.get("north_label", "")),
        ("south_label", setup.get("south_label", "")),
        ("east_label", setup.get("east_label", "")),
        ("west_label", setup.get("west_label", "")),
        ("north_road", setup.get("north_road", "")),
        ("south_road", setup.get("south_road", "")),
        ("east_road", setup.get("east_road", "")),
        ("west_road", setup.get("west_road", "")),
        ("caption_text", setup.get("caption_text", "")),
        ("show_u_turn", setup.get("show_u_turn", "")),
    ]
    return pd.DataFrame(rows, columns=["field", "value"])


class _WorkbookAdapter:
    def __init__(self, workbook) -> None:
        self.book = workbook
        self.sheets = {worksheet.title: worksheet for worksheet in workbook.worksheets}


def _setup_frame(setup: dict[str, Any]) -> pd.DataFrame:
    rows = [{"field": key, "value": value} for key, value in setup.items()]
    rows.append(
        {
            "field": "phf_15min_note",
            "value": "PHF_15min ใช้เฉพาะรายการที่ include_in_peak = TRUE จึงอาจแตกต่างจาก Hourly_Movement_PCU หากไม่รวม U-turn, overpass หรือการเคลื่อนที่อื่น",
        }
    )
    return pd.DataFrame(rows)


def _time_text(value: Any) -> str:
    if pd.isna(value):
        return ""
    if isinstance(value, pd.Timestamp):
        value = value.time()
    if isinstance(value, time):
        return value.strftime("%H:%M")
    text = str(value).strip()
    return text[:5] if len(text) >= 5 and text[2:3] == ":" else text


def _peak_row(peaks: pd.DataFrame, period: str) -> pd.Series | None:
    if peaks.empty or "period" not in peaks.columns:
        return None
    matches = peaks[peaks["period"].astype(str).str.upper() == period]
    if matches.empty:
        return None
    return matches.iloc[0]


def _peak_value(row: pd.Series | None, column: str) -> Any:
    if row is None or column not in row:
        return None
    return row[column]


def _resolved_peaks_for_export(setup: dict[str, Any], normalized: pd.DataFrame, peaks: pd.DataFrame) -> pd.DataFrame:
    periods = confirmed_peak_periods_from_setup(setup)
    if str(setup.get("peak_selection_source") or "").casefold() == PEAK_SELECTION_USER_CONFIRMED and periods:
        return confirmed_peak_phf(
            normalized,
            peak_periods=periods,
            peak_mode=str(setup.get("peak_mode") or DEFAULT_PEAK_MODE),
            peak_selection_source=PEAK_SELECTION_USER_CONFIRMED,
        )
    return peaks


def _window_text(setup: dict[str, Any], start_key: str, end_key: str) -> str:
    start = _time_text(setup.get(start_key))
    end = _time_text(setup.get(end_key))
    return f"{start}-{end}" if start and end else ""


def _peak_report_frame(setup: dict[str, Any], peaks: pd.DataFrame) -> pd.DataFrame:
    am_peak = _peak_row(peaks, "AM")
    pm_peak = _peak_row(peaks, "PM")
    peak_mode = setup.get("peak_mode") or _peak_value(am_peak, "peak_mode") or _peak_value(pm_peak, "peak_mode") or ""
    peak_selection_source = (
        setup.get("peak_selection_source")
        or _peak_value(am_peak, "peak_selection_source")
        or _peak_value(pm_peak, "peak_selection_source")
        or "auto_suggested"
    )

    return pd.DataFrame(
        [
            {
                "peak_mode": peak_mode,
                "am_peak_search_window": _window_text(setup, "am_peak_window_start", "am_peak_window_end"),
                "pm_peak_search_window": _window_text(setup, "pm_peak_window_start", "pm_peak_window_end"),
                "am_peak_start": _time_text(_peak_value(am_peak, "peak_start")),
                "am_peak_end": _time_text(_peak_value(am_peak, "peak_end")),
                "am_peak_pcu": _peak_value(am_peak, "hourly_pcu"),
                "am_phf": _peak_value(am_peak, "phf"),
                "pm_peak_start": _time_text(_peak_value(pm_peak, "peak_start")),
                "pm_peak_end": _time_text(_peak_value(pm_peak, "peak_end")),
                "pm_peak_pcu": _peak_value(pm_peak, "hourly_pcu"),
                "pm_phf": _peak_value(pm_peak, "phf"),
                "peak_selection_source": peak_selection_source,
            }
        ]
    )


def _report_text(normalized: pd.DataFrame, peaks: pd.DataFrame, vehicle: pd.DataFrame) -> pd.DataFrame:
    total_pcu = normalized["pcu"].sum() if "pcu" in normalized else 0
    lines = [
        f"ปริมาณจราจรรวมตลอดช่วงเวลาสำรวจ เท่ากับประมาณ {total_pcu:,.0f} หน่วยรถยนต์นั่งเทียบเท่า (PCU)"
    ]

    period_labels = {"AM": "ช่วงเวลาเร่งด่วนเช้า", "PM": "ช่วงเวลาเร่งด่วนเย็น"}
    for period, label in period_labels.items():
        peak = _peak_row(peaks, period)
        if peak is None:
            lines.append(f"{label} ไม่พบช่วงเวลาที่มีข้อมูลครบถ้วนสำหรับการคำนวณ")
            continue
        lines.append(
            f"{label}อยู่ระหว่างเวลา {_time_text(peak['peak_start'])}-{_time_text(peak['peak_end'])} น. "
            f"มีปริมาณจราจรสูงสุด {peak['hourly_pcu']:,.0f} PCU/ชั่วโมง "
            f"และมีค่า PHF เท่ากับ {peak['phf']:.2f}"
        )

    if not vehicle.empty and {"vehicle_class", "count_share"}.issubset(vehicle.columns):
        top = vehicle.sort_values(["count_share", "count"], ascending=False).head(3)
        parts = [f"{row.vehicle_class} ร้อยละ {row.count_share * 100:.1f}" for row in top.itertuples(index=False)]
        if parts:
            lines.append(f"ประเภทยานพาหนะที่มีสัดส่วนสูงสุด 3 อันดับแรก ได้แก่ {', '.join(parts)} ของจำนวนยานพาหนะทั้งหมด")
    if {"vehicle_class", "pce_factor"}.issubset(normalized.columns):
        overrides = []
        for vehicle_class, factor in normalized.groupby("vehicle_class")["pce_factor"].first().dropna().items():
            default = DEFAULT_PCE_FACTORS.get(str(vehicle_class))
            if default is not None and abs(float(factor) - float(default)) > 1e-12:
                overrides.append(f"{vehicle_class}={float(factor):g}")
        if overrides:
            lines.append(f"PCE overrides used for PCU calculations: {', '.join(overrides)}")
    return pd.DataFrame({"report_text": lines})


def _number_format(column_name: str) -> str | None:
    lowered = column_name.lower()
    if "%" in column_name or lowered == "percent":
        return "0.00%"
    if lowered in {"count", "value"} or lowered.endswith("_count") or column_name in {"จำนวนคัน", "Total (คัน)"}:
        return "#,##0"
    if "share" in lowered or lowered.endswith("phf"):
        return "0.000"
    if "pce" in lowered:
        return "0.000"
    if "pcu" in lowered:
        return "#,##0"
    if column_name in {"15", "30", "45", "60", "Total"}:
        return "#,##0"
    if lowered.endswith("_start") or lowered.endswith("_end") or lowered.startswith("time_"):
        return "hh:mm"
    return None


def _format_worksheet(worksheet, sheet_name: str, create_excel_tables: bool = DEFAULT_CREATE_EXCEL_TABLES) -> None:
    worksheet.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_blue = Side(style="thin", color="B7C9DA")
    border = Border(left=thin_blue, right=thin_blue, top=thin_blue, bottom=thin_blue)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "Export_Metadata": {"A": 24, "B": 72},
        "Setup": {"A": 28, "B": 36},
        "PCE_Factors": {"A": 20, "B": 14, "C": 18},
        "Mapping": {"A": 24, "B": 16, "C": 18, "D": 12, "E": 12, "F": 16, "G": 18, "H": 16, "I": 18},
        "Movement_Aggregation_Audit": {"A": 20, "B": 18, "C": 18, "D": 24, "E": 18, "F": 16, "G": 20, "H": 14},
        "QC_Check": {"A": 14, "B": 24, "C": 30, "D": 56, "E": 64, "F": 20, "G": 24, "H": 18},
        "Hourly_Summary": {"A": 14, "B": 14, "C": 14, "D": 14},
        "Movement_Summary": {"A": 18, "B": 12, "C": 12, "D": 16, "E": 18, "F": 16, "G": 18, "H": 14, "I": 14},
        "Vehicle_Composition": {"A": 20, "B": 14, "C": 14, "D": 16},
        "Hourly_Movement_PCU": {"A": 18},
        "Hourly_Vehicle_Class": {"A": 18, "B": 14, "C": 14},
        "Vehicle_Composition_Report": {"A": 42, "B": 16, "C": 14, "D": 14},
        "Vehicle_Group_PCE": {"A": 24, "B": 14, "C": 14, "D": 14, "E": 34},
        "PHF_15min": {"A": 18, "B": 12, "C": 12, "D": 12, "E": 12, "F": 14, "G": 12},
        "Peak_PHF": {"A": 18, "B": 22, "C": 22, "D": 14, "E": 14, "F": 14, "G": 12, "H": 14, "I": 14, "J": 14, "K": 12},
        "Report_Text": {"A": 110},
    }

    headers = {cell.column: str(cell.value or "") for cell in worksheet[1]}
    for column_cells in worksheet.columns:
        column_letter = column_cells[0].column_letter
        configured_width = widths.get(sheet_name, {}).get(column_letter)
        if configured_width is None:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            configured_width = min(max(max_length + 2, 12), 42)
        worksheet.column_dimensions[column_letter].width = configured_width

        number_format = _number_format(headers.get(column_cells[0].column, ""))
        for cell in column_cells[1:]:
            horizontal = "center" if headers.get(column_cells[0].column, "") in {"เวลา"} else None
            cell.alignment = Alignment(horizontal=horizontal, vertical="top", wrap_text=sheet_name == "Report_Text")
            cell.border = border
            if number_format:
                cell.number_format = number_format

    if create_excel_tables and sheet_name == "Normalized_Data" and worksheet.max_row >= 2 and worksheet.max_column >= 1:
        from openpyxl.worksheet.table import Table, TableStyleInfo

        worksheet.auto_filter.ref = worksheet.dimensions
        table_ref = worksheet.dimensions
        table_name = f"{sheet_name.replace('_', '')}Table"
        table = Table(displayName=table_name, ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        worksheet.add_table(table)


def _write_dataframe_sheet(
    workbook,
    sheet_name: str,
    dataframe: pd.DataFrame,
    create_excel_tables: bool = DEFAULT_CREATE_EXCEL_TABLES,
) -> None:
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    worksheet = workbook.create_sheet(sheet_name)
    for row in dataframe_to_rows(dataframe, index=False, header=True):
        worksheet.append(["" if pd.isna(value) else value for value in row])
    _format_worksheet(worksheet, sheet_name, create_excel_tables=create_excel_tables)


def _find_header_column(worksheet, header: str) -> int | None:
    for cell in worksheet[1]:
        if str(cell.value or "") == header:
            return cell.column
    return None


def _last_detail_row(worksheet, label_column: int = 1) -> int:
    last_row = worksheet.max_row
    if last_row >= 2 and str(worksheet.cell(last_row, label_column).value or "").strip().lower() in {"รวม", "total"}:
        return last_row - 1
    return last_row


def _apply_formula_summaries(writer) -> None:
    vehicle_sheet = writer.sheets["Vehicle_Composition_Report"]
    vehicle_count_column = _find_header_column(vehicle_sheet, "จำนวนคัน")
    vehicle_percent_column = _find_header_column(vehicle_sheet, "สัดส่วน (%)")
    if vehicle_count_column and vehicle_percent_column and vehicle_sheet.max_row >= 2:
        total_row = vehicle_sheet.max_row
        total_count_ref = f"${get_column_letter(vehicle_count_column)}${total_row}"
        for row in range(2, total_row):
            count_ref = f"{get_column_letter(vehicle_count_column)}{row}"
            vehicle_sheet.cell(row, vehicle_percent_column).value = f"=IFERROR({count_ref}/{total_count_ref},0)"
        vehicle_sheet.cell(total_row, vehicle_percent_column).value = f'=IF({total_count_ref}>0,1,0)'

    group_sheet = writer.sheets["Vehicle_Group_PCE"]
    group_value_column = _find_header_column(group_sheet, "value")
    group_percent_column = _find_header_column(group_sheet, "percent")
    if group_value_column and group_percent_column and vehicle_count_column and group_sheet.max_row >= 2:
        denominator = f"'Vehicle_Composition_Report'!${get_column_letter(vehicle_count_column)}${vehicle_sheet.max_row}"
        for row in range(2, group_sheet.max_row + 1):
            value_ref = f"{get_column_letter(group_value_column)}{row}"
            group_sheet.cell(row, group_percent_column).value = f"=IFERROR({value_ref}/{denominator},0)"

    phf_sheet = writer.sheets["PHF_15min"]
    total_column = _find_header_column(phf_sheet, "Total")
    phf_column = _find_header_column(phf_sheet, "PHF")
    quarter_columns = [_find_header_column(phf_sheet, header) for header in ["15", "30", "45", "60"]]
    if total_column and phf_column and all(quarter_columns) and phf_sheet.max_row >= 2:
        quarter_start = get_column_letter(min(quarter_columns))
        quarter_end = get_column_letter(max(quarter_columns))
        total_letter = get_column_letter(total_column)
        for row in range(2, phf_sheet.max_row + 1):
            phf_sheet.cell(row, phf_column).value = f"=IFERROR({total_letter}{row}/(4*MAX({quarter_start}{row}:{quarter_end}{row})),0)"


def _add_excel_native_charts(worksheet, workbook) -> None:
    hourly_sheet = workbook["Hourly_Movement_PCU"]
    hourly_total_column = _find_header_column(hourly_sheet, "Total")
    hourly_last_row = _last_detail_row(hourly_sheet)
    if hourly_total_column and hourly_last_row >= 2:
        hourly_chart = LineChart()
        hourly_chart.title = "Hourly PCU"
        hourly_chart.y_axis.title = "PCU/hour"
        hourly_chart.x_axis.title = "Time"
        hourly_chart.height = 9
        hourly_chart.width = 20
        data = Reference(hourly_sheet, min_col=hourly_total_column, min_row=1, max_row=hourly_last_row)
        categories = Reference(hourly_sheet, min_col=1, min_row=2, max_row=hourly_last_row)
        hourly_chart.add_data(data, titles_from_data=True)
        hourly_chart.set_categories(categories)
        hourly_chart.legend = None
        worksheet.add_chart(hourly_chart, "A6")

    vehicle_sheet = workbook["Vehicle_Composition_Report"]
    vehicle_percent_column = _find_header_column(vehicle_sheet, "สัดส่วน (%)")
    vehicle_last_row = _last_detail_row(vehicle_sheet)
    if vehicle_percent_column and vehicle_last_row >= 2:
        vehicle_chart = BarChart()
        vehicle_chart.type = "bar"
        vehicle_chart.title = "Vehicle Composition (%)"
        vehicle_chart.x_axis.title = "Percent"
        vehicle_chart.y_axis.title = "Vehicle type"
        vehicle_chart.x_axis.numFmt = "0.00%"
        vehicle_chart.height = 12
        vehicle_chart.width = 20
        data = Reference(vehicle_sheet, min_col=vehicle_percent_column, min_row=1, max_row=vehicle_last_row)
        categories = Reference(vehicle_sheet, min_col=1, min_row=2, max_row=vehicle_last_row)
        vehicle_chart.add_data(data, titles_from_data=True)
        vehicle_chart.set_categories(categories)
        vehicle_chart.legend = None
        vehicle_chart.dLbls = DataLabelList()
        vehicle_chart.dLbls.showVal = True
        vehicle_chart.dLbls.numFmt = "0.00%"
        worksheet.add_chart(vehicle_chart, "A25")


def _insert_charts_sheet(writer, chart_pngs: dict[str, bytes]) -> None:
    worksheet = writer.book.create_sheet(CHART_SHEET_NAME)
    worksheet.sheet_view.showGridLines = False
    worksheet["A1"] = "Report Charts"
    worksheet["A1"].font = Font(bold=True, size=16, color="1F4E78")
    worksheet["A2"] = "Excel native charts are linked to summary tables and update when the workbook recalculates."
    worksheet["A3"] = "PNG charts are static export images generated by the app."
    worksheet["A2"].alignment = Alignment(wrap_text=True)
    worksheet["A3"].alignment = Alignment(wrap_text=True)
    worksheet.column_dimensions["A"].width = 18
    worksheet.column_dimensions["B"].width = 18
    worksheet.column_dimensions["C"].width = 18
    worksheet.column_dimensions["D"].width = 18
    worksheet.column_dimensions["E"].width = 18
    worksheet.row_dimensions[1].height = 26

    _add_excel_native_charts(worksheet, writer.book)

    placements = {
        "hourly_pcu": ("A48", "Static PNG - Hourly PCU"),
        "vehicle_composition": ("A80", "Static PNG - Vehicle composition"),
    }
    for key, (anchor, title) in placements.items():
        png = chart_pngs.get(key)
        if not png:
            continue
        label_cell = worksheet.cell(row=worksheet[anchor].row - 1, column=1)
        label_cell.value = title
        label_cell.font = Font(bold=True, color="1F4E78")
        image = OpenpyxlImage(BytesIO(png))
        image.width = 900
        image.height = 460 if key == "hourly_pcu" else 540
        worksheet.add_image(image, anchor)


def _style_range(
    worksheet,
    cell_range: str,
    fill: PatternFill | None = None,
    font: Font | None = None,
    alignment: Alignment | None = None,
    border: Border | None = None,
) -> None:
    for row in worksheet[cell_range]:
        for cell in row:
            if fill is not None:
                cell.fill = fill
            if font is not None:
                cell.font = font
            if alignment is not None:
                cell.alignment = alignment
            if border is not None:
                cell.border = border


def _source_formula(sheet_name: str, cell: str) -> str:
    return f'=IFERROR(\'{sheet_name}\'!{cell},"")'


def _setup_lookup_expr(field: str) -> str:
    return f'IFERROR(INDEX(Setup!$B$1:$B$80,MATCH("{field}",Setup!$A$1:$A$80,0)),"")'


def _setup_lookup_formula(field: str) -> str:
    return f"={_setup_lookup_expr(field)}"


def _setup_first_lookup_formula(*fields: str) -> str:
    formula = '""'
    for field in reversed(fields):
        expression = _setup_lookup_expr(field)
        formula = f"IF(LEN({expression})>0,{expression},{formula})"
    return f"={formula}"


def _setup_join_formula(*fields: str) -> str:
    parts = [_setup_lookup_expr(field) for field in fields]
    separator = '&" "&'
    return f'=TRIM({separator.join(parts)})'


def _setup_survey_point_formula() -> str:
    survey_point = _setup_lookup_expr("survey_point")
    joined_legacy = _setup_join_formula("tmc_id", "tmc_name")[1:]
    return f"=IF(LEN({survey_point})>0,{survey_point},{joined_legacy})"


def _setup_report_title_formula() -> str:
    return _setup_first_lookup_formula("report_title", "tmc_title", "tmc_name")


def _diagram_data_formula(
    value_column: str,
    movement_cell: str,
    hourly_last_row: int,
    hourly_last_column: int,
) -> str:
    last_letter = get_column_letter(hourly_last_column)
    data_range = f"'Hourly_Movement_PCU'!$B$2:${last_letter}${hourly_last_row}"
    time_range = f"'Hourly_Movement_PCU'!$A$2:$A${hourly_last_row}"
    movement_headers = f"'Hourly_Movement_PCU'!$B$1:${last_letter}$1"
    movement_match = f"MATCH({movement_cell},{movement_headers},0)"
    if value_column == "total":
        total_match = f'MATCH("รวม",{time_range},0)'
        total_fallback = f'MATCH("Total",{time_range},0)'
        return (
            f"=IFERROR(INDEX({data_range},{total_match},{movement_match}),"
            f"IFERROR(INDEX({data_range},{total_fallback},{movement_match}),0))"
        )

    peak_cell = "'Peak_PHF'!$H$2" if value_column == "pm" else "'Peak_PHF'!$D$2"
    peak_key = f'IF(ISNUMBER({peak_cell}),TEXT({peak_cell},"hh:mm"),LEFT({peak_cell},5))&"*"'
    return f"=IFERROR(INDEX({data_range},MATCH({peak_key},{time_range},0),{movement_match}),0)"


def _insert_diagram_data_sheet(writer) -> None:
    workbook = writer.book
    worksheet = workbook.create_sheet(DIAGRAM_DATA_SHEET_NAME)
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"
    worksheet.append(["movement_code", "total_pcu", "pm_peak_pcu", "am_peak_pcu"])

    hourly_sheet = workbook["Hourly_Movement_PCU"]
    hourly_last_row = max(hourly_sheet.max_row, 2)
    hourly_last_column = max(hourly_sheet.max_column, 2)
    for row_index, movement_code in enumerate(MOVEMENT_CODES, start=2):
        worksheet.cell(row_index, 1).value = movement_code
        worksheet.cell(row_index, 2).value = _diagram_data_formula("total", f"$A{row_index}", hourly_last_row, hourly_last_column)
        worksheet.cell(row_index, 3).value = _diagram_data_formula("pm", f"$A{row_index}", hourly_last_row, hourly_last_column)
        worksheet.cell(row_index, 4).value = _diagram_data_formula("am", f"$A{row_index}", hourly_last_row, hourly_last_column)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_blue = Side(style="thin", color="B7C9DA")
    border = Border(left=thin_blue, right=thin_blue, top=thin_blue, bottom=thin_blue)
    _style_range(
        worksheet,
        "A1:D1",
        fill=header_fill,
        font=header_font,
        alignment=Alignment(horizontal="center", vertical="center"),
        border=border,
    )
    _style_range(worksheet, "A2:D17", border=border, alignment=Alignment(horizontal="center", vertical="center"))
    for row in range(2, 18):
        for column in range(2, 5):
            worksheet.cell(row, column).number_format = "#,##0"
    for column, width in {"A": 16, "B": 14, "C": 14, "D": 14}.items():
        worksheet.column_dimensions[column].width = width


def _diagram_lookup_formula(movement_code: str, value_column: str) -> str:
    column_map = {"total": "B", "pm": "C", "am": "D"}
    column = column_map[value_column]
    return f'=IFERROR(INDEX(Diagram_Data!${column}:${column},MATCH("{movement_code}",Diagram_Data!$A:$A,0)),0)'


def _write_movement_block(worksheet, start_row: int, start_column: int, title: str, movements: list[str]) -> None:
    blue_fill = PatternFill("solid", fgColor="1F4E78")
    light_fill = PatternFill("solid", fgColor="EAF2F8")
    header_font = Font(color="FFFFFF", bold=True, size=8)
    label_font = Font(bold=True, size=8)
    thin = Side(style="thin", color="8EAADB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    end_column = start_column + len(movements)
    worksheet.merge_cells(
        start_row=start_row,
        start_column=start_column,
        end_row=start_row,
        end_column=end_column,
    )
    title_cell = worksheet.cell(start_row, start_column)
    title_cell.value = title
    title_cell.fill = blue_fill
    title_cell.font = header_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    labels = [("Total 12-hour PCU", "total"), ("PM peak PCU", "pm"), ("AM peak PCU", "am")]
    worksheet.cell(start_row + 1, start_column).value = "movement"
    for index, movement in enumerate(movements, start=start_column + 1):
        worksheet.cell(start_row + 1, index).value = movement
    for offset, (label, key) in enumerate(labels, start=2):
        row = start_row + offset
        worksheet.cell(row, start_column).value = label
        worksheet.cell(row, start_column).font = label_font
        worksheet.cell(row, start_column).fill = light_fill
        for col_offset, movement in enumerate(movements, start=start_column + 1):
            cell = worksheet.cell(row, col_offset)
            cell.value = _diagram_lookup_formula(movement, key)
            cell.number_format = "#,##0"

    for row in range(start_row, start_row + 5):
        for column in range(start_column, end_column + 1):
            cell = worksheet.cell(row, column)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if row == start_row + 1:
                cell.fill = light_fill
                cell.font = label_font


def _copy_linked_table(
    worksheet,
    source_sheet,
    start_row: int,
    start_column: int,
    title: str,
    peak_start_labels: set[str] | None = None,
) -> tuple[int, int]:
    peak_start_labels = peak_start_labels or set()
    source_rows = source_sheet.max_row
    source_columns = source_sheet.max_column
    title_end = start_column + source_columns - 1
    worksheet.merge_cells(start_row=start_row, start_column=start_column, end_row=start_row, end_column=title_end)
    title_cell = worksheet.cell(start_row, start_column)
    title_cell.value = title
    title_cell.font = Font(bold=True, size=11, color="1F4E78")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    table_start = start_row + 1
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True, size=8)
    total_fill = PatternFill("solid", fgColor="D9EAF7")
    peak_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="B7C9DA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_offset in range(source_rows):
        source_row = row_offset + 1
        report_row = table_start + row_offset
        row_label = str(source_sheet.cell(source_row, 1).value or "")
        is_total = row_label.strip().lower() in {"รวม", "total"}
        is_peak = any(row_label.startswith(label) for label in peak_start_labels if label)
        for column_offset in range(source_columns):
            source_column = column_offset + 1
            report_column = start_column + column_offset
            cell = worksheet.cell(report_row, report_column)
            cell.value = _source_formula(source_sheet.title, f"{get_column_letter(source_column)}{source_row}")
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if source_column == 1 or source_row == 1 else "right",
                vertical="center",
                wrap_text=True,
            )
            if source_row == 1:
                cell.fill = header_fill
                cell.font = header_font
            elif is_total:
                cell.fill = total_fill
                cell.font = Font(bold=True, size=8)
            elif is_peak:
                cell.fill = peak_fill
                cell.font = Font(size=8)
            else:
                cell.font = Font(size=8)
            if source_column > 1 and source_row > 1:
                cell.number_format = "#,##0"
    return table_start, table_start + source_rows - 1


def _peak_start_labels(peaks: pd.DataFrame) -> set[str]:
    labels = set()
    if peaks.empty or "peak_start" not in peaks.columns:
        return labels
    for value in peaks["peak_start"].dropna():
        text = _time_text(value)
        if text:
            labels.add(text)
    return labels


def _try_add_report_image(worksheet, png: bytes | None, anchor: str, width: int, height: int) -> None:
    if not png:
        return
    try:
        image = OpenpyxlImage(BytesIO(png))
        image.width = width
        image.height = height
        worksheet.add_image(image, anchor)
    except Exception:
        return


def _insert_tmc_report_sheet(writer, setup: dict[str, Any], peaks: pd.DataFrame, chart_pngs: dict[str, bytes]) -> None:
    workbook = writer.book
    worksheet = workbook.create_sheet(TMC_REPORT_SHEET_NAME)
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "Q5"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_margins.left = 0.25
    worksheet.page_margins.right = 0.25
    worksheet.page_margins.top = 0.35
    worksheet.page_margins.bottom = 0.35
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0

    widths = {
        "A": 4,
        "B": 10,
        "C": 10,
        "D": 10,
        "E": 10,
        "F": 3,
        "G": 8,
        "H": 8,
        "I": 8,
        "J": 3,
        "K": 10,
        "L": 10,
        "M": 10,
        "N": 10,
        "O": 4,
        "P": 2,
    }
    for column in range(1, 35):
        letter = get_column_letter(column)
        worksheet.column_dimensions[letter].width = widths.get(letter, 9)
    for row in range(1, 80):
        worksheet.row_dimensions[row].height = 18

    dark_blue = PatternFill("solid", fgColor="1F4E78")
    medium_blue = PatternFill("solid", fgColor="D9EAF7")
    pale_yellow = PatternFill("solid", fgColor="FFF2CC")
    road_fill = PatternFill("solid", fgColor="D9D9D9")
    thin = Side(style="thin", color="7F7F7F")
    medium = Side(style="medium", color="000000")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    worksheet.merge_cells("A1:AH1")
    worksheet["A1"] = _setup_report_title_formula()
    worksheet["A1"].fill = dark_blue
    worksheet["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 28

    worksheet.merge_cells("A2:AH2")
    worksheet["A2"] = _setup_survey_point_formula()
    worksheet["A2"].font = Font(bold=True, size=13, color="1F4E78")
    worksheet["A2"].alignment = Alignment(horizontal="center", vertical="center")

    metadata = [
        ("Project", _setup_first_lookup_formula("project_name", "project")),
        ("Survey point", _setup_survey_point_formula()),
        ("Survey date", _setup_lookup_formula("survey_date")),
        ("Weather", _setup_lookup_formula("weather")),
        ("Responsible party", _setup_lookup_formula("responsible_party")),
        ("Survey period", _setup_first_lookup_formula("survey_period", "survey_period_text")),
    ]
    for row_offset, (label, formula) in enumerate(metadata, start=4):
        worksheet.cell(row_offset, 1).value = label
        worksheet.cell(row_offset, 1).font = Font(bold=True, color="1F4E78")
        worksheet.cell(row_offset, 1).fill = medium_blue
        worksheet.cell(row_offset, 1).border = thin_border
        worksheet.merge_cells(start_row=row_offset, start_column=2, end_row=row_offset, end_column=8)
        value_cell = worksheet.cell(row_offset, 2)
        value_cell.value = formula
        value_cell.border = thin_border
        value_cell.alignment = Alignment(horizontal="left", vertical="center")

    worksheet.merge_cells("A11:O11")
    worksheet["A11"] = "Intersection movement schematic"
    worksheet["A11"].fill = dark_blue
    worksheet["A11"].font = Font(color="FFFFFF", bold=True, size=11)
    worksheet["A11"].alignment = Alignment(horizontal="center")
    _style_range(worksheet, "A12:O45", border=thin_border, alignment=Alignment(horizontal="center", vertical="center", wrap_text=True))

    # Lightweight cell-based road schematic.
    _style_range(worksheet, "G22:I28", fill=road_fill, border=Border(left=medium, right=medium, top=medium, bottom=medium))
    _style_range(worksheet, "A25:O25", fill=road_fill)
    _style_range(worksheet, "H12:H45", fill=road_fill)
    worksheet["H20"] = "↑"
    worksheet["H30"] = "↓"
    worksheet["F25"] = "←"
    worksheet["J25"] = "→"
    for cell_ref in ["H20", "H30", "F25", "J25"]:
        worksheet[cell_ref].font = Font(bold=True, size=14, color="C00000")

    labels = {
        "H12": _setup_lookup_formula("north_label"),
        "H44": _setup_lookup_formula("south_label"),
        "A25": _setup_lookup_formula("west_label"),
        "O25": _setup_lookup_formula("east_label"),
        "I17": _setup_lookup_formula("north_road"),
        "G36": _setup_lookup_formula("south_road"),
        "C24": _setup_lookup_formula("west_road"),
        "M26": _setup_lookup_formula("east_road"),
    }
    for cell_ref, formula in labels.items():
        worksheet[cell_ref] = formula
        worksheet[cell_ref].font = Font(bold=True, size=8)
        worksheet[cell_ref].fill = pale_yellow
        worksheet[cell_ref].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    _write_movement_block(worksheet, 13, 2, "North approach", ["NU", "NW", "NS", "NE"])
    _write_movement_block(worksheet, 37, 2, "South approach", ["SW", "SN", "SE", "SU"])
    _write_movement_block(worksheet, 28, 1, "West approach", ["WN", "WE", "WS"])
    _write_movement_block(worksheet, 28, 11, "East approach", ["EN", "EW", "ES"])

    worksheet.merge_cells("B43:E43")
    worksheet["B43"] = "Summary"
    worksheet["B43"].fill = dark_blue
    worksheet["B43"].font = Font(color="FFFFFF", bold=True, size=9)
    worksheet["B43"].alignment = Alignment(horizontal="center")
    summary_rows = [
        ("AM peak PCU", "=IFERROR(Peak_PHF!$F$2,0)"),
        ("PM peak PCU", "=IFERROR(Peak_PHF!$J$2,0)"),
        ("Total PCU", "=SUM(Diagram_Data!$B$2:$B$17)"),
    ]
    for offset, (label, formula) in enumerate(summary_rows, start=44):
        worksheet.cell(offset, 2).value = label
        worksheet.cell(offset, 2).fill = medium_blue
        worksheet.cell(offset, 2).font = Font(bold=True, size=8)
        worksheet.merge_cells(start_row=offset, start_column=3, end_row=offset, end_column=5)
        worksheet.cell(offset, 3).value = formula
        worksheet.cell(offset, 3).number_format = "#,##0"
        worksheet.cell(offset, 3).alignment = Alignment(horizontal="right")
        _style_range(worksheet, f"B{offset}:E{offset}", border=thin_border)

    worksheet.merge_cells("A47:O47")
    survey_period = str(setup.get("survey_period") or setup.get("survey_period_text") or "07.00-19.00 น.")
    worksheet["A47"] = f"ปริมาณจราจรบนทางแยก PCU/12 ชม. ({survey_period})"
    if setup.get("caption_text"):
        worksheet["A47"] = setup["caption_text"]
    worksheet["A47"].font = Font(bold=True, size=11)
    worksheet["A47"].alignment = Alignment(horizontal="center")

    peak_start_labels = _peak_start_labels(peaks)
    hourly_start, hourly_end = _copy_linked_table(
        worksheet,
        workbook["Hourly_Movement_PCU"],
        4,
        17,
        "Hourly movement PCU",
        peak_start_labels,
    )
    vehicle_start_row = hourly_end + 3
    _copy_linked_table(
        worksheet,
        workbook["Hourly_Vehicle_Class"],
        vehicle_start_row,
        17,
        "Hourly vehicle-class table",
        peak_start_labels,
    )

    worksheet["A50"] = "Hourly PCU chart"
    worksheet["A50"].font = Font(bold=True, size=11, color="1F4E78")
    _try_add_report_image(worksheet, chart_pngs.get("hourly_pcu"), "A51", 600, 310)
    worksheet["Q50"] = "Vehicle composition chart"
    worksheet["Q50"].font = Font(bold=True, size=11, color="1F4E78")
    _try_add_report_image(worksheet, chart_pngs.get("vehicle_composition"), "Q51", 600, 340)

    worksheet.print_area = "A1:AH72"


def _diagram_config_from_setup(setup: dict[str, Any]) -> DiagramConfig:
    setup = setup_with_metadata(setup)
    return DiagramConfig(
        tmc_id=str(setup.get("tmc_id", "") or ""),
        tmc_name=str(setup.get("tmc_title") or setup.get("tmc_name", "") or ""),
        survey_date_text=str(setup.get("survey_date_text", "") or ""),
        north_label=str(setup.get("north_label", "") or ""),
        south_label=str(setup.get("south_label", "") or ""),
        east_label=str(setup.get("east_label", "") or ""),
        west_label=str(setup.get("west_label", "") or ""),
        north_road=str(setup.get("north_road", "") or ""),
        south_road=str(setup.get("south_road", "") or ""),
        east_road=str(setup.get("east_road", "") or ""),
        west_road=str(setup.get("west_road", "") or ""),
        survey_period_text=str(setup.get("survey_period") or setup.get("survey_period_text") or "07.00-19.00 น."),
        caption_text=str(setup.get("caption_text", "") or ""),
        show_u_turn=bool(setup.get("show_u_turn", True)),
    )


def _insert_diagram_sheet(writer, diagram_png: bytes) -> None:
    worksheet = writer.book.create_sheet(DIAGRAM_SHEET_NAME)
    worksheet.sheet_view.showGridLines = False
    worksheet["A1"] = "TMC Movement Diagram"
    worksheet["A1"].font = Font(bold=True, size=16, color="1F4E78")
    worksheet["A2"] = "Static PNG movement diagram generated from Hourly_Movement_PCU and Peak_PHF."
    worksheet["A2"].alignment = Alignment(wrap_text=True)
    for column in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]:
        worksheet.column_dimensions[column].width = 15
    worksheet.row_dimensions[1].height = 26
    image = OpenpyxlImage(BytesIO(diagram_png))
    image.width = 1180
    image.height = 910
    worksheet.add_image(image, "A4")


def _export_workbook_from_template(
    setup: dict[str, Any],
    sheets: dict[str, pd.DataFrame],
    peaks: pd.DataFrame,
    hourly_movement: pd.DataFrame,
    chart_pngs: dict[str, bytes],
    include_charts: bool,
    include_diagram: bool,
    template_path: str | None,
    template_map_path: str | None,
    create_excel_tables: bool,
) -> bytes:
    resources = load_report_template_resources(
        template_path or DEFAULT_TEMPLATE_PATH,
        template_map_path or DEFAULT_TEMPLATE_MAP_PATH,
    )
    workbook = load_workbook(resources.template_path)
    adapter = _WorkbookAdapter(workbook)
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    for sheet_name in EXPORT_SHEETS:
        _write_dataframe_sheet(
            workbook,
            sheet_name,
            sheets[sheet_name],
            create_excel_tables=create_excel_tables,
        )
    adapter.sheets = {worksheet.title: worksheet for worksheet in workbook.worksheets}
    _apply_formula_summaries(adapter)
    _insert_diagram_data_sheet(adapter)
    adapter.sheets = {worksheet.title: worksheet for worksheet in workbook.worksheets}

    populate_template_report_sheet(
        workbook=workbook,
        mapping=resources.mapping,
        setup=setup,
        hourly_movement_pcu=hourly_movement,
        hourly_vehicle_class=sheets["Hourly_Vehicle_Class"],
        vehicle_composition_report=sheets["Vehicle_Composition_Report"],
        chart_pngs=chart_pngs if include_charts else {},
        report_sheet_name=TMC_REPORT_SHEET_NAME,
        # Native chart preservation via OOXML package patching is disabled because it can corrupt worksheet XML.
        use_native_template_charts=False,
    )

    if include_charts:
        _insert_charts_sheet(adapter, chart_pngs)
    if include_diagram:
        diagram_png = generate_four_leg_tmc_diagram(hourly_movement, peaks, _diagram_config_from_setup(setup))
        _insert_diagram_sheet(adapter, diagram_png)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _non_total_rows(dataframe: pd.DataFrame) -> pd.DataFrame:
    return hourly_interval_rows(dataframe)


def _native_chart_source_data(
    hourly_movement: pd.DataFrame,
    vehicle_composition_for_report: pd.DataFrame,
) -> dict[str, dict[str, list[Any]]]:
    hourly_rows = _non_total_rows(hourly_movement)
    time_column = hourly_rows.columns[0] if len(hourly_rows.columns) else None
    hourly_categories = hourly_rows[time_column].tolist() if time_column is not None else []
    hourly_values = hourly_rows["Total"].tolist() if "Total" in hourly_rows.columns else []

    vehicle_lookup: dict[str, Any] = {}
    if {"vehicle_class", "à¸ªà¸±à¸”à¸ªà¹ˆà¸§à¸™ (%)"}.issubset(vehicle_composition_for_report.columns):
        for _, row in vehicle_composition_for_report.iterrows():
            vehicle_class = str(row["vehicle_class"])
            if vehicle_class != "Total":
                vehicle_lookup[vehicle_class] = row["à¸ªà¸±à¸”à¸ªà¹ˆà¸§à¸™ (%)"]

    return {
        "hourly_pcu": {
            "categories": hourly_categories,
            "values": hourly_values,
        },
        "vehicle_composition": {
            "categories": list(VEHICLE_CLASSES),
            "values": [vehicle_lookup.get(vehicle_class, 0) for vehicle_class in VEHICLE_CLASSES],
        },
    }


def _export_workbook_with_excel_com(
    setup: dict[str, Any],
    sheets: dict[str, pd.DataFrame],
    hourly_movement: pd.DataFrame,
    template_path: str | None,
    template_map_path: str | None,
) -> bytes:
    from .excel_com_export import export_with_excel_com

    resources = load_report_template_resources(
        template_path or DEFAULT_TEMPLATE_PATH,
        template_map_path or DEFAULT_TEMPLATE_MAP_PATH,
    )
    chart_source_data = _native_chart_source_data(hourly_movement, sheets["Vehicle_Composition_Report"])
    report_data = {
        "sheets": sheets,
        "hourly_movement_pcu": hourly_movement,
        "hourly_vehicle_class": sheets["Hourly_Vehicle_Class"],
        "vehicle_composition_report": sheets["Vehicle_Composition_Report"],
    }
    with TemporaryDirectory() as temporary_directory:
        output_path = Path(temporary_directory) / "tmc_report_excel_com.xlsx"
        export_with_excel_com(
            template_path=resources.template_path,
            output_path=output_path,
            template_map=resources.mapping,
            report_data=report_data,
            metadata=setup,
            chart_source_data=chart_source_data,
        )
        return output_path.read_bytes()


def export_workbook(
    setup: dict[str, Any],
    mapping: pd.DataFrame,
    normalized: pd.DataFrame,
    qc: pd.DataFrame,
    hourly: pd.DataFrame,
    movement: pd.DataFrame,
    vehicle: pd.DataFrame,
    peaks: pd.DataFrame,
    include_charts: bool = True,
    include_diagram: bool = True,
    use_template_report_layout: bool = False,
    use_excel_com_native_charts: bool = False,
    template_path: str | None = None,
    template_map_path: str | None = None,
    create_excel_tables: bool = DEFAULT_CREATE_EXCEL_TABLES,
    pce_factors: dict[str, float] | None = None,
    export_mode: str | None = None,
    source_file_name: str | None = None,
    generated_at: datetime | str | None = None,
    template_version: str = TEMPLATE_VERSION,
) -> bytes:
    setup = setup_with_metadata(setup)
    if normalize_movement_code_scheme(setup.get("movement_code_scheme")) == MOVEMENT_SCHEME_V2:
        raise ValueError(
            "approach_movement v2 export/report generation is not supported yet; use the v2 dry-run path for summaries only."
        )
    mapping = clean_mapping(mapping)
    peaks = _resolved_peaks_for_export(setup, normalized, peaks)
    buffer = BytesIO()
    hourly_movement = hourly_movement_pcu(normalized, mapping)
    vehicle_composition_for_report = vehicle_composition_report(normalized)
    sheets = {
        "Export_Metadata": _export_metadata_frame(
            setup,
            export_mode=export_mode,
            source_file_name=source_file_name,
            generated_at=generated_at,
            template_version=template_version,
        ),
        "Setup": _setup_frame(setup),
        "PCE_Factors": pce_factor_traceability_frame(pce_factors),
        "Mapping": mapping,
        "Movement_Aggregation_Audit": movement_aggregation_audit(normalized, mapping),
        "Normalized_Data": normalized,
        "QC_Check": qc,
        "Hourly_Summary": hourly,
        "Movement_Summary": movement,
        "Vehicle_Composition": vehicle,
        "Hourly_Movement_PCU": hourly_movement,
        "Hourly_Vehicle_Class": hourly_vehicle_class(normalized, hourly),
        "Vehicle_Composition_Report": vehicle_composition_for_report,
        "Vehicle_Group_PCE": vehicle_group_pce(normalized),
        "PHF_15min": phf_15min(normalized),
        "Peak_PHF": _peak_report_frame(setup, peaks),
        "Report_Text": _report_text(normalized, peaks, vehicle),
    }
    chart_pngs = dict(report_chart_pngs(hourly_movement, vehicle_composition_for_report, setup=setup)) if include_charts else {}
    if use_excel_com_native_charts:
        try:
            from .excel_com_export import ExcelComUnavailable, require_excel_com

            require_excel_com()
            return _export_workbook_with_excel_com(
                setup=setup,
                sheets=sheets,
                hourly_movement=hourly_movement,
                template_path=template_path,
                template_map_path=template_map_path,
            )
        except ExcelComUnavailable as exc:
            warnings.warn(
                f"Excel COM unavailable; falling back to safe openpyxl export with PNG charts: {exc}",
                RuntimeWarning,
                stacklevel=2,
            )
        except Exception as exc:
            warnings.warn(
                f"Excel COM native-chart export failed after COM was available; falling back to safe openpyxl export with PNG charts: {exc}",
                RuntimeWarning,
                stacklevel=2,
            )

    if use_template_report_layout:
        try:
            return _export_workbook_from_template(
                setup=setup,
                sheets=sheets,
                peaks=peaks,
                hourly_movement=hourly_movement,
                chart_pngs=chart_pngs,
                include_charts=include_charts,
                include_diagram=include_diagram,
                template_path=template_path,
                template_map_path=template_map_path,
                create_excel_tables=create_excel_tables,
            )
        except (OSError, ValueError, KeyError, ReportTemplateUnavailable) as exc:
            warnings.warn(
                f"Template report layout unavailable; falling back to generated report sheet: {exc}",
                RuntimeWarning,
                stacklevel=2,
            )

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        writer.book.calculation.calcMode = "auto"
        writer.book.calculation.fullCalcOnLoad = True
        writer.book.calculation.forceFullCalc = True
        for sheet_name in EXPORT_SHEETS:
            sheets[sheet_name].to_excel(writer, sheet_name=sheet_name, index=False)
            _format_worksheet(
                writer.sheets[sheet_name],
                sheet_name,
                create_excel_tables=create_excel_tables,
            )
        _apply_formula_summaries(writer)
        _insert_diagram_data_sheet(writer)
        _insert_tmc_report_sheet(writer, setup, peaks, chart_pngs)
        if include_charts:
            _insert_charts_sheet(writer, chart_pngs)
        if include_diagram:
            diagram_png = generate_four_leg_tmc_diagram(hourly_movement, peaks, _diagram_config_from_setup(setup))
            _insert_diagram_sheet(writer, diagram_png)
    return buffer.getvalue()
